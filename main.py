# -*- coding: utf-8 -*-
"""
Gestor de Proyectos API - Versión Limpia
API principal para gestión de proyectos con Firebase
Arquitectura modular optimizada para NextJS
Soporte completo para UTF-8 y caracteres especiales en español
"""

import os
import sys
import logging
from contextlib import asynccontextmanager
from dotenv import load_dotenv

# Cargar variables de entorno desde .env
load_dotenv()

# Configurar logger
logger = logging.getLogger(__name__)

# Configurar encoding UTF-8 para todo el sistema
if sys.platform.startswith('win'):
    # En Windows, asegurar UTF-8
    import locale
    try:
        locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_ALL, 'Spanish_Spain.1252')
        except:
            pass

# Configurar stdout y stderr para UTF-8
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')
from fastapi import FastAPI, HTTPException, Query, Request, status, Form, UploadFile, File, Path
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from typing import Dict, Any, Optional, Union, List
import uvicorn
import asyncio
from datetime import datetime
import json
import re
import uuid

# Rate limiting (opcional, con fallback)
try:
    from slowapi import Limiter, _rate_limit_exceeded_handler
    from slowapi.util import get_remote_address
    from slowapi.errors import RateLimitExceeded
    SLOWAPI_AVAILABLE = True
    print("✅ SlowAPI loaded successfully")
except ImportError as e:
    print(f"⚠️ Warning: SlowAPI not available: {e} - Rate limiting disabled")
    SLOWAPI_AVAILABLE = False
    Limiter = None
    _rate_limit_exceeded_handler = None
    get_remote_address = None
    RateLimitExceeded = None

# Monitoring with Prometheus (DESHABILITADO temporalmente por conflictos en Railway)
# TODO: Habilitar cuando se configure correctamente prometheus_multiproc_dir
PROMETHEUS_AVAILABLE = False
Counter = None
Histogram = None
Gauge = None
generate_latest = None
CONTENT_TYPE_LATEST = None
print("⚠️ Prometheus metrics disabled (temporarily disabled for Railway compatibility)")

# Importar para manejar tipos de Firebase
try:
    from google.cloud.firestore_v1._helpers import DatetimeWithNanoseconds
    FIREBASE_TYPES_AVAILABLE = True
except ImportError:
    FIREBASE_TYPES_AVAILABLE = False
    DatetimeWithNanoseconds = None

# Importar sistema de autenticación y autorización
try:
    from auth_system import (
        ROLES,
        DEFAULT_USER_ROLE,
        ROLE_HIERARCHY,
        PUBLIC_PATHS as AUTH_PUBLIC_PATHS
    )
    from auth_system.middleware import AuthorizationMiddleware, AuditLogMiddleware
    AUTH_SYSTEM_AVAILABLE = True
    print("✅ Auth system loaded successfully")
except ImportError as e:
    print(f"⚠️ Warning: Auth system not available: {e}")
    AUTH_SYSTEM_AVAILABLE = False
    ROLES = {}
    DEFAULT_USER_ROLE = "visualizador"
    ROLE_HIERARCHY = {}
    AUTH_PUBLIC_PATHS = []
    AuthorizationMiddleware = None
    AuditLogMiddleware = None

# Importar Firebase con configuración automática
try:
    from database.firebase_config import (
        PROJECT_ID, 
        FIREBASE_AVAILABLE, 
        ensure_firebase_configured, 
        configure_firebase,
        validate_firebase_connection,
        get_firestore_client
    )
    print(f"✅ Firebase auto-config loaded successfully - FIREBASE_AVAILABLE: {FIREBASE_AVAILABLE}")
except Exception as e:
    print(f"❌ Warning: Firebase import failed: {e}")
    FIREBASE_AVAILABLE = False
    PROJECT_ID = os.getenv("FIREBASE_PROJECT_ID", "NOT_CONFIGURED")
    configure_firebase = lambda: (False, {"error": "Not available"})
    ensure_firebase_configured = lambda: False
    validate_firebase_connection = lambda: {"connected": False, "error": "Not available"}
    get_firestore_client = lambda: None

# Importar scripts de forma segura
try:
    from api.scripts import (
        # Firebase operations
        get_collections_info,
        test_firebase_connection,
        get_collections_summary,
        get_proyectos_presupuestales,
        get_unique_nombres_centros_gestores,
        get_proyectos_presupuestales_by_bpin,
        get_proyectos_presupuestales_by_bp,
        get_proyectos_presupuestales_by_centro_gestor,
        # Unidades proyecto operations (funciones especializadas y optimizadas)
        get_unidades_proyecto_geometry,
        get_unidades_proyecto_attributes,
        get_filter_options,
        validate_unidades_proyecto_collection,
        # Contratos operations
        get_contratos_init_data,
        get_contratos_emprestito_all,
        get_contratos_emprestito_by_referencia,
        get_contratos_emprestito_by_centro_gestor,
        # Bancos operations
        get_bancos_emprestito_all,
        get_procesos_emprestito_all,
        # Empréstito operations completas
        obtener_datos_secop_completos,
        actualizar_proceso_emprestito_completo,
        procesar_todos_procesos_emprestito_completo,
        # Nuevas funciones para proyecciones de empréstito
        crear_tabla_proyecciones_desde_sheets,
        leer_proyecciones_emprestito,
        leer_proyecciones_no_guardadas,
        get_proyecciones_sin_proceso,
        actualizar_proyeccion_emprestito,
        # Reportes contratos operations
        create_reporte_contrato,
        get_reportes_contratos,
        get_reporte_contrato_by_id,
        get_reportes_by_centro_gestor,
        get_reportes_by_referencia_contrato,
        setup_google_drive_service,
        # User management operations
        validate_email,
        validate_fullname,
        validate_password,
        validate_cellphone,
        check_user_session,
        create_user_account,
        update_user_password,
        delete_user_account,
        list_users,
        # Auth operations
        authenticate_email_password,
        validate_user_session,
        # Proyectos presupuestales operations
        process_proyectos_presupuestales_json,
        # Availability flags
        USER_MANAGEMENT_AVAILABLE,
        AUTH_OPERATIONS_AVAILABLE,
        EMPRESTITO_OPERATIONS_AVAILABLE,
        REPORTES_CONTRATOS_AVAILABLE,
        PROYECTOS_PRESUPUESTALES_OPERATIONS_AVAILABLE,
        # Flujo caja operations
        process_flujo_caja_excel,
        save_flujo_caja_to_firebase,
        get_flujo_caja_from_firebase,
        FLUJO_CAJA_OPERATIONS_AVAILABLE,
    )
    SCRIPTS_AVAILABLE = True
    print(f"✅ Scripts imported successfully - SCRIPTS_AVAILABLE: {SCRIPTS_AVAILABLE}")
except Exception as e:
    print(f"❌ Warning: Scripts import failed: {e}")
    SCRIPTS_AVAILABLE = False
    USER_MANAGEMENT_AVAILABLE = False
    AUTH_OPERATIONS_AVAILABLE = False
    FLUJO_CAJA_OPERATIONS_AVAILABLE = False

# Importar modelos Pydantic
try:
    from api.models import (
        UserRegistrationRequest,
        UserLoginRequest,
        PasswordUpdateRequest,
        GoogleAuthRequest,
        SessionValidationRequest,
        UserListFilters,
        StandardResponse,
        ValidationErrorResponse,
        EmprestitoRequest,
        EmprestitoResponse,
        ProyeccionEmprestitoUpdateRequest,
        ProyeccionEmprestitoUpdateResponse,
        ProyeccionEmprestitoRegistroRequest,
        ProyeccionEmprestitoRegistroResponse,
        USER_MODELS_AVAILABLE,
        # Reportes contratos models
        ReporteContratoRequest,
        ReporteContratoResponse,
        REPORTE_MODELS_AVAILABLE,
        # Proyectos presupuestales models
        PROYECTOS_PRESUPUESTALES_MODELS_AVAILABLE,
        # Flujo de caja models
        FlujoCajaRequest,
        FlujoCajaResponse,
        FlujoCajaUploadRequest,
        FlujoCajaFilters,
        FLUJO_CAJA_MODELS_AVAILABLE,
    )
    print(f"✅ User models imported successfully - USER_MODELS_AVAILABLE: {USER_MODELS_AVAILABLE}")
except Exception as e:
    print(f"❌ Warning: User models import failed: {e}")
    USER_MODELS_AVAILABLE = False
    
    # Crear clases dummy para evitar errores de NameError
    from pydantic import BaseModel
    from typing import Optional
    
    class UserRegistrationRequest(BaseModel):
        email: str
        password: str
        confirmPassword: str
        name: str
        cellphone: str
        nombre_centro_gestor: str
    
    class UserLoginRequest(BaseModel):
        email: str
        password: str
    
    class PasswordUpdateRequest(BaseModel):
        uid: str
        new_password: str
    
    class GoogleAuthRequest(BaseModel):
        id_token: str
    
    class SessionValidationRequest(BaseModel):
        id_token: str
    
    class UserListFilters(BaseModel):
        pass
    
    class StandardResponse(BaseModel):
        success: bool
        message: Optional[str] = None
    
    class ValidationErrorResponse(BaseModel):
        success: bool = False
        error: str
    
    class EmprestitoRequest(BaseModel):
        referencia_proceso: str
        nombre_centro_gestor: str
        nombre_banco: str
        bp: str
        plataforma: str
        nombre_resumido_proceso: Optional[str] = None
        id_paa: Optional[str] = None
        valor_proyectado: Optional[float] = None
    
    class EmprestitoResponse(BaseModel):
        success: bool
        message: Optional[str] = None



# Configurar el lifespan de la aplicación
@asynccontextmanager
async def lifespan(app: FastAPI):
    """Gestionar el ciclo de vida de la aplicación"""
    # Startup
    print("Starting API...")
    print(f"Port: {os.getenv('PORT', '8000')}")
    print(f"Environment: {os.getenv('ENVIRONMENT', 'development')}")
    print(f"Firebase Project: {PROJECT_ID}")
    
    # Inicializar Firebase de forma segura
    if ensure_firebase_configured():
        print("✅ Firebase initialized successfully")
    else:
        print("❌ Firebase initialization failed")
    
    # Inicializar Firebase automáticamente (sin fallar la app)
    firebase_initialized = False
    if FIREBASE_AVAILABLE:
        try:
            firebase_initialized, status = configure_firebase()
            if firebase_initialized:
                print("✅ Firebase initialized successfully")
            else:
                print(f"⚠️ Firebase initialization failed: {status.get('error', 'Unknown error')}")
        except Exception as e:
            print(f"⚠️ Firebase setup error: {e} - API will run in limited mode")
            firebase_initialized = False
    else:
        print("⚠️ Firebase not available - API running in limited mode")
        firebase_initialized = False
    
    print(f"🚀 API starting with Firebase: {'✅ Connected' if firebase_initialized else '❌ Limited mode'}")
    
    yield
    
    # Shutdown
    print("Stopping API...")

# ============================================
# 📊 MÉTRICAS DE PROMETHEUS PARA MONITOREO APM
# ============================================
# Inicializar métricas como None por defecto
REQUEST_COUNT = None
REQUEST_LATENCY = None
ACTIVE_REQUESTS = None
FIREBASE_QUERIES = None
CACHE_HITS = None
CACHE_MISSES = None

if PROMETHEUS_AVAILABLE and Counter is not None:
    try:
        # Configurar Prometheus para modo multi-proceso si está disponible
        # Esto previene errores cuando Railway usa múltiples workers
        import os
        if 'prometheus_multiproc_dir' not in os.environ:
            # Si no está configurado multi-proceso, usar registro normal
            pass
        
        REQUEST_COUNT = Counter(
            'gestor_api_requests_total', 
            'Total de requests por endpoint',
            ['method', 'endpoint', 'status']
        )

        REQUEST_LATENCY = Histogram(
            'gestor_api_request_duration_seconds',
            'Latencia de requests en segundos',
            ['method', 'endpoint']
        )

        ACTIVE_REQUESTS = Gauge(
            'gestor_api_requests_active',
            'Número de requests activos',
            ['method', 'endpoint']
        )

        FIREBASE_QUERIES = Counter(
            'gestor_api_firebase_queries_total',
            'Total de queries a Firebase/Firestore',
            ['collection']
        )

        CACHE_HITS = Counter(
            'gestor_api_cache_hits_total',
            'Total de cache hits',
            ['endpoint']
        )

        CACHE_MISSES = Counter(
            'gestor_api_cache_misses_total',
            'Total de cache misses',
            ['endpoint']
        )
        print("✅ Prometheus metrics initialized")
    except ValueError as e:
        # ValueError típicamente ocurre cuando la métrica ya está registrada (múltiples workers)
        print(f"⚠️ Warning: Prometheus metrics already registered (multi-worker): {e}")
        print("   Metrics will be disabled for this worker to prevent conflicts")
        REQUEST_COUNT = None
        REQUEST_LATENCY = None
        ACTIVE_REQUESTS = None
        FIREBASE_QUERIES = None
        CACHE_HITS = None
        CACHE_MISSES = None
    except Exception as e:
        print(f"⚠️ Warning: Failed to initialize Prometheus metrics: {e}")
        print("   Continuing without metrics...")
        REQUEST_COUNT = None
        REQUEST_LATENCY = None
        ACTIVE_REQUESTS = None
        FIREBASE_QUERIES = None
        CACHE_HITS = None
        CACHE_MISSES = None
else:
    print("⚠️ Prometheus metrics disabled")

# ============================================
# 🚦 RATE LIMITER PARA PREVENIR ABUSO
# ============================================
if SLOWAPI_AVAILABLE:
    limiter = Limiter(key_func=get_remote_address)
    print("✅ Rate limiter initialized")
else:
    limiter = None
    print("⚠️ Rate limiting disabled")

# Crear instancia de FastAPI con lifespan y soporte UTF-8
app = FastAPI(
    title="Gestor de Proyectos API",
    description="API para gestión de proyectos con Firebase/Firestore - Soporte completo UTF-8 🇪🇸",
    version="1.0.0",
    lifespan=lifespan,
    swagger_ui_parameters={
        "defaultModelsExpandDepth": 1,
        "displayRequestDuration": True,
        "filter": True,
        "tryItOutEnabled": True,
        "requestSnippetsEnabled": True,
        "defaultModelRendering": "example",
        "showExtensions": True,
        "showCommonExtensions": True
    }
)

# Registrar el rate limiter con FastAPI (solo si está disponible)
if SLOWAPI_AVAILABLE and limiter is not None and RateLimitExceeded is not None and _rate_limit_exceeded_handler is not None:
    app.state.limiter = limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
    print("✅ Rate limiting registered with FastAPI")
else:
    print("⚠️ Rate limiting disabled - SlowAPI not available")

# Función decorador opcional para rate limiting
def optional_rate_limit(limit_string: str):
    """Decorador que aplica rate limiting solo si SlowAPI está disponible"""
    def decorator(func):
        if SLOWAPI_AVAILABLE and limiter is not None:
            try:
                return limiter.limit(limit_string)(func)
            except Exception as e:
                print(f"⚠️ Warning: Could not apply rate limit to {func.__name__}: {e}")
                return func
        return func
    return decorator

# 🚀 CACHE SIMPLE EN MEMORIA PARA OPTIMIZACIÓN
from functools import lru_cache
from datetime import timedelta
import hashlib

# Cache simple en memoria (usar Redis en producción)
_simple_cache = {}
_cache_timestamps = {}

def get_cache_key(func_name: str, *args, **kwargs) -> str:
    """Generar clave de caché única"""
    key_data = f"{func_name}:{str(args)}:{str(sorted(kwargs.items()))}"
    return hashlib.md5(key_data.encode()).hexdigest()

def get_from_cache(cache_key: str, max_age_seconds: int = 300):
    """Obtener del caché si existe y es válido"""
    if cache_key in _simple_cache:
        cached_time = _cache_timestamps.get(cache_key)
        if cached_time and (datetime.now() - cached_time).total_seconds() < max_age_seconds:
            return _simple_cache[cache_key], True
    return None, False

def set_in_cache(cache_key: str, value):
    """Guardar en caché"""
    _simple_cache[cache_key] = value
    _cache_timestamps[cache_key] = datetime.now()

def async_cache(ttl_seconds: int = 300):
    """
    Decorador para cachear funciones async con TTL (Time To Live)
    Uso: @async_cache(ttl_seconds=600)
    
    IMPORTANTE: Cachea el resultado ANTES de cualquier middleware (gzip, etc)
    """
    def decorator(func):
        from functools import wraps
        import copy
        
        @wraps(func)
        async def wrapper(*args, **kwargs):
            # Generar clave de caché única basada en función y argumentos
            cache_key = get_cache_key(func.__name__, *args, **kwargs)
            
            # Intentar obtener del caché
            cached_value, is_valid = get_from_cache(cache_key, ttl_seconds)
            if is_valid:
                logger.info(f"✅ Cache hit for {func.__name__}")
                # Retornar copia profunda para evitar mutaciones
                try:
                    return copy.deepcopy(cached_value)
                except:
                    return cached_value
            
            # Si no está en caché, ejecutar función
            logger.info(f"⚠️ Cache miss for {func.__name__} - ejecutando función")
            result = await func(*args, **kwargs)
            
            # Guardar en caché solo si es serializable
            try:
                set_in_cache(cache_key, result)
            except Exception as e:
                logger.warning(f"No se pudo cachear resultado de {func.__name__}: {e}")
            
            return result
        
        return wrapper
    return decorator

# Configurar CORS - Optimizado para Vercel + Railway + Netlify + Live Server
def get_cors_origins():
    """Obtener orígenes CORS desde variables de entorno de forma segura"""
    origins = []
    
    # Orígenes de desarrollo local (incluye Live Server)
    local_origins = [
        "http://localhost:3000",
        "http://localhost:3001", 
        "http://localhost:5173",  # Vite dev server default port
        "http://localhost:5500",  # Live Server default port
        "http://localhost:8080",  # Webpack dev server
        "http://127.0.0.1:3000",
        "http://127.0.0.1:3001",
        "http://127.0.0.1:5173",  # Vite dev server con 127.0.0.1
        "http://127.0.0.1:5500",  # Live Server con 127.0.0.1
        "http://127.0.0.1:8080",
    ]
    
    # Dominios específicos de producción/hosting
    production_origins = [
        # Netlify apps
        "https://captura-emprestito.netlify.app",
        # Vercel apps
        "https://gestor-proyectos-vercel.vercel.app",
        "https://gestor-proyectos-vercel-5ogb5wph8-juan-pablos-projects-56fe2e60.vercel.app",
        # Artefacto CaliTrack 360 Frontend - Producción y variantes de Vercel
        "https://artefacto-calitrack-360-frontend-production-dbcd9wrsi.vercel.app",
        "https://artefacto-calitrack-360-frontend-production.vercel.app",
        "https://artefacto-calitrack-360-frontend.vercel.app",
        # Agrega aquí otros dominios específicos de producción según sea necesario
    ]
    
    # Siempre incluir dominios de producción
    origins.extend(production_origins)
    
    # Siempre incluir dominios locales (para desarrollo)
    origins.extend(local_origins)
    
    # Orígenes desde variables de entorno
    frontend_url = os.getenv("FRONTEND_URL")
    if frontend_url:
        origins.append(frontend_url)
    
    # Orígenes adicionales (separados por coma)
    additional_origins = os.getenv("CORS_ORIGINS", "")
    if additional_origins:
        origins.extend([origin.strip() for origin in additional_origins.split(",")])
    
    # Eliminar duplicados
    origins = list(set(origins))
    
    return origins

def get_cors_origin_regex():
    """
    Obtener patrón regex para permitir variantes de Vercel dinámicamente.
    Vercel genera URLs como: project-name-hash-team.vercel.app
    """
    # Patrones para proyectos de Vercel que necesitan acceso
    vercel_patterns = [
        r"https://artefacto-calitrack-360-frontend.*\.vercel\.app",
        r"https://gestor-proyectos-vercel.*\.vercel\.app",
    ]
    # Combinar todos los patrones en uno solo
    combined_pattern = "|".join(f"({pattern})" for pattern in vercel_patterns)
    return combined_pattern

# 🔤 MIDDLEWARE UTF-8 PARA CARACTERES ESPECIALES
@app.middleware("http")
async def utf8_middleware(request: Request, call_next):
    """Middleware para asegurar encoding UTF-8 en todas las respuestas"""
    response = await call_next(request)
    
    # Asegurar que las respuestas JSON tengan charset UTF-8
    if response.headers.get("content-type", "").startswith("application/json"):
        response.headers["content-type"] = "application/json; charset=utf-8"
    
    return response

# ⚡ MIDDLEWARE DE PERFORMANCE PARA AGREGAR HEADERS Y MEDIR TIEMPOS
@app.middleware("http")
async def performance_middleware(request: Request, call_next):
    """Middleware para mejorar performance y agregar headers útiles"""
    import time
    start_time = time.time()
    
    response = await call_next(request)
    
    # Calcular tiempo de procesamiento
    process_time = time.time() - start_time
    
    # Agregar headers de performance
    response.headers["X-Process-Time"] = f"{process_time:.3f}"
    response.headers["X-Content-Type-Options"] = "nosniff"
    
    # Sugerir cache para endpoints GET de lectura
    if request.method == "GET" and response.status_code == 200:
        # Cache público para endpoints de datos que no cambian frecuentemente
        if any(path in request.url.path for path in [
            "/centros-gestores/", "/firebase/collections", "/proyectos-presupuestales/",
            "/unidades-proyecto/filters", "/bancos_emprestito", "/auth/config"
        ]):
            response.headers["Cache-Control"] = "public, max-age=300"  # 5 minutos
    
    return response

# 🌐 CONFIGURACIÓN DE CORS
cors_origins = get_cors_origins()
cors_origin_regex = get_cors_origin_regex()
print(f"🌐 CORS configured for {len(cors_origins)} specific origins + regex patterns for Vercel variants")

# Configuración restrictiva con orígenes específicos + regex para variantes de Vercel
# Permite credentials (cookies, tokens) de manera segura
app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,  # Lista específica de orígenes permitidos
    allow_origin_regex=cors_origin_regex,  # Regex para variantes de Vercel
    allow_credentials=True,  # Permitir cookies y headers de autenticación
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "HEAD", "PATCH"],
    allow_headers=[
        "Authorization",
        "Content-Type", 
        "Accept",
        "Accept-Charset",
        "Accept-Encoding",
        "Accept-Language",
        "Origin", 
        "X-Requested-With",
        "Cache-Control",
        "Pragma",
        "X-CSRF-Token",
    ],
    expose_headers=["Content-Type", "Authorization"],
    max_age=600,  # Cache de preflight requests por 10 minutos
)

# 🗜️ GZIP COMPRESSION HABILITADO (optimiza respuestas grandes)
from fastapi.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=1000)  # Comprimir respuestas > 1KB
print("🗜️ GZIP compression enabled for responses > 1KB")

# 🔐 MIDDLEWARE DE AUTENTICACIÓN Y AUTORIZACIÓN
if AUTH_SYSTEM_AVAILABLE and AuthorizationMiddleware is not None:
    # Definir rutas públicas (combinar con las del sistema de auth)
    public_paths = [
        "/",
        "/docs",
        "/openapi.json",
        "/redoc",
        "/ping",
        "/health",
        "/cors-test",
        "/test/utf8",
        "/debug/railway",
        "/metrics",
        "/auth/login",
        "/auth/register",
        "/auth/google",
        "/auth/config",
        "/auth/validate-session",
        "/auth/workload-identity/status"
    ]
    
    app.add_middleware(
        AuthorizationMiddleware,
        public_paths=public_paths
    )
    print("✅ Authorization middleware enabled")
    
    # Middleware de auditoría (opcional, configurar según necesidad)
    if AuditLogMiddleware is not None:
        app.add_middleware(
            AuditLogMiddleware,
            enable_logging=True  # Cambiar a False para deshabilitar logging automático
        )
        print("✅ Audit log middleware enabled")
else:
    print("⚠️ Authorization middleware disabled - Auth system not available")

# ⏱️ MIDDLEWARE DE TIMING Y MONITOREO APM
import time

@app.middleware("http")
async def monitoring_middleware(request: Request, call_next):
    """
    Middleware para monitoreo APM: métricas de latencia, contador de requests, requests activos
    También agrega X-Response-Time header y loguea endpoints lentos
    """
    method = request.method
    endpoint = request.url.path
    
    # Incrementar gauge de requests activos (solo si Prometheus disponible)
    if PROMETHEUS_AVAILABLE and ACTIVE_REQUESTS is not None:
        ACTIVE_REQUESTS.labels(method=method, endpoint=endpoint).inc()
    
    # Medir tiempo de ejecución
    start_time = time.time()
    
    try:
        response = await call_next(request)
        status_code = response.status_code
    except Exception as e:
        status_code = 500
        logger.error(f"Error en {endpoint}: {str(e)}")
        raise
    finally:
        # Decrementar gauge de requests activos (solo si Prometheus disponible)
        if PROMETHEUS_AVAILABLE and ACTIVE_REQUESTS is not None:
            ACTIVE_REQUESTS.labels(method=method, endpoint=endpoint).dec()
        
        # Calcular latencia
        process_time = time.time() - start_time
        
        # Registrar métricas en Prometheus (solo si disponible)
        if PROMETHEUS_AVAILABLE and REQUEST_COUNT is not None and REQUEST_LATENCY is not None:
            REQUEST_COUNT.labels(method=method, endpoint=endpoint, status=status_code).inc()
            REQUEST_LATENCY.labels(method=method, endpoint=endpoint).observe(process_time)
    
    # Agregar header de tiempo de respuesta
    response.headers["X-Response-Time"] = f"{process_time:.3f}s"
    
    # Log solo endpoints lentos (> 3s)
    if process_time > 3.0:
        logger.warning(f"⚠️ Slow endpoint: {endpoint} - {process_time:.3f}s (status: {status_code})")
    
    return response

print("⏱️ Monitoring middleware enabled (APM + Timing)")

# � FUNCIONES UTILITARIAS PARA UTF-8
def create_utf8_response(content: Dict[str, Any], status_code: int = 200) -> JSONResponse:
    """Crear respuesta JSON con encoding UTF-8 explícito"""
    return JSONResponse(
        content=content,
        status_code=status_code,
        headers={"Content-Type": "application/json; charset=utf-8"},
        media_type="application/json"
    )

def handle_utf8_text(text: str) -> str:
    """Asegurar que el texto mantenga caracteres UTF-8"""
    if isinstance(text, str):
        return text.encode('utf-8').decode('utf-8')
    return str(text)

def clean_firebase_data(data):
    """
    Limpia datos de Firebase para serialización JSON
    Convierte DatetimeWithNanoseconds y otros tipos no serializables
    """
    if isinstance(data, dict):
        cleaned = {}
        for key, value in data.items():
            cleaned[key] = clean_firebase_data(value)
        return cleaned
    elif isinstance(data, list):
        return [clean_firebase_data(item) for item in data]
    elif FIREBASE_TYPES_AVAILABLE and isinstance(data, DatetimeWithNanoseconds):
        return data.isoformat()
    elif isinstance(data, datetime):
        return data.isoformat()
    else:
        return data

# �🛠️ MIDDLEWARE DE TIMEOUT PARA PREVENIR COLGADAS
@app.middleware("http")
async def timeout_middleware(request: Request, call_next):
    """Middleware para prevenir que las requests se cuelguen"""
    try:
        # Timeout extendido para endpoints de procesamiento masivo
        if request.url.path == "/emprestito/obtener-procesos-secop":
            # 5 minutos para procesamiento masivo de SECOP
            timeout_seconds = 300.0
        elif request.url.path == "/emprestito/obtener-contratos-secop":
            # 10 minutos para procesamiento masivo de contratos
            timeout_seconds = 600.0
        else:
            # Timeout de 30 segundos para todas las otras requests
            timeout_seconds = 30.0
            
        return await asyncio.wait_for(call_next(request), timeout=timeout_seconds)
    except asyncio.TimeoutError:
        return JSONResponse(
            status_code=504,
            content={
                "error": "Request timeout",
                "message": f"The request took too long to process (timeout: {timeout_seconds}s)",
                "fallback": True,
                "timestamp": datetime.now().isoformat(),
                "endpoint": str(request.url.path)
            }
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "error": "Internal server error", 
                "message": "An unexpected error occurred",
                "fallback": True,
                "timestamp": datetime.now().isoformat()
            }
        )

# Swagger UI configurado automáticamente con parámetros optimizados

# ============================================================================
# ENDPOINTS GENERALES
# ============================================================================

@app.get("/")
async def read_root():
    """Endpoint raíz con información básica de la API"""
    response_data = {
        "message": "Gestor de Proyectos API 🇪🇸",
        "description": "API con soporte completo para UTF-8 y caracteres en español",
        "version": "1.0.0",
        "timestamp": datetime.now().isoformat(),
        "last_updated": "2025-10-04T00:00:00Z",  # API last update date
        "firebase_project": PROJECT_ID,
        "status": "funcionando ✅",
        "encoding": "UTF-8",
        "spanish_support": "Sí - Acentos: á é í ó ú, Ñ, diéresis: ü",
        "documentation": "/docs",
        "environment_debug": {
            "firebase_project_id": os.getenv("FIREBASE_PROJECT_ID", "NOT_SET"),
            "has_service_account_key": bool(os.getenv("FIREBASE_SERVICE_ACCOUNT_KEY")),
            "environment": os.getenv("ENVIRONMENT", "NOT_SET"),
            "port": os.getenv("PORT", "NOT_SET")
        },
        "endpoints": {
            "general": ["/", "/health", "/ping", "/centros-gestores/nombres-unicos"],
            "firebase": ["/firebase/status", "/firebase/collections"], 
            "proyectos_de_inversion": [
                "/proyectos-presupuestales/all",
                "/proyectos-presupuestales/bpin/{bpin}",
                "/proyectos-presupuestales/bp/{bp}",
                "/proyectos-presupuestales/centro-gestor/{nombre_centro_gestor}",
                "/proyectos-presupuestales/cargar-json (POST)"
            ],
            "unidades_proyecto": [
                "/unidades-proyecto/geometry", 
                "/unidades-proyecto/attributes",
                "/unidades-proyecto/dashboard",
                "/unidades-proyecto/filters",
                "/unidades-proyecto/download-geojson",
                "/unidades-proyecto/download-table",
                "/unidades-proyecto/download-table_by_centro_gestor"
            ],
            "gestion_contractual": [
                "/contratos/init_contratos_seguimiento"
            ],
            "gestion_emprestito": [
                "/emprestito/cargar-proceso",
                "/emprestito/cargar-orden-compra",
                "/emprestito/cargar-pago (POST - Registrar pago de empréstito con timestamp automático)",
                "/contratos_pagos_all (GET - Obtener todos los pagos de empréstito)",
                "/emprestito/obtener-procesos-secop (POST - Procesamiento masivo)",
                "/emprestito/proceso/{referencia_proceso}",
                "/emprestito/obtener-contratos-secop",
                "/contratos_emprestito_all",
                "/contratos_emprestito/referencia/{referencia_contrato}",
                "/contratos_emprestito/centro-gestor/{nombre_centro_gestor}",
                "/bancos_emprestito_all",
                "/procesos_emprestito_all",
                "/emprestito/flujo-caja/cargar-excel (POST - Cargar flujos de caja desde Excel)",
                "/emprestito/flujo-caja/all (GET - Consultar flujos de caja con filtros)",
                "/emprestito/crear-tabla-proyecciones (POST - Crear tabla desde Google Sheets)",
                "/emprestito/leer-tabla-proyecciones (GET - Leer proyecciones cargadas)"
            ],
            "administracion_usuarios": [
                "/auth/validate-session",
                "/auth/login", 
                "/auth/register",
                "/auth/change-password",
                "/auth/google",
                "/auth/user/{uid}",
                "/admin/users"
            ]
        },
        "new_features": {
            "user_management": "Sistema completo de gestión de usuarios con Firebase Authentication",
            "auth_methods": "Soporte para email/password, Google (@cali.gov.co), y autenticación telefónica",
            "user_roles": "Sistema de roles y permisos (admin, gestor, viewer, editor)",
            "utf8_support": "Soporte completo para caracteres especiales en español: ñ, á, é, í, ó, ú, ü",
            "filters": "Todos los endpoints de Unidades de Proyecto soportan filtros avanzados",
            "supported_filters": [
                "nombre_centro_gestor", "tipo_intervencion", "estado", "upid", 
                "comuna_corregimiento", "barrio_vereda", "nombre_up", "direccion",
                "referencia_contrato", "referencia_proceso", "include_bbox", "limit", "offset"
            ],
            "dashboard": "Endpoint de dashboard con métricas agregadas y análisis estadístico",
            "workload_identity": "Autenticación automática usando Google Cloud Workload Identity Federation",
            "emprestito_management": "Sistema de gestión de empréstito con integración SECOP y TVEC APIs",
            "duplicate_prevention": "Validación automática de duplicados por referencia_proceso",
            "platform_detection": "Detección automática de plataforma (SECOP/TVEC) y enrutamiento inteligente",
            "external_apis": "Integración con APIs oficiales: SECOP (p6dx-8zbt) y TVEC (rgxm-mmea)",
            "encoding": "UTF-8 completo para español: ñáéíóúü ¡¿"
        }
    }
    
    return create_utf8_response(response_data)

@app.get("/metrics", tags=["Monitoring"])
async def metrics():
    """
    📊 Endpoint de Métricas de Prometheus
    
    Expone métricas de la aplicación en formato Prometheus para monitoreo APM:
    - gestor_api_requests_total: Contador de requests por endpoint, método y status
    - gestor_api_request_duration_seconds: Histograma de latencia de requests
    - gestor_api_requests_active: Gauge de requests activos
    - gestor_api_firebase_queries_total: Contador de queries a Firestore
    - gestor_api_cache_hits_total: Contador de cache hits
    - gestor_api_cache_misses_total: Contador de cache misses
    
    Usar con Grafana + Prometheus para dashboards de monitoreo
    """
    if not PROMETHEUS_AVAILABLE or generate_latest is None or CONTENT_TYPE_LATEST is None:
        raise HTTPException(status_code=503, detail="Prometheus metrics not available")
    
    from fastapi.responses import Response
    return Response(content=generate_latest(), media_type=CONTENT_TYPE_LATEST)

@app.get("/ping", tags=["General"], summary="🔵 Ping Simple")
async def ping():
    """🔵 GET | ❤️ Health Check | Health check super simple para Railway con soporte UTF-8"""
    response_data = {
        "status": "ok ✅", 
        "message": "Servidor funcionando correctamente",
        "encoding": "UTF-8",
        "spanish_test": "ñáéíóúü ¡¿",
        "timestamp": datetime.now().isoformat(),
        "last_updated": "2025-10-04T00:00:00Z"  # Endpoint creation/update date
    }
    return create_utf8_response(response_data)

@app.get("/cors-test", tags=["General"])
async def cors_test(request: Request):
    """Endpoint específico para probar configuración CORS"""
    origin = request.headers.get("origin", "No origin header")
    user_agent = request.headers.get("user-agent", "No user-agent")
    
    response_data = {
        "success": True,
        "message": "CORS test successful ✅",
        "origin": origin,
        "user_agent": user_agent[:100] + "..." if len(user_agent) > 100 else user_agent,
        "cors_configured": True,
        "allowed_methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS", "HEAD", "PATCH"],
        "timestamp": datetime.now().isoformat(),
        "server_info": {
            "environment": os.getenv("ENVIRONMENT", "development"),
            "port": os.getenv("PORT", "8000"),
            "cors_origins_count": len(cors_origins)
        }
    }
    
    # Crear respuesta con headers CORS explícitos adicionales
    response = JSONResponse(
        content=response_data,
        status_code=200,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "Access-Control-Allow-Origin": origin if origin != "No origin header" else "*",
            "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS, HEAD, PATCH",
            "Access-Control-Allow-Headers": "Authorization, Content-Type, Accept, Origin, X-Requested-With",
            "Access-Control-Allow-Credentials": "true"
        }
    )
    
    return response

@app.options("/cors-test", tags=["General"])
async def cors_test_options(request: Request):
    """OPTIONS handler específico para CORS test"""
    origin = request.headers.get("origin", "*")
    
    return JSONResponse(
        content={"message": "CORS preflight OK"},
        status_code=200,
        headers={
            "Access-Control-Allow-Origin": origin,
            "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS, HEAD, PATCH",
            "Access-Control-Allow-Headers": "Authorization, Content-Type, Accept, Origin, X-Requested-With",
            "Access-Control-Allow-Credentials": "true",
            "Access-Control-Max-Age": "86400"
        }
    )

@app.get("/test/utf8", tags=["General"])
async def test_utf8():
    """Endpoint de prueba específico para caracteres UTF-8 en español"""
    test_data = {
        "encoding": "UTF-8",
        "status": "Funcionando correctamente ✅",
        "test_cases": {
            "vocales_acentuadas": "á é í ó ú",
            "vocales_mayusculas": "Á É Í Ó Ú",
            "enie": "ñ Ñ",
            "dieresis": "ü Ü",
            "signos_interrogacion": "¿Cómo estás?",
            "signos_exclamacion": "¡Excelente!",
            "nombres_espanoles": [
                "José María",
                "Ángela Rodríguez", 
                "Peña Nieto",
                "Núñez",
                "Güell"
            ],
            "ciudades_colombia": [
                "Bogotá",
                "Medellín", 
                "Cali",
                "Barranquilla",
                "Cartagena",
                "Cúcuta",
                "Ibagué",
                "Pereira",
                "Santa Marta",
                "Manizales"
            ],
            "texto_completo": "La niña soñó con un colibrí que volaba sobre el jardín donde crecían las flores más hermosas de España.",
            "caracteres_especiales": "°ª€£¢¥§¨©®™",
            "test_json": "Prueba de JSON con acentos: María José fue a Bogotá"
        },
        "timestamp": datetime.now().isoformat()
    }
    
    return create_utf8_response(test_data)



@app.get("/debug/railway", tags=["General"])
async def railway_debug():
    """Debug específico para Railway - Diagnóstico simplificado"""
    try:
        # Variables de entorno
        env_info = {
            "FIREBASE_PROJECT_ID": os.getenv("FIREBASE_PROJECT_ID", "NOT_SET"),
            "HAS_FIREBASE_SERVICE_ACCOUNT_KEY": bool(os.getenv("FIREBASE_SERVICE_ACCOUNT_KEY")),
            "RAILWAY_ENVIRONMENT": os.getenv("RAILWAY_ENVIRONMENT", "NOT_SET"),
            "PORT": os.getenv("PORT", "NOT_SET")
        }
        
        # Test de Service Account
        sa_test = {"status": "not_tested"}
        if os.getenv("FIREBASE_SERVICE_ACCOUNT_KEY"):
            try:
                import json
                import base64
                
                decoded = base64.b64decode(os.getenv("FIREBASE_SERVICE_ACCOUNT_KEY")).decode('utf-8')
                creds_data = json.loads(decoded)
                
                sa_test = {
                    "status": "success",
                    "client_email": creds_data.get("client_email", "missing"),
                    "project_id": creds_data.get("project_id", "missing"),
                    "has_private_key": bool(creds_data.get("private_key"))
                }
            except Exception as e:
                sa_test = {
                    "status": "failed",
                    "error": str(e)
                }
        
        # Test Firebase directly
        firebase_test = None
        if FIREBASE_AVAILABLE:
            try:
                firebase_test = validate_firebase_connection()
            except Exception as e:
                firebase_test = {"error": str(e)}
        
        return {
            "status": "debug_info",
            "timestamp": datetime.now().isoformat(),
            "environment_variables": env_info,
            "service_account_test": sa_test,
            "firebase_test": firebase_test,
            "firebase_available": FIREBASE_AVAILABLE,
            "scripts_available": SCRIPTS_AVAILABLE,
            "project_id_detected": PROJECT_ID,
            "recommendation": "Check service_account_test and firebase_test for issues"
        }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

@app.get("/health", tags=["General"], summary="🔵 Estado de Salud API")
async def health_check():
    """🔵 GET | ❤️ Health Check | Verificar estado de salud de la API"""
    
    # Intentar obtener del cache (TTL 30 segundos para health check)
    cache_key = get_cache_key("health_check")
    cached_data, is_valid = get_from_cache(cache_key, max_age_seconds=30)
    if is_valid:
        # Actualizar timestamp en cada llamada pero mantener resto del cache
        cached_data["timestamp"] = datetime.now().isoformat()
        return cached_data
    
    try:
        basic_response = {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "services": {
                "api": "running"
            },
            "port": os.getenv("PORT", "8000"),
            "environment": os.getenv("ENVIRONMENT", "development"),
            "project_id": PROJECT_ID
        }
        
        # Verificar Firebase usando configuración funcional
        if FIREBASE_AVAILABLE:
            # Test default project
            firebase_status = validate_firebase_connection()
            basic_response["services"]["firebase"] = firebase_status
            basic_response["services"]["scripts"] = {"available": SCRIPTS_AVAILABLE}
            
            # Debug info for Railway
            basic_response["debug"] = {
                "firebase_project_env": os.getenv("FIREBASE_PROJECT_ID", "NOT_SET"),
                "has_sa_key": bool(os.getenv("FIREBASE_SERVICE_ACCOUNT_KEY")),
                "firebase_available": FIREBASE_AVAILABLE,
                "scripts_available": SCRIPTS_AVAILABLE,
                "environment": os.getenv("ENVIRONMENT", "development")
            }
            
            if not firebase_status["connected"]:
                basic_response["status"] = "degraded"
                
        else:
            basic_response["services"]["firebase"] = {
                "available": False, 
                "message": "Firebase SDK not available"
            }

            basic_response["status"] = "degraded"
        
        # Guardar en cache
        set_in_cache(cache_key, basic_response)
        
        return basic_response
        
    except Exception as e:
        print(f"Health check error: {e}")
        return {
            "status": "partial",
            "timestamp": datetime.now().isoformat(),
            "error": str(e)[:100],
            "services": {
                "api": "running"
            }
        }

@app.get("/centros-gestores/nombres-unicos", tags=["General"])
async def get_all_nombres_centros_gestores_unique():
    """
    ## Obtener Nombres Únicos de Centros Gestores
    
    **Propósito**: Retorna una lista de valores únicos del campo "nombre_centro_gestor" 
    de la colección "proyectos_presupuestales".
    
    ### ✅ Casos de uso:
    - Poblar dropdowns y selectores en formularios
    - Filtros dinámicos en dashboards
    - Validación de centros gestores existentes
    - Reportes por centro gestor
    - Análisis de distribución institucional
    
    ### 📊 Características:
    - Valores únicos ordenados alfabéticamente
    - Filtrado automático de valores vacíos o nulos
    - Conteo total de centros gestores únicos
    - Optimizado para carga rápida
    
    ### 🔧 Optimizaciones:
    - Eliminación de duplicados usando set()
    - Normalización de espacios en blanco
    - Ordenamiento alfabético para mejor UX
    - Filtrado de valores vacíos
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const response = await fetch('/centros-gestores/nombres-unicos');
    const data = await response.json();
    if (data.success) {
        console.log('Centros gestores encontrados:', data.count);
        const dropdown = data.data.map(nombre => ({
            value: nombre,
            label: nombre
        }));
    }
    ```
    
    ### 💡 Casos de uso prácticos:
    - **Formularios**: Autocomplete de centros gestores
    - **Dashboards**: Filtros dinámicos por institución
    - **Reportes**: Agrupación por centro gestor
    - **Validación**: Verificar centros gestores válidos
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    # Intentar obtener del cache (TTL 5 minutos)
    cache_key = get_cache_key("centros_gestores_nombres_unicos")
    cached_data, is_valid = get_from_cache(cache_key, max_age_seconds=300)
    if is_valid:
        return cached_data
    
    try:
        result = await get_unique_nombres_centros_gestores()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo nombres únicos de centros gestores: {result.get('error', 'Error desconocido')}"
            )
        
        response_data = {
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "field": result["field"],
            "collection": result["collection"],
            "timestamp": result["timestamp"],
            "last_updated": "2025-10-04T00:00:00Z",  # Endpoint creation date
            "message": f"Se obtuvieron {result['count']} nombres únicos de centros gestores",
            "metadata": {
                "sorted": True,
                "filtered_empty": True,
                "normalized": True,
                "cache_recommended": True,
                "utf8_enabled": True
            }
        }
        
        # Guardar en cache
        set_in_cache(cache_key, response_data)
        
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando nombres únicos de centros gestores: {str(e)}"
        )

# ============================================================================
# ENDPOINTS DE FIREBASE
# ============================================================================

@app.get("/firebase/status", tags=["Firebase"])
async def firebase_status():
    """Verificar estado de la conexión con Firebase"""
    try:
        # Cache corto para evitar consultas repetidas a Firebase en ráfagas
        cache_key = get_cache_key("firebase_status")
        cached = _simple_cache.get(cache_key)
        if cached:
            cached_time = _cache_timestamps.get(cache_key)
            if cached_time and (datetime.now() - cached_time).total_seconds() < 30:
                return cached
        if not FIREBASE_AVAILABLE:
            return {
                "connected": False,
                "error": "Firebase SDK not available",
                "available": False,
                "status": "unavailable",
                "last_updated": "2025-10-02T00:00:00Z"
            }
        
        if not SCRIPTS_AVAILABLE:
            return {
                "connected": False,
                "error": "Scripts not available",
                "available": FIREBASE_AVAILABLE,
                "status": "limited",
                "last_updated": "2025-10-02T00:00:00Z"
            }
        # Realizar comprobación activa de Firebase
        connection_result = await test_firebase_connection()
        connection_result["last_updated"] = "2025-10-02T00:00:00Z"
        # Guardar en cache corto
        set_in_cache(cache_key, connection_result)
        return connection_result
        
    except Exception as e:
        return {
            "connected": False,
            "error": f"Error checking Firebase: {str(e)}",
            "available": FIREBASE_AVAILABLE,
            "status": "error",
            "last_updated": "2025-10-02T00:00:00Z"
        }

@app.get("/firebase/collections", tags=["Firebase"])
@optional_rate_limit("30/minute")  # Máximo 30 requests por minuto
async def get_firebase_collections(request: Request):
    """Obtener información completa de todas las colecciones de Firestore"""
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    # Intentar obtener del cache (TTL 5 minutos)
    cache_key = get_cache_key("firebase_collections")
    cached_data, is_valid = get_from_cache(cache_key, max_age_seconds=300)
    if is_valid:
        return cached_data
    
    try:
        # OPTIMIZACIÓN: Reducir muestreo a 10 documentos por colección para velocidad
        collections_data = await get_collections_info(limit_docs_per_collection=10)
        
        if not collections_data["success"]:
            raise HTTPException(
                status_code=500, 
                detail=f"Error obteniendo información de colecciones: {collections_data.get('error', 'Error desconocido')}"
            )
        
        # Add timestamp for endpoint tracking
        collections_data["last_updated"] = "2025-10-02T00:00:00Z"  # Endpoint creation/update date
        
        # Guardar en cache
        set_in_cache(cache_key, collections_data)
        
        return collections_data
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Error interno del servidor: {str(e)}"
        )

@app.get("/firebase/collections/summary", tags=["Firebase"])
@optional_rate_limit("30/minute")  # Máximo 30 requests por minuto
async def get_firebase_collections_summary(request: Request):
    """Obtener resumen estadístico de las colecciones"""
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    # Intentar obtener del cache (TTL 5 minutos)
    cache_key = get_cache_key("firebase_collections_summary")
    cached_data, is_valid = get_from_cache(cache_key, max_age_seconds=300)
    if is_valid:
        return cached_data
    
    try:
        summary_data = await get_collections_summary()
        
        if not summary_data["success"]:
            raise HTTPException(
                status_code=500, 
                detail=f"Error obteniendo resumen: {summary_data.get('error', 'Error desconocido')}"
            )
        
        # Add timestamp for endpoint tracking
        summary_data["last_updated"] = "2025-10-02T00:00:00Z"  # Endpoint creation/update date
        
        # Guardar en cache
        set_in_cache(cache_key, summary_data)
        
        return summary_data
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo resumen: {str(e)}")

# ============================================================================
# ENDPOINTS DE PROYECTOS DE INVERSIÓN
# ============================================================================

@app.get("/proyectos-presupuestales/all", tags=["Proyectos de Inversión"], summary="🔵 Todos los Proyectos Presupuestales")
@optional_rate_limit("40/minute")  # Máximo 40 requests por minuto (endpoint costoso)
@async_cache(ttl_seconds=300)  # Cache de 5 minutos para proyectos
async def get_proyectos_all(request: Request):
    """
    ## 🔵 GET | 📋 Listados | Obtener Todos los Proyectos Presupuestales
    
    **Propósito**: Retorna todos los documentos de la colección "proyectos_presupuestales".
    
    ### ✅ Casos de uso:
    - Obtener listado completo de proyectos presupuestales
    - Exportación de datos para análisis
    - Integración con sistemas externos
    - Reportes y dashboards de proyectos de inversión
    
    ### 📊 Información incluida:
    - Todos los campos disponibles en la colección
    - ID del documento para referencia
    - Conteo total de registros
    - Timestamp de la consulta
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const response = await fetch('/proyectos-presupuestales/all');
    const data = await response.json();
    if (data.success) {
        console.log('Proyectos encontrados:', data.count);
        console.log('Datos:', data.data);
    }
    ```
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        result = await get_proyectos_presupuestales()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo proyectos presupuestales: {result.get('error', 'Error desconocido')}"
            )
        
        return {
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "timestamp": result["timestamp"],
            "last_updated": "2025-10-04T00:00:00Z",  # Endpoint creation date
            "message": f"Se obtuvieron {result['count']} proyectos presupuestales exitosamente"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando proyectos presupuestales: {str(e)}"
        )

@app.get("/proyectos-presupuestales/bpin/{bpin}", tags=["Proyectos de Inversión"], summary="🔵 Proyectos por BPIN")
async def get_proyectos_by_bpin(bpin: str):
    """
    ## 🔵 GET | 🔍 Consultas | Obtener Proyectos por BPIN
    
    **Propósito**: Retorna proyectos presupuestales filtrados por código BPIN específico.
    
    ### ✅ Casos de uso:
    - Búsqueda de proyectos por código BPIN específico
    - Consulta de detalles de proyecto individual
    - Validación de existencia de BPIN
    - Integración con sistemas de seguimiento presupuestal
    
    ### 🔍 Filtrado:
    - **Campo**: `bpin` (coincidencia exacta)
    - **Tipo**: String - Código único del proyecto
    - **Sensible a mayúsculas**: Sí
    
    ### 📊 Información incluida:
    - Todos los campos del proyecto que coincida con el BPIN
    - ID del documento para referencia
    - Conteo de registros encontrados
    - Información del filtro aplicado
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const bpin = "2023000123456";
    const response = await fetch(`/proyectos-presupuestales/bpin/${bpin}`);
    const data = await response.json();
    if (data.success && data.count > 0) {
        console.log('Proyecto encontrado:', data.data[0]);
    } else {
        console.log('No se encontró proyecto con BPIN:', bpin);
    }
    ```
    
    ### 💡 Notas:
    - Si no se encuentra ningún proyecto, retorna array vacío
    - El BPIN debe ser exacto (sin espacios adicionales)
    - Típicamente retorna 0 o 1 resultado (BPIN único)
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        result = await get_proyectos_presupuestales_by_bpin(bpin)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo proyectos por BPIN: {result.get('error', 'Error desconocido')}"
            )
        
        return {
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "filter": result["filter"],
            "timestamp": result["timestamp"],
            "last_updated": "2025-10-04T00:00:00Z",  # Endpoint creation date
            "message": f"Se encontraron {result['count']} proyectos con BPIN '{bpin}'"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta por BPIN: {str(e)}"
        )

@app.get("/proyectos-presupuestales/bp/{bp}", tags=["Proyectos de Inversión"])
async def get_proyectos_by_bp(bp: str):
    """
    ## Obtener Proyectos Presupuestales por BP
    
    **Propósito**: Retorna proyectos presupuestales filtrados por código BP específico.
    
    ### ✅ Casos de uso:
    - Búsqueda de proyectos por código BP específico
    - Consulta de proyectos relacionados por BP
    - Análisis de agrupación presupuestal
    - Reportes por código de proyecto base
    
    ### 🔍 Filtrado:
    - **Campo**: `bp` (coincidencia exacta)
    - **Tipo**: String - Código base del proyecto
    - **Sensible a mayúsculas**: Sí
    
    ### 📊 Información incluida:
    - Todos los campos de los proyectos que coincidan con el BP
    - ID del documento para referencia
    - Conteo de registros encontrados
    - Información del filtro aplicado
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const bp = "BP-2023-001";
    const response = await fetch(`/proyectos-presupuestales/bp/${bp}`);
    const data = await response.json();
    if (data.success && data.count > 0) {
        console.log(`Encontrados ${data.count} proyectos con BP:`, bp);
        data.data.forEach(proyecto => {
            console.log('Proyecto:', proyecto.nombre_proyecto);
        });
    }
    ```
    
    ### 💡 Notas:
    - Puede retornar múltiples proyectos (un BP puede tener varios proyectos)
    - Si no se encuentra ningún proyecto, retorna array vacío
    - El BP debe ser exacto (sin espacios adicionales)
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        result = await get_proyectos_presupuestales_by_bp(bp)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo proyectos por BP: {result.get('error', 'Error desconocido')}"
            )
        
        return {
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "filter": result["filter"],
            "timestamp": result["timestamp"],
            "last_updated": "2025-10-04T00:00:00Z",  # Endpoint creation date
            "message": f"Se encontraron {result['count']} proyectos con BP '{bp}'"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta por BP: {str(e)}"
        )

@app.get("/proyectos-presupuestales/centro-gestor/{nombre_centro_gestor}", tags=["Proyectos de Inversión"])
async def get_proyectos_by_centro_gestor(nombre_centro_gestor: str):
    """
    ## Obtener Proyectos Presupuestales por Centro Gestor
    
    **Propósito**: Retorna proyectos presupuestales filtrados por nombre del centro gestor específico.
    
    ### ✅ Casos de uso:
    - Consulta de proyectos por dependencia responsable
    - Reportes por entidad gestora
    - Dashboard por centro de responsabilidad
    - Análisis de distribución institucional
    - Seguimiento de proyectos por secretaría/departamento
    
    ### 🔍 Filtrado:
    - **Campo**: `nombre_centro_gestor` (coincidencia exacta)
    - **Tipo**: String - Nombre completo del centro gestor
    - **Sensible a mayúsculas**: Sí
    - **Espacios**: Sensible a espacios adicionales
    
    ### 📊 Información incluida:
    - Todos los campos de los proyectos del centro gestor
    - ID del documento para referencia
    - Conteo de registros encontrados
    - Información del filtro aplicado
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const centroGestor = "Secretaría de Salud";
    const response = await fetch(`/proyectos-presupuestales/centro-gestor/${encodeURIComponent(centroGestor)}`);
    const data = await response.json();
    if (data.success && data.count > 0) {
        console.log(`${data.count} proyectos encontrados para:`, centroGestor);
        const totalPresupuesto = data.data.reduce((sum, p) => sum + (p.presupuesto || 0), 0);
        console.log('Presupuesto total:', totalPresupuesto);
    }
    ```
    
    ### 💡 Notas:
    - Típicamente retorna múltiples proyectos por centro gestor
    - El nombre debe ser exacto (use `/centros-gestores/nombres-unicos` para obtener nombres válidos)
    - Para nombres con espacios, usar `encodeURIComponent()` en el frontend
    - Si no se encuentra ningún proyecto, retorna array vacío
    
    ### 🔗 Endpoint relacionado:
    - `GET /centros-gestores/nombres-unicos` - Para obtener lista de centros gestores válidos
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        result = await get_proyectos_presupuestales_by_centro_gestor(nombre_centro_gestor)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo proyectos por centro gestor: {result.get('error', 'Error desconocido')}"
            )
        
        return {
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "filter": result["filter"],
            "timestamp": result["timestamp"],
            "last_updated": "2025-10-04T00:00:00Z",  # Endpoint creation date
            "message": f"Se encontraron {result['count']} proyectos para el centro gestor '{nombre_centro_gestor}'"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta por centro gestor: {str(e)}"
        )

@app.post("/proyectos-presupuestales/cargar-json", tags=["Proyectos de Inversión"], summary="🟢 Cargar JSON Proyectos")
async def cargar_proyectos_presupuestales_json(
    archivo_json: UploadFile = File(..., description="Archivo JSON con proyectos presupuestales"),
    update_mode: str = Form(default="merge", description="Modo de actualización: merge, replace, append")
):
    """
    ## � POST | �📊 Carga de Archivos | Cargar Proyectos desde JSON
    
    Endpoint POST para subir un archivo JSON con información de proyectos presupuestales 
    y cargarlo en la colección "proyectos_presupuestales".
    
    ### 📁 Archivo JSON esperado:
    ```json
    [
        {
            "nombre_proyecto": "Construcción de Puente",
            "bpin": "2023000123456",
            "bp": "BP-2023-001", 
            "nombre_centro_gestor": "Secretaría de Infraestructura",
            "valor_proyecto": 500000000
        },
        {
            "nombre_proyecto": "Otro Proyecto",
            "bpin": "2023000789012"
        }
    ]
    ```
    
    ### 🔧 Modos de actualización:
    - **merge**: Actualiza existentes y crea nuevos (por defecto)
    - **replace**: Reemplaza toda la colección
    - **append**: Solo agrega nuevos
    
    ### 🎯 Cómo usar:
    1. Haz clic en "Choose File" 
    2. Selecciona tu archivo .json
    3. Selecciona el modo de actualización
    4. Haz clic en "Execute"
    
    ### ✅ Validaciones:
    - Solo archivos .json
    - Cada proyecto debe tener "nombre_proyecto"
    - Tamaño máximo: 10MB
    """
    # Verificar disponibilidad de servicios
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase o scripts no disponibles")
    
    if not PROYECTOS_PRESUPUESTALES_OPERATIONS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Operaciones de proyectos presupuestales no disponibles")
    
    # Validar modo de actualización
    if update_mode not in ["merge", "replace", "append"]:
        raise HTTPException(status_code=400, detail="update_mode debe ser: merge, replace o append")
    
    # Validar tipo de archivo
    if not archivo_json.filename.lower().endswith('.json'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos JSON (.json)")
    
    # Validar tamaño del archivo (10MB máximo)
    max_size = 10 * 1024 * 1024  # 10MB
    if archivo_json.size and archivo_json.size > max_size:
        raise HTTPException(status_code=400, detail="El archivo no puede exceder 10MB")
    
    try:
        # Leer el contenido del archivo
        contenido = await archivo_json.read()
        
        # Decodificar como JSON
        try:
            json_data = json.loads(contenido.decode('utf-8'))
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"Error al leer JSON: {str(e)}")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="El archivo debe estar codificado en UTF-8")
        
        # Validar que sea una lista
        if not isinstance(json_data, list):
            raise HTTPException(status_code=400, detail="El JSON debe ser una lista de proyectos")
        
        if len(json_data) == 0:
            raise HTTPException(status_code=400, detail="La lista no puede estar vacía")
        
        # Validar que cada proyecto tenga nombre_proyecto
        for i, proyecto in enumerate(json_data):
            if not isinstance(proyecto, dict):
                raise HTTPException(status_code=400, detail=f"El elemento {i} debe ser un objeto")
            if not proyecto.get("nombre_proyecto"):
                raise HTTPException(status_code=400, detail=f"El proyecto {i} debe tener 'nombre_proyecto'")
        
        # Procesar proyectos
        result = await process_proyectos_presupuestales_json(
            proyectos_data=json_data,
            update_mode=update_mode
        )
        
        if not result["success"]:
            raise HTTPException(status_code=400, detail=result.get('error', 'Error desconocido'))
        
        # Agregar información del archivo procesado
        result["archivo_info"] = {
            "nombre_archivo": archivo_json.filename,
            "tamaño_bytes": len(contenido),
            "proyectos_en_archivo": len(json_data),
            "update_mode_usado": update_mode
        }
        
        return create_utf8_response(result)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")



# ============================================================================
# ENDPOINTS DE UNIDADES DE PROYECTO
# ============================================================================

@app.get("/unidades-proyecto/geometry", tags=["Unidades de Proyecto"], summary="🔵 Geometrías Completas")
@optional_rate_limit("60/minute")  # Máximo 60 requests por minuto (endpoint pesado)
async def export_geometry_for_nextjs(
    request: Request,
    # Filtros server-side optimizados
    nombre_centro_gestor: Optional[str] = Query(None, description="Centro gestor responsable"),
    tipo_intervencion: Optional[str] = Query(None, description="Tipo de intervención"),
    estado: Optional[str] = Query(None, description="Estado del proyecto"),
    upid: Optional[str] = Query(None, description="ID específico de unidad"),
    clase_up: Optional[str] = Query(None, description="Clase de la unidad de proyecto"),
    tipo_equipamiento: Optional[str] = Query(None, description="Tipo de equipamiento del proyecto"),
    
    # Filtros geográficos adicionales
    comuna_corregimiento: Optional[str] = Query(None, description="Comuna o corregimiento específico"),
    barrio_vereda: Optional[str] = Query(None, description="Barrio o vereda específico"),
    
    # Filtros de visualización y análisis
    presupuesto_base: Optional[float] = Query(None, ge=0, description="Presupuesto mínimo del proyecto"),
    avance_obra: Optional[float] = Query(None, ge=0, le=100, description="Porcentaje mínimo de avance de obra"),
    frente_activo: Optional[str] = Query(None, description="Frente activo del proyecto"),
    
    # Configuración geográfica
    include_bbox: Optional[bool] = Query(False, description="Calcular y incluir bounding box"),
    limit: Optional[int] = Query(None, ge=1, le=10000, description="Límite de registros"),
    
    # Parámetros de mantenimiento y debug
    force_refresh: Optional[str] = Query(None, description="Forzar limpieza de cache (debug)"),
    debug: Optional[bool] = Query(False, description="Modo debug con información adicional")
):
    """
    ## 🔵 GET | 🗺️ Datos Geoespaciales | Datos Geoespaciales Completos
    
    **Propósito**: Retorna TODOS los registros de proyectos en formato GeoJSON con soporte completo para:
    - LineString, MultiLineString, Polygon, MultiPolygon
    - GeometryCollection (geometrías unificadas)
    - Todas las propiedades del proyecto (nombre_up, centro_gestor, etc.)
    
    ### Geometrías Soportadas
    
    **Simples**: Point, LineString, Polygon
    **Multi**: MultiPoint, MultiLineString, MultiPolygon  
    **Complejas**: GeometryCollection (resultado de unificación de features)
    
    ### Parámetros de Filtrado
    
    | Filtro | Descripción |
    |--------|-------------|
    | upid | ID específico de unidad (ej: UNP-1000) |
    | nombre_centro_gestor | Centro gestor responsable |
    | tipo_equipamiento | Tipo de equipamiento (ej: Vías) |
    | comuna_corregimiento | Comuna o corregimiento |
    | limit | Límite de resultados (1-10000) |
    | debug | Incluir información de depuración |
    """
    # Verificación robusta de Firebase
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        return create_utf8_response({
            "type": "FeatureCollection",
            "features": [],
            "properties": {
                "success": False,
                "error": "Firebase not available",
                "count": 0
            }
        }, status_code=503)
    
    try:
        # Construir filtros optimizados para geometrías
        filters = {}
        
        if nombre_centro_gestor:
            filters["nombre_centro_gestor"] = nombre_centro_gestor
        if tipo_intervencion:
            filters["tipo_intervencion"] = tipo_intervencion
        if estado:
            filters["estado"] = estado
        if upid:
            filters["upid"] = upid
        if clase_up:
            filters["clase_up"] = clase_up
        if tipo_equipamiento:
            filters["tipo_equipamiento"] = tipo_equipamiento
        if comuna_corregimiento:
            filters["comuna_corregimiento"] = comuna_corregimiento
        if barrio_vereda:
            filters["barrio_vereda"] = barrio_vereda
        if presupuesto_base is not None:
            filters["presupuesto_base"] = presupuesto_base
        if avance_obra is not None:
            filters["avance_obra"] = avance_obra
        if frente_activo:
            filters["frente_activo"] = frente_activo
        if limit:
            filters["limit"] = limit
        if include_bbox:
            filters["include_bbox"] = include_bbox
        if force_refresh:
            filters["force_refresh"] = force_refresh
        
        result = await get_unidades_proyecto_geometry(filters)
        
        # Agregar información de debug si se solicita
        if debug and result.get("type") == "FeatureCollection":
            result["properties"]["debug"] = {
                "filters_applied": filters,
                "server_version": "2.0-geometry-collection-support",
                "timestamp": datetime.utcnow().isoformat()
            }
        
        # Manejar el formato correcto de respuesta
        if result.get("type") == "FeatureCollection":
            # Respuesta GeoJSON exitosa - retornar directamente
            return create_utf8_response(result)
        elif result.get("success") is False:
            # Respuesta de error
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo geometrías: {result.get('error', 'Error desconocido')}"
            )
        else:
            # Formato inesperado
            raise HTTPException(
                status_code=500,
                detail="Formato de respuesta inesperado del servicio de geometrías"
            )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando geometrías: {str(e)}"
        )

@app.get("/unidades-proyecto/attributes", tags=["Unidades de Proyecto"], summary="🔵 GET | 📊 Datos Tabulares | Atributos Tabulares")
@optional_rate_limit("60/minute")  # Máximo 60 requests por minuto
async def export_attributes_for_nextjs(
    request: Request,
    # Filtros básicos originales
    nombre_centro_gestor: Optional[str] = Query(None, description="Centro gestor responsable"),
    tipo_intervencion: Optional[str] = Query(None, description="Tipo de intervención"),
    estado: Optional[str] = Query(None, description="Estado del proyecto"),
    upid: Optional[str] = Query(None, description="ID específico de unidad"),
    clase_obra: Optional[str] = Query(None, description="Clase de obra del proyecto"),
    tipo_equipamiento: Optional[str] = Query(None, description="Tipo de equipamiento del proyecto"),
    nombre_up: Optional[str] = Query(None, description="Búsqueda parcial en nombre (contiene texto)"),
    comuna_corregimiento: Optional[str] = Query(None, description="Comuna o corregimiento"),
    barrio_vereda: Optional[str] = Query(None, description="Barrio o vereda"),
    direccion: Optional[str] = Query(None, description="Búsqueda parcial en dirección (contiene texto)"),
    referencia_contrato: Optional[str] = Query(None, description="Referencia del contrato"),
    referencia_proceso: Optional[str] = Query(None, description="Referencia del proceso"),
    frente_activo: Optional[str] = Query(None, description="Frente activo del proyecto"),
    
    # Paginación
    limit: Optional[int] = Query(None, ge=1, le=1000, description="Máximo de resultados"),
    offset: Optional[int] = Query(None, ge=0, description="Saltar registros para paginación")
):
    """
    ## 🔵 GET | 📊 Datos Tabulares | Atributos Tabulares
    
    **Propósito**: Retorna atributos completos de proyectos excluyendo datos geográficos.
    
    ### Optimización de Datos
    
    **Campos incluidos**: Todos los atributos del proyecto (nombres, estados, referencias, etc.)
    **Campos excluidos**: coordinates, geometry, linestring, polygon, lat, lng y similares
    **Paginación**: Sistema limit/offset para manejo eficiente de grandes volúmenes
    
    ### Estrategia de Filtrado
    
    **Sin filtros**: Dataset completo de atributos  
    **Con filtros**: Optimización server-side + filtros client-side específicos
    
    **Server-side**: upid, estado, tipo_intervencion, nombre_centro_gestor  
    **Client-side**: search, nombre_up, direccion, ubicación geográfica
    
    ### Parámetros
    
    | Filtro | Descripción |
    |--------|-------------|
    | nombre_centro_gestor | Centro gestor responsable |
    | tipo_intervencion | Tipo de intervención |
    | estado | Estado del proyecto |
    | upid | ID específico de unidad |
    | clase_up | Clase de la unidad de proyecto |
    | tipo_equipamiento | Tipo de equipamiento del proyecto |
    | nombre_up | Búsqueda parcial en nombre |
    | comuna_corregimiento | Comuna o corregimiento |
    | barrio_vereda | Barrio o vereda |
    | direccion | Búsqueda parcial en dirección |
    | referencia_contrato | Referencia del contrato |
    | referencia_proceso | Referencia del proceso |
    | **limit** | Máximo resultados (1-1000) |
    | **offset** | Registros a omitir |
    
    ### Aplicaciones
    
    - Grillas de datos y tablas administrativas
    - Reportes tabulares con filtros múltiples
    - Exportación a formatos estructurados
    - Interfaces de búsqueda avanzada
    """
    # Verificación robusta de Firebase con reintentos
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        # Intentar reconfigurar Firebase como último recurso
        try:
            print("⚠️ Attempting Firebase reconfiguration...")
            firebase_initialized, status = configure_firebase()
            if firebase_initialized:
                print("✅ Firebase reconfiguration successful")
            else:
                print(f"❌ Firebase reconfiguration failed: {status.get('error', 'Unknown error')}")
                return {
                    "success": False,
                    "error": "Firebase not available - check Railway environment variables",
                    "data": [],
                    "count": 0,
                    "type": "attributes",
                    "help": "Verify FIREBASE_SERVICE_ACCOUNT_KEY or GOOGLE_APPLICATION_CREDENTIALS_JSON",
                    "railway_fix": "Run generate_railway_fallback.py to create Service Account fallback"
                }
        except Exception as e:
            return {
                "success": False,
                "error": f"Firebase configuration failed: {str(e)}",
                "data": [],
                "count": 0,
                "type": "attributes",
                "help": "Check Railway environment variables or use Service Account fallback"
            }
    
    try:
        # Construir filtros
        filters = {}
        
        if nombre_centro_gestor:
            filters["nombre_centro_gestor"] = nombre_centro_gestor
        if tipo_intervencion:
            filters["tipo_intervencion"] = tipo_intervencion
        if estado:
            filters["estado"] = estado
        if upid:
            filters["upid"] = upid
        if clase_obra:
            filters["clase_obra"] = clase_obra
        if tipo_equipamiento:
            filters["tipo_equipamiento"] = tipo_equipamiento
        if nombre_up:
            filters["nombre_up"] = nombre_up
        if comuna_corregimiento:
            filters["comuna_corregimiento"] = comuna_corregimiento
        if barrio_vereda:
            filters["barrio_vereda"] = barrio_vereda
        if direccion:
            filters["direccion"] = direccion
        if referencia_contrato:
            filters["referencia_contrato"] = referencia_contrato
        if referencia_proceso:
            filters["referencia_proceso"] = referencia_proceso
        if frente_activo:
            filters["frente_activo"] = frente_activo
        
        result = await get_unidades_proyecto_attributes(
            filters=filters,
            limit=limit,
            offset=offset
        )
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo atributos: {result.get('error', 'Error desconocido')}"
            )
        
        response_data = {
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "total_before_limit": result.get("total_before_limit"),
            "type": "attributes",
            "collection": "unidades-proyecto",
            "filters_applied": result.get("filters_applied", {}),
            "pagination": result.get("pagination", {}),
            "timestamp": datetime.now().isoformat(),
            "last_updated": "2025-10-02T00:00:00Z",  # Endpoint creation/update date
            "message": result.get("message", "Atributos obtenidos exitosamente")
        }
        
        return create_utf8_response(response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando atributos: {str(e)}"
        )

# ============================================================================
# ENDPOINT PARA ARTEFACTO DE CAPTURA #360
# ============================================================================

@app.get("/unidades-proyecto/init-360", tags=["Artefacto de Captura #360"], summary="🔵 GET | 📋 Listados | Datos Iniciales para Captura #360")
@optional_rate_limit("60/minute")
async def get_unidades_proyecto_init_360(request: Request):
    """
    ## 🔵 GET | 📋 Listados | Obtener Datos Iniciales para Artefacto de Captura #360
    
    **Propósito**: Retorna registros de la colección "unidades_proyecto" filtrados según 
    criterios específicos para el artefacto de captura #360.
    
    ### ✅ Campos retornados:
    - upid
    - nombre_up
    - nombre_up_detalle
    - tipo_equipamiento
    - tipo_intervencion
    - estado
    - avance_obra
    - presupuesto_base
    - geometry (datos geoespaciales del registro)
    - direccion
    
    ### 🚫 Exclusiones aplicadas:
    
    **Por clase_up**:
    - "Interventoría"
    - "Estudios y diseños"
    - "Subsidios"
    
    **Por tipo_equipamiento**:
    - "Fuentes y monumentos"
    - "Parques y zonas verdes"
    - "Vivienda mejoramiento"
    - "Vivienda nueva"
    - "Adquisición predios"
    
    **Por tipo_intervencion**:
    - "Estudios y diseños"
    - "Transferencia directa"
    
    ### 📊 Información incluida en la respuesta:
    - Lista de registros que cumplen los criterios
    - Conteo total de registros retornados
    - Timestamp de la consulta
    - Criterios de exclusión aplicados
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const response = await fetch('/unidades-proyecto/init-360');
    const data = await response.json();
    if (data.success) {
        console.log('Registros encontrados:', data.count);
        console.log('Datos:', data.data);
    }
    ```
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        # Conectar a Firestore
        db = get_firestore_client()
        if db is None:
            raise HTTPException(
                status_code=503,
                detail="No se pudo conectar a Firestore"
            )
        
        # Definir criterios de exclusión
        exclusion_clase_up = ["Interventoría", "Estudios y diseños", "Subsidios"]
        exclusion_tipo_equipamiento = [
            "Fuentes y monumentos",
            "Parques y zonas verdes",
            "Vivienda mejoramiento",
            "Vivienda nueva",
            "Adquisición predios"
        ]
        exclusion_tipo_intervencion = ["Estudios y diseños", "Transferencia directa"]
        
        # Campos a retornar
        campos_requeridos = [
            'upid',
            'nombre_up',
            'nombre_up_detalle',
            'tipo_equipamiento',
            'tipo_intervencion',
            'estado',
            'avance_obra',
            'presupuesto_base',
            'geometry',
            'direccion'
        ]
        
        # Consultar colección
        query = db.collection('unidades_proyecto')
        docs = query.stream()
        
        # Procesar documentos
        registros_filtrados = []
        
        for doc in docs:
            doc_data = doc.to_dict()
            
            # Extraer campos, buscando en el nivel raíz y en properties
            def get_field_value(field_name):
                """Obtener valor del campo desde el documento o properties"""
                if field_name in doc_data:
                    return doc_data[field_name]
                elif 'properties' in doc_data and field_name in doc_data['properties']:
                    return doc_data['properties'][field_name]
                return None
            
            # Obtener valores para filtrado
            clase_up = get_field_value('clase_up')
            tipo_equipamiento = get_field_value('tipo_equipamiento')
            tipo_intervencion = get_field_value('tipo_intervencion')
            
            # Aplicar filtros de exclusión
            # Excluir si clase_up está en la lista de exclusión
            if clase_up and clase_up in exclusion_clase_up:
                continue
            
            # Excluir si tipo_equipamiento está en la lista de exclusión
            if tipo_equipamiento and tipo_equipamiento in exclusion_tipo_equipamiento:
                continue
            
            # Excluir si tipo_intervencion está en la lista de exclusión
            if tipo_intervencion and tipo_intervencion in exclusion_tipo_intervencion:
                continue
            
            # Si pasa todos los filtros, extraer campos requeridos
            registro = {}
            for campo in campos_requeridos:
                valor = get_field_value(campo)
                registro[campo] = valor
            
            registros_filtrados.append(registro)
        
        # Preparar respuesta
        response_data = {
            "success": True,
            "data": registros_filtrados,
            "count": len(registros_filtrados),
            "collection": "unidades_proyecto",
            "timestamp": datetime.now().isoformat(),
            "last_updated": "2025-11-26T00:00:00Z",
            "message": f"Se obtuvieron {len(registros_filtrados)} registros que cumplen los criterios del artefacto #360",
            "filters_applied": {
                "excluded_clase_up": exclusion_clase_up,
                "excluded_tipo_equipamiento": exclusion_tipo_equipamiento,
                "excluded_tipo_intervencion": exclusion_tipo_intervencion
            },
            "fields_returned": campos_requeridos
        }
        
        return create_utf8_response(response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta init-360: {str(e)}"
        )

# ============================================================================
# ENDPOINT PARA OPCIONES DE FILTROS
# ============================================================================

@app.get("/unidades-proyecto/filters", tags=["Unidades de Proyecto"], response_class=JSONResponse)
async def get_filters_endpoint(
    field: Optional[str] = Query(
        None, 
        description="Campo específico para obtener valores únicos (opcional)",
        enum=[
            "estado", "tipo_intervencion", "nombre_centro_gestor", 
            "comuna_corregimiento", "barrio_vereda", "fuente_financiacion", 
            "ano", "clase_up", "frente_activo"
        ]
    ),
    limit: Optional[int] = Query(
        None, 
        description="Límite de valores únicos a retornar (opcional)", 
        ge=1,
        le=100
    )
):
    """
    **Obtener valores únicos para filtros de Unidades de Proyecto**
    
    Endpoint optimizado para poblar controles de filtrado en dashboards y interfaces.
    Diseñado específicamente para aplicaciones NextJS con carga eficiente de opciones.
    
    **Características principales:**
    - **Filtrado inteligente**: Especifica un campo para cargar solo sus valores
    - **Control de volumen**: Aplica límites para evitar sobrecarga de datos  
    - **Optimización server-side**: Usa queries eficientes de Firestore
    - **Cache-friendly**: Estructura optimizada para sistemas de caché
    
    **Casos de uso:**
    - Poblar dropdowns y selectores en dashboards
    - Cargar opciones de filtrado dinámicamente
    - Implementar autocomplete y búsqueda predictiva
    - Validar valores disponibles antes de filtrar
    
    **Campos disponibles:**
    - `estado`: Estados de proyecto (activo, completado, etc.)
    - `tipo_intervencion`: Tipos de intervención urbana
    - `nombre_centro_gestor`: Centros gestores responsables
    - `comuna_corregimiento`: Ubicaciones por comuna/corregimiento
    - `barrio_vereda`: Ubicaciones por barrio/vereda
    - `fuente_financiacion`: Fuentes de financiación del proyecto
    - `ano`: Años de ejecución disponibles
    - `departamento`: Departamentos con proyectos
    - `municipio`: Municipios con proyectos
    
    **Optimizaciones aplicadas:**
    - Sampling inteligente de documentos para reducir latencia
    - Filtros server-side en Firestore para mejor rendimiento
    - Límites configurables para controlar payload
    - Estructura de respuesta optimizada para frontend
    """
    # Verificación robusta de Firebase con reintentos
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        # Intentar reconfigurar Firebase como último recurso
        try:
            print("⚠️ Attempting Firebase reconfiguration...")
            firebase_initialized, status = configure_firebase()
            if firebase_initialized:
                print("✅ Firebase reconfiguration successful")
            else:
                print(f"❌ Firebase reconfiguration failed: {status.get('error', 'Unknown error')}")
                return {
                    "success": False,
                    "error": "Firebase not available - check Railway environment variables",
                    "filters": {},
                    "type": "filters",
                    "help": "Verify FIREBASE_SERVICE_ACCOUNT_KEY or GOOGLE_APPLICATION_CREDENTIALS_JSON",
                    "railway_fix": "Run generate_railway_fallback.py to create Service Account fallback"
                }
        except Exception as e:
            return {
                "success": False,
                "error": f"Firebase configuration failed: {str(e)}",
                "filters": {},
                "type": "filters",
                "help": "Check Railway environment variables or use Service Account fallback"
            }
    
    # Intentar obtener del cache (TTL 5 minutos)
    cache_key = get_cache_key(f"unidades_filters_{field}_{limit}")
    cached_data, is_valid = get_from_cache(cache_key, max_age_seconds=300)
    if is_valid:
        return cached_data
    
    try:
        result = await get_filter_options(field=field, limit=limit)
        
        if not result.get("success", False):
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo filtros: {result.get('error', 'Error desconocido')}"
            )
        
        response_data = {
            "success": True,
            "filters": result["filters"],
            "metadata": {
                "total_fields": result.get("total_fields", 0),
                "field_requested": result.get("field_requested"),
                "limit_applied": result.get("limit_applied"),
                "optimized_query": True,
                "cache_recommended": True,
                "utf8_enabled": True,
                "spanish_support": True
            },
            "type": "filters",
            "collection": "unidades-proyecto", 
            "timestamp": datetime.now().isoformat(),
            "last_updated": "2025-10-02T00:00:00Z",  # Endpoint creation/update date
            "message": f"Filtros obtenidos exitosamente"
        }
        
        # Guardar en cache
        set_in_cache(cache_key, response_data)
        
        return create_utf8_response(response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando filtros: {str(e)}"
        )


# ============================================================================
# NUEVOS ENDPOINTS PARA ESTRUCTURA CON INTERVENCIONES
# ============================================================================

@app.get("/unidades-proyecto/{upid}", tags=["Unidades de Proyecto"], summary="🔵 GET | Unidad Específica con Intervenciones")
@optional_rate_limit("60/minute")
async def get_unidad_by_upid(
    upid: str = Path(..., description="ID único de la unidad de proyecto (ej: UNP-1978)")
):
    """
    ## 🔵 GET | Obtener Unidad de Proyecto Específica
    
    **Propósito**: Retorna una unidad de proyecto específica con todas sus intervenciones.
    
    ### Estructura de Respuesta
    
    Retorna un GeoJSON Feature con:
    - **geometry**: Geometría de la unidad (Point, LineString, etc.)
    - **properties.intervenciones**: Array de intervenciones en esta unidad
    - **properties.n_intervenciones**: Conteo de intervenciones
    
    ### Ejemplo de Uso
    
    ```javascript
    // Obtener unidad UNP-1978
    const response = await fetch('/unidades-proyecto/UNP-1978');
    const unidad = await response.json();
    
    console.log(unidad.properties.nombre_up);
    console.log(unidad.properties.n_intervenciones); // 1
    console.log(unidad.properties.intervenciones[0].estado); // "Terminado"
    ```
    
    ### Campos Retornados
    
    **Unidad:**
    - upid, nombre_up, direccion, barrio_vereda, comuna_corregimiento
    - tipo_equipamiento, clase_up, nombre_centro_gestor
    
    **Intervenciones (array):**
    - intervencion_id, ano, estado, tipo_intervencion
    - presupuesto_base, avance_obra, frente_activo
    - fecha_inicio, fecha_fin, referencias
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase not available")
    
    try:
        from api.scripts.unidades_proyecto import get_unidades_proyecto_geometry
        
        result = await get_unidades_proyecto_geometry({"upid": upid})
        
        if result.get("type") == "FeatureCollection":
            features = result["features"]
            if features:
                return create_utf8_response(features[0])
        
        raise HTTPException(status_code=404, detail=f"Unidad {upid} no encontrada")
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error obteniendo unidad: {str(e)}"
        )


@app.get("/intervenciones/{intervencion_id}", tags=["Unidades de Proyecto"], summary="🔵 GET | Intervención Específica")
@optional_rate_limit("60/minute")
async def get_intervencion_by_id_endpoint(
    intervencion_id: str = Path(..., description="ID de la intervención (ej: UNP-1978-0)")
):
    """
    ## 🔵 GET | Obtener Intervención Específica
    
    **Propósito**: Buscar una intervención específica dentro de todas las unidades.
    
    ### Estructura de Respuesta
    
    ```json
    {
      "unidad": {
        "upid": "UNP-1978",
        "nombre_up": "Carrera 118 Entre Calle 15 Y 16",
        "direccion": "...",
        "geometry": {...}
      },
      "intervencion": {
        "intervencion_id": "UNP-1978-0",
        "ano": 2024,
        "estado": "Terminado",
        "presupuesto_base": 55041504.84,
        "avance_obra": 100.0
      }
    }
    ```
    
    ### Ejemplo de Uso
    
    ```javascript
    const response = await fetch('/intervenciones/UNP-1978-0');
    const data = await response.json();
    
    console.log(data.unidad.nombre_up);
    console.log(data.intervencion.estado);
    ```
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase not available")
    
    try:
        from api.scripts.unidades_proyecto import get_intervencion_by_id
        
        result = await get_intervencion_by_id(intervencion_id)
        
        if not result:
            raise HTTPException(
                status_code=404,
                detail=f"Intervención {intervencion_id} no encontrada"
            )
        
        return create_utf8_response(result)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error obteniendo intervención: {str(e)}"
        )


@app.get("/intervenciones", tags=["Unidades de Proyecto"], summary="🔵 GET | Filtrar Intervenciones")
@optional_rate_limit("60/minute")
async def get_intervenciones_filtradas_endpoint(
    estado: Optional[str] = Query(None, description="Estado de la intervención"),
    tipo_intervencion: Optional[str] = Query(None, description="Tipo de intervención"),
    ano: Optional[int] = Query(None, description="Año de la intervención"),
    frente_activo: Optional[str] = Query(None, description="Estado del frente activo")
):
    """
    ## 🔵 GET | Filtrar Intervenciones
    
    **Propósito**: Filtrar intervenciones dentro de todas las unidades y retornar
    solo las unidades que tienen intervenciones que cumplen los criterios.
    
    ### Filtros Disponibles
    
    - **estado**: "En ejecución", "Terminado", "En alistamiento", etc.
    - **tipo_intervencion**: Tipo de obra o intervención
    - **ano**: Año específico (ej: 2024)
    - **frente_activo**: "Frente activo", "Inactivo", "No aplica"
    
    ### Estructura de Respuesta
    
    GeoJSON FeatureCollection donde:
    - Cada feature es una unidad que tiene intervenciones que cumplen los filtros
    - `properties.intervenciones` contiene SOLO las intervenciones filtradas
    - `properties.n_intervenciones` es el conteo de intervenciones filtradas
    
    ### Ejemplo de Uso
    
    ```javascript
    // Obtener todas las intervenciones en ejecución de 2024
    const response = await fetch('/intervenciones?estado=En ejecución&ano=2024');
    const data = await response.json();
    
    console.log(data.properties.total_intervenciones); // Total de intervenciones encontradas
    console.log(data.features.length); // Unidades con intervenciones que cumplen
    ```
    
    ### Casos de Uso
    
    - Ver todas las intervenciones activas
    - Filtrar por año para análisis temporal
    - Buscar frentes activos específicos
    - Combinar múltiples filtros para búsquedas precisas
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase not available")
    
    try:
        from api.scripts.unidades_proyecto import get_intervenciones_filtradas
        
        result = await get_intervenciones_filtradas(
            estado=estado,
            tipo_intervencion=tipo_intervencion,
            ano=ano,
            frente_activo=frente_activo
        )
        
        return create_utf8_response(result)
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error filtrando intervenciones: {str(e)}"
        )


@app.get("/frentes-activos", tags=["Unidades de Proyecto"], summary="🔵 GET | Frentes Activos")
@optional_rate_limit("60/minute")
async def get_frentes_activos_endpoint():
    """
    ## 🔵 GET | Obtener Frentes Activos
    
    **Propósito**: Retornar todas las unidades que tienen intervenciones
    con frente activo.
    
    ### Estructura de Respuesta
    
    GeoJSON FeatureCollection con:
    - **features**: Unidades con frentes activos
    - **properties.total_frentes_activos**: Conteo total de intervenciones con frente activo
    - **properties.total_unidades_con_frentes**: Número de unidades que tienen frentes activos
    
    ### Ejemplo de Uso
    
    ```javascript
    const response = await fetch('/frentes-activos');
    const data = await response.json();
    
    console.log(data.properties.total_frentes_activos); // Total de frentes activos
    console.log(data.properties.total_unidades_con_frentes); // Unidades con frentes
    
    // Renderizar en mapa con icono especial
    data.features.forEach(feature => {
      const marker = L.marker([...], {
        icon: iconFrenteActivo
      });
      marker.addTo(map);
    });
    ```
    
    ### Aplicaciones
    
    - Visualización de frentes activos en mapa
    - Dashboard de seguimiento de obras activas
    - Alertas y notificaciones sobre frentes activos
    - Reportes de avance de obra
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase not available")
    
    try:
        from api.scripts.unidades_proyecto import get_frentes_activos
        
        result = await get_frentes_activos()
        
        return create_utf8_response(result)
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error obteniendo frentes activos: {str(e)}"
        )


# ============================================================================
# ENDPOINT PARA DESCARGA DE GEOJSON
# ============================================================================

@app.get("/unidades-proyecto/download-geojson", tags=["Unidades de Proyecto"], summary="🔵 Descarga GeoJSON")
@optional_rate_limit("30/minute")  # Rate limiting para descargas pesadas
async def download_unidades_proyecto_geojson(
    request: Request,
    # Filtros de contenido
    nombre_centro_gestor: Optional[str] = Query(None, description="Centro gestor responsable"),
    tipo_intervencion: Optional[str] = Query(None, description="Tipo de intervención"),
    estado: Optional[str] = Query(None, description="Estado del proyecto"),
    upid: Optional[str] = Query(None, description="ID específico de unidad"),
    
    # Filtros geográficos
    comuna_corregimiento: Optional[str] = Query(None, description="Comuna o corregimiento específico"),
    barrio_vereda: Optional[str] = Query(None, description="Barrio o vereda específico"),
    
    # Configuración de descarga
    include_all_records: Optional[bool] = Query(True, description="Incluir todos los registros (con y sin geometría)"),
    only_with_geometry: Optional[bool] = Query(False, description="Solo registros con geometría válida"),
    limit: Optional[int] = Query(1000, ge=1, le=10000, description="Límite de registros (default: 1000 para performance)"),
    
    # Parámetros de formato
    include_metadata: Optional[bool] = Query(True, description="Incluir metadata en el GeoJSON")
):
    """
    ## 🔵 GET | 📁 Descarga | Descargar Unidades de Proyecto en formato GeoJSON
    
    **Propósito**: Descarga datos de la colección "unidades_proyecto" en formato .geojson 
    estándar para uso en aplicaciones SIG y herramientas geoespaciales.
    
    ### ✅ Características principales:
    - **Formato estándar**: GeoJSON compatible con QGIS, ArcGIS, Leaflet, etc.
    - **Filtros flexibles**: Permite filtrar por centro gestor, tipo, estado, ubicación
    - **Geometría configurable**: Opción de incluir todos los registros o solo los que tienen geometría
    - **Campos optimizados**: Incluye todos los campos relevantes para análisis SIG
    - **Encoding UTF-8**: Soporte completo para caracteres especiales en español
    
    ### 🗺️ Estrategia de geometría:
    - **include_all_records=true** (por defecto): Incluye todos los registros, los sin geometría usan coordenadas [0,0]
    - **only_with_geometry=true**: Solo registros con coordenadas válidas
    - Campo **has_valid_geometry** indica si las coordenadas son reales o placeholder
    
    ### 📊 Campos incluidos:
    - **upid**: Identificador único del proyecto
    - **nombre_up**: Nombre del proyecto
    - **estado**: Estado actual del proyecto
    - **tipo_intervencion**: Tipo de intervención urbana
    - **nombre_centro_gestor**: Entidad responsable
    - **comuna_corregimiento**: Ubicación administrativa
    - **barrio_vereda**: Ubicación específica
    - **presupuesto_base**: Valor del proyecto
    - **avance_obra**: Porcentaje de avance
    - **bpin**: Código BPIN del proyecto
    - **has_valid_geometry**: Indica si tiene coordenadas reales
    
    ### 🎯 Casos de uso:
    - **Análisis SIG**: Importar en QGIS, ArcGIS para análisis espacial
    - **Mapas web**: Cargar en Leaflet, Mapbox, OpenLayers
    - **Visualización**: Crear mapas temáticos y dashboards geográficos
    - **Integración**: Conectar con otras plataformas geoespaciales
    - **Backup**: Exportar datos para respaldo
    
    ### 📝 Ejemplo de uso:
    ```bash
    # Descargar todos los proyectos
    GET /unidades-proyecto/download-geojson
    
    # Solo proyectos de una secretaría
    GET /unidades-proyecto/download-geojson?nombre_centro_gestor=Secretaría de Infraestructura
    
    # Solo proyectos con geometría válida
    GET /unidades-proyecto/download-geojson?only_with_geometry=true
    
    # Proyectos de una comuna específica
    GET /unidades-proyecto/download-geojson?comuna_corregimiento=Comuna 1
    ```
    
    ### 💡 Nota técnica:
    - El archivo se descarga directamente como .geojson
    - Content-Type: application/geo+json
    - Encoding: UTF-8 para caracteres especiales
    - Compatible con estándares RFC 7946 (GeoJSON)
    """
    
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        # Construir filtros
        filters = {}
        
        if nombre_centro_gestor:
            filters["nombre_centro_gestor"] = nombre_centro_gestor
        if tipo_intervencion:
            filters["tipo_intervencion"] = tipo_intervencion
        if estado:
            filters["estado"] = estado
        if upid:
            filters["upid"] = upid
        if comuna_corregimiento:
            filters["comuna_corregimiento"] = comuna_corregimiento
        if barrio_vereda:
            filters["barrio_vereda"] = barrio_vereda
        if limit:
            filters["limit"] = limit
        
        # Obtener datos geoespaciales
        result = await get_unidades_proyecto_geometry(filters)
        
        # Verificar si el resultado es exitoso
        if result.get("type") != "FeatureCollection":
            if result.get("success") is False:
                raise HTTPException(
                    status_code=500,
                    detail=f"Error obteniendo datos: {result.get('error', 'Error desconocido')}"
                )
            else:
                raise HTTPException(
                    status_code=500,
                    detail="Formato de respuesta inesperado del servicio de geometrías"
                )
        
        # Extraer features
        features = result.get("features", [])
        
        # Aplicar filtro de geometría si se solicita
        if only_with_geometry and not include_all_records:
            features = [
                feature for feature in features 
                if feature.get("properties", {}).get("has_valid_geometry", False)
            ]
        
        # Crear GeoJSON final
        geojson_response = {
            "type": "FeatureCollection",
            "features": features
        }
        
        # Agregar metadata si se solicita
        if include_metadata:
            geojson_response["metadata"] = {
                "source": "unidades_proyecto collection",
                "exported_at": datetime.now().isoformat(),
                "total_features": len(features),
                "filters_applied": filters,
                "has_valid_geometry_count": len([
                    f for f in features 
                    if f.get("properties", {}).get("has_valid_geometry", False)
                ]),
                "coordinate_system": "WGS84 (EPSG:4326)",
                "format": "GeoJSON (RFC 7946)",
                "encoding": "UTF-8",
                "api_version": "1.0.0",
                "last_updated": "2025-10-28T00:00:00Z"
            }
        
        # Retornar como respuesta JSON con headers apropiados para descarga
        return JSONResponse(
            content=geojson_response,
            status_code=200,
            headers={
                "Content-Type": "application/geo+json; charset=utf-8",
                "Content-Disposition": "attachment; filename=unidades_proyecto.geojson",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando descarga GeoJSON: {str(e)}"
        )


# ============================================================================
# ENDPOINT PARA CARGAR GEOJSON A FIRESTORE
# ============================================================================

@app.get("/unidades-proyecto/download-table", tags=["Unidades de Proyecto"], summary="🔵 Descarga Tabla Excel")
@optional_rate_limit("20/minute")  # Rate limiting para descargas pesadas
async def download_unidades_proyecto_table(
    request: Request,
    # Filtros de contenido
    nombre_centro_gestor: Optional[str] = Query(None, description="Centro gestor responsable"),
    tipo_intervencion: Optional[str] = Query(None, description="Tipo de intervención"),
    estado: Optional[str] = Query(None, description="Estado del proyecto"),
    upid: Optional[str] = Query(None, description="ID específico de unidad"),
    clase_obra: Optional[str] = Query(None, description="Clase de obra"),
    tipo_equipamiento: Optional[str] = Query(None, description="Tipo de equipamiento"),
    
    # Filtros geográficos
    comuna_corregimiento: Optional[str] = Query(None, description="Comuna o corregimiento"),
    barrio_vereda: Optional[str] = Query(None, description="Barrio o vereda"),
    
    # Configuración de descarga
    limit: Optional[int] = Query(None, ge=1, le=10000, description="Límite de registros (None = todos)")
):
    """
    ## 🔵 GET | 📁 Descarga | Tabla Excel de Unidades de Proyecto
    
    **Propósito**: Descarga todos los datos de la colección "unidades_proyecto" en formato Excel (.xlsx)
    con todos los campos tabulares para análisis, reportes y gestión de proyectos.
    
    ### ✅ Características:
    - **Formato Excel**: Compatible con Microsoft Excel, Google Sheets, LibreOffice
    - **Todos los campos**: Incluye toda la información tabular de proyectos
    - **Filtros disponibles**: Por centro gestor, estado, ubicación, etc.
    - **Encoding UTF-8**: Soporte completo para caracteres especiales
    - **Headers descriptivos**: Nombres de columnas legibles
    
    ### 📊 Campos incluidos:
    - **UPID**: Identificador único
    - **Nombre UP**: Nombre del proyecto
    - **Estado**: Estado actual
    - **Tipo Intervención**: Categoría de intervención
    - **Clase Obra**: Clasificación de obra
    - **Tipo Equipamiento**: Tipo de equipamiento
    - **Centro Gestor**: Entidad responsable
    - **Comuna/Corregimiento**: Ubicación administrativa
    - **Barrio/Vereda**: Ubicación específica
    - **Dirección**: Dirección del proyecto
    - **Presupuesto Base**: Valor inicial del proyecto
    - **Presupuesto Total UP**: Presupuesto total
    - **Avance Obra**: Porcentaje de avance
    - **BPIN**: Código BPIN
    - **Año**: Año del proyecto
    - **Fuente Financiación**: Origen de recursos
    - **Referencia Contrato**: Referencias de contratos
    - **Plataforma**: Plataforma de contratación
    - **Fechas**: Fecha inicio y fin
    
    ### 🎯 Casos de uso:
    - **Reportes**: Crear informes gerenciales y ejecutivos
    - **Análisis**: Análisis de datos en Excel/Power BI
    - **Seguimiento**: Control y seguimiento de proyectos
    - **Auditoría**: Revisión y verificación de información
    - **Integración**: Importar a otros sistemas de gestión
    
    ### 📝 Ejemplos:
    ```bash
    # Descargar todos los proyectos
    GET /unidades-proyecto/download-table
    
    # Proyectos de una secretaría
    GET /unidades-proyecto/download-table?nombre_centro_gestor=Secretaría de Infraestructura
    
    # Proyectos activos de una comuna
    GET /unidades-proyecto/download-table?estado=Activo&comuna_corregimiento=COMUNA 01
    
    # Primeros 500 registros
    GET /unidades-proyecto/download-table?limit=500
    ```
    """
    
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Construir filtros
        filters = {}
        
        if nombre_centro_gestor:
            filters["nombre_centro_gestor"] = nombre_centro_gestor
        if tipo_intervencion:
            filters["tipo_intervencion"] = tipo_intervencion
        if estado:
            filters["estado"] = estado
        if upid:
            filters["upid"] = upid
        if clase_up:
            filters["clase_up"] = clase_up
        if tipo_equipamiento:
            filters["tipo_equipamiento"] = tipo_equipamiento
        if comuna_corregimiento:
            filters["comuna_corregimiento"] = comuna_corregimiento
        if barrio_vereda:
            filters["barrio_vereda"] = barrio_vereda
        if limit:
            filters["limit"] = limit
        
        # Obtener datos de atributos (sin geometría para mejor performance)
        result = await get_unidades_proyecto_attributes(filters=filters, limit=limit)
        
        # Verificar si el resultado es exitoso
        if not result.get("success"):
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo datos: {result.get('error', 'Error desconocido')}"
            )
        
        # Extraer datos
        data = result.get("data", [])
        
        if not data:
            raise HTTPException(
                status_code=404,
                detail="No se encontraron registros con los filtros especificados"
            )
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Unidades Proyecto"
        
        # Definir columnas (en orden lógico)
        columns = [
            ("upid", "UPID"),
            ("nombre_up", "Nombre UP"),
            ("nombre_up_detalle", "Nombre UP Detalle"),
            ("estado", "Estado"),
            ("tipo_intervencion", "Tipo Intervención"),
            ("clase_up", "Clase UP"),
            ("tipo_equipamiento", "Tipo Equipamiento"),
            ("nombre_centro_gestor", "Centro Gestor"),
            ("centro_gestor", "Centro Gestor (Código)"),
            ("comuna_corregimiento", "Comuna/Corregimiento"),
            ("barrio_vereda", "Barrio/Vereda"),
            ("direccion", "Dirección"),
            ("presupuesto_base", "Presupuesto Base"),
            ("presupuesto_total_up", "Presupuesto Total UP"),
            ("avance_obra", "Avance Obra (%)"),
            ("bpin", "BPIN"),
            ("ano", "Año"),
            ("fuente_financiacion", "Fuente Financiación"),
            ("referencia_contrato", "Referencia Contrato"),
            ("referencia_proceso", "Referencia Proceso"),
            ("plataforma", "Plataforma"),
            ("url_proceso", "URL Proceso"),
            ("fecha_inicio", "Fecha Inicio"),
            ("fecha_inicio_std", "Fecha Inicio Estandarizada"),
            ("fecha_fin", "Fecha Fin"),
            ("identificador", "Identificador"),
            ("cantidad", "Cantidad"),
            ("unidad_medida", "Unidad Medida"),
            ("fuera_rango", "Fuera Rango"),
            ("has_geometry", "Tiene Geometría"),
            ("created_at", "Fecha Creación"),
            ("updated_at", "Fecha Actualización"),
            ("processed_timestamp", "Timestamp Procesamiento")
        ]
        
        # Estilo del encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Escribir encabezados
        for col_idx, (field_key, field_name) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_idx, record in enumerate(data, start=2):
            for col_idx, (field_key, _) in enumerate(columns, start=1):
                value = record.get(field_key)
                
                # Formatear valores especiales
                if value is not None:
                    # Convertir listas a string separado por comas
                    if isinstance(value, list):
                        value = ", ".join(str(v) for v in value if v)
                    # Convertir booleanos a texto
                    elif isinstance(value, bool):
                        value = "Sí" if value else "No"
                    # Formatear fechas
                    elif field_key in ["created_at", "updated_at", "processed_timestamp", "fecha_inicio_std"]:
                        value = str(value) if value else ""
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_idx)
            # Ancho basado en el contenido (máximo 50)
            max_length = 15  # Ancho mínimo
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Congelar primera fila (encabezados)
        ws.freeze_panes = "A2"
        
        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"unidades_proyecto_{timestamp}.xlsx"
        
        # Retornar archivo Excel
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"❌ ERROR en download_table: {str(e)}")
        print(f"❌ TRACEBACK: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando descarga de tabla: {str(e)}"
        )


@app.get("/unidades-proyecto/download-table_by_centro_gestor", tags=["Unidades de Proyecto"], summary="🔵 Descarga Tabla Excel por Centro Gestor")
@optional_rate_limit("20/minute")  # Rate limiting para descargas pesadas
async def download_unidades_proyecto_table_by_centro_gestor(
    request: Request,
    nombre_centro_gestor: str = Query(..., description="Centro gestor responsable (requerido)"),
    
    # Filtros adicionales opcionales
    tipo_intervencion: Optional[str] = Query(None, description="Tipo de intervención"),
    estado: Optional[str] = Query(None, description="Estado del proyecto"),
    upid: Optional[str] = Query(None, description="ID específico de unidad"),
    clase_obra: Optional[str] = Query(None, description="Clase de obra"),
    tipo_equipamiento: Optional[str] = Query(None, description="Tipo de equipamiento"),
    
    # Filtros geográficos
    comuna_corregimiento: Optional[str] = Query(None, description="Comuna o corregimiento"),
    barrio_vereda: Optional[str] = Query(None, description="Barrio o vereda"),
    
    # Configuración de descarga
    limit: Optional[int] = Query(None, ge=1, le=10000, description="Límite de registros (None = todos)")
):
    """
    ## 🔵 GET | 📁 Descarga | Tabla Excel de Unidades de Proyecto por Centro Gestor
    
    **Propósito**: Descarga datos de la colección "unidades_proyecto" filtrados por "nombre_centro_gestor"
    en formato Excel (.xlsx) con todos los campos tabulares para análisis y reportes específicos por entidad.
    
    ### ✅ Características:
    - **Filtro obligatorio**: Requiere especificar el centro gestor
    - **Formato Excel**: Compatible con Microsoft Excel, Google Sheets, LibreOffice
    - **Todos los campos**: Incluye toda la información tabular de proyectos
    - **Filtros adicionales**: Combinar con otros filtros (estado, ubicación, etc.)
    - **Encoding UTF-8**: Soporte completo para caracteres especiales
    - **Headers descriptivos**: Nombres de columnas legibles
    
    ### 📊 Campos incluidos:
    - **UPID**: Identificador único
    - **Nombre UP**: Nombre del proyecto
    - **Estado**: Estado actual
    - **Tipo Intervención**: Categoría de intervención
    - **Clase Obra**: Clasificación de obra
    - **Tipo Equipamiento**: Tipo de equipamiento
    - **Centro Gestor**: Entidad responsable
    - **Comuna/Corregimiento**: Ubicación administrativa
    - **Barrio/Vereda**: Ubicación específica
    - **Dirección**: Dirección del proyecto
    - **Presupuesto Base**: Valor inicial del proyecto
    - **Presupuesto Total UP**: Presupuesto total
    - **Avance Obra**: Porcentaje de avance
    - **BPIN**: Código BPIN
    - **Año**: Año del proyecto
    - **Fuente Financiación**: Origen de recursos
    - **Referencia Contrato**: Referencias de contratos
    - **Plataforma**: Plataforma de contratación
    - **Fechas**: Fecha inicio y fin
    
    ### 🎯 Casos de uso:
    - **Reportes por entidad**: Informes específicos por secretaría o entidad
    - **Seguimiento sectorial**: Control de proyectos por sector
    - **Análisis comparativo**: Comparar gestión entre diferentes centros gestores
    - **Auditoría específica**: Revisión de proyectos de una entidad particular
    - **Informes gerenciales**: Reportes ejecutivos por dependencia
    
    ### 📝 Ejemplos:
    ```bash
    # Descargar todos los proyectos de una secretaría
    GET /unidades-proyecto/download-table_by_centro_gestor?nombre_centro_gestor=Secretaría de Infraestructura
    
    # Proyectos activos de una secretaría
    GET /unidades-proyecto/download-table_by_centro_gestor?nombre_centro_gestor=Secretaría de Educación&estado=Activo
    
    # Proyectos de una secretaría en una comuna específica
    GET /unidades-proyecto/download-table_by_centro_gestor?nombre_centro_gestor=Secretaría de Salud&comuna_corregimiento=COMUNA 01
    
    # Primeros 100 registros de una secretaría
    GET /unidades-proyecto/download-table_by_centro_gestor?nombre_centro_gestor=Secretaría de Hacienda&limit=100
    ```
    """
    
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Construir filtros (nombre_centro_gestor es obligatorio)
        filters = {
            "nombre_centro_gestor": nombre_centro_gestor
        }
        
        # Agregar filtros opcionales
        if tipo_intervencion:
            filters["tipo_intervencion"] = tipo_intervencion
        if estado:
            filters["estado"] = estado
        if upid:
            filters["upid"] = upid
        if clase_up:
            filters["clase_up"] = clase_up
        if tipo_equipamiento:
            filters["tipo_equipamiento"] = tipo_equipamiento
        if comuna_corregimiento:
            filters["comuna_corregimiento"] = comuna_corregimiento
        if barrio_vereda:
            filters["barrio_vereda"] = barrio_vereda
        if limit:
            filters["limit"] = limit
        
        # Obtener datos de atributos (sin geometría para mejor performance)
        result = await get_unidades_proyecto_attributes(filters=filters, limit=limit)
        
        # Verificar si el resultado es exitoso
        if not result.get("success"):
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo datos: {result.get('error', 'Error desconocido')}"
            )
        
        # Extraer datos
        data = result.get("data", [])
        
        if not data:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontraron registros para el centro gestor '{nombre_centro_gestor}' con los filtros especificados"
            )
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Unidades Proyecto"
        
        # Definir columnas (en orden lógico)
        columns = [
            ("upid", "UPID"),
            ("nombre_up", "Nombre UP"),
            ("nombre_up_detalle", "Nombre UP Detalle"),
            ("estado", "Estado"),
            ("tipo_intervencion", "Tipo Intervención"),
            ("clase_up", "Clase UP"),
            ("tipo_equipamiento", "Tipo Equipamiento"),
            ("nombre_centro_gestor", "Centro Gestor"),
            ("centro_gestor", "Centro Gestor (Código)"),
            ("comuna_corregimiento", "Comuna/Corregimiento"),
            ("barrio_vereda", "Barrio/Vereda"),
            ("direccion", "Dirección"),
            ("presupuesto_base", "Presupuesto Base"),
            ("presupuesto_total_up", "Presupuesto Total UP"),
            ("avance_obra", "Avance Obra (%)"),
            ("bpin", "BPIN"),
            ("ano", "Año"),
            ("fuente_financiacion", "Fuente Financiación"),
            ("referencia_contrato", "Referencia Contrato"),
            ("referencia_proceso", "Referencia Proceso"),
            ("plataforma", "Plataforma"),
            ("url_proceso", "URL Proceso"),
            ("fecha_inicio", "Fecha Inicio"),
            ("fecha_inicio_std", "Fecha Inicio Estandarizada"),
            ("fecha_fin", "Fecha Fin"),
            ("identificador", "Identificador"),
            ("cantidad", "Cantidad"),
            ("unidad_medida", "Unidad Medida"),
            ("fuera_rango", "Fuera Rango"),
            ("has_geometry", "Tiene Geometría"),
            ("created_at", "Fecha Creación"),
            ("updated_at", "Fecha Actualización"),
            ("processed_timestamp", "Timestamp Procesamiento")
        ]
        
        # Estilo del encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Escribir encabezados
        for col_idx, (field_key, field_name) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_idx, record in enumerate(data, start=2):
            for col_idx, (field_key, _) in enumerate(columns, start=1):
                value = record.get(field_key)
                
                # Formatear valores especiales
                if value is not None:
                    # Convertir listas a string separado por comas
                    if isinstance(value, list):
                        value = ", ".join(str(v) for v in value if v)
                    # Convertir booleanos a texto
                    elif isinstance(value, bool):
                        value = "Sí" if value else "No"
                    # Formatear fechas
                    elif field_key in ["created_at", "updated_at", "processed_timestamp", "fecha_inicio_std"]:
                        value = str(value) if value else ""
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_idx)
            # Ancho basado en el contenido (máximo 50)
            max_length = 15  # Ancho mínimo
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Congelar primera fila (encabezados)
        ws.freeze_panes = "A2"
        
        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generar nombre de archivo con timestamp y nombre del centro gestor
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Limpiar nombre del centro gestor para usarlo en el nombre del archivo
        centro_gestor_safe = nombre_centro_gestor.replace(" ", "_").replace("/", "-")
        filename = f"unidades_proyecto_{centro_gestor_safe}_{timestamp}.xlsx"
        
        # Retornar archivo Excel
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"❌ ERROR en download_table_by_centro_gestor: {str(e)}")
        print(f"❌ TRACEBACK: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando descarga de tabla por centro gestor: {str(e)}"
        )


@app.post("/unidades-proyecto/cargar-geojson", tags=["Unidades de Proyecto"], summary="🟢 Cargar GeoJSON a Firestore (UPSERT)")
async def cargar_geojson_a_firestore(
    geojson_file: UploadFile = File(..., description="Archivo GeoJSON con unidades de proyecto"),
    batch_size: int = Query(500, ge=1, le=500, description="Tamaño de lote para operaciones batch"),
    override_existing: bool = Query(False, description="[DEPRECADO] Parámetro legacy, ahora siempre hace UPSERT"),
    override_upid: bool = Query(False, description="Generar nuevos UPIDs aunque existan"),
    dry_run: bool = Query(False, description="Simular carga sin escribir en Firebase")
):
    """
    ## 🟢 POST | 📤 UPSERT | Importar/Actualizar Unidades de Proyecto desde GeoJSON a Firestore
    
    **Propósito**: Cargar o actualizar masivamente datos de unidades de proyecto desde un archivo GeoJSON 
    a la colección 'unidades_proyecto' en Firebase Firestore usando estrategia **UPSERT**.
    
    ### ✅ Características principales:
    - **🔄 UPSERT automático**: Si el documento existe, actualiza solo los campos modificados. Si no existe, lo crea.
    - **Importación masiva**: Procesa múltiples features en un solo archivo
    - **Validación automática**: Verifica estructura GeoJSON y campos requeridos
    - **Procesamiento por lotes**: Optimizado para grandes volúmenes (hasta 500 por batch)
    - **Generación de UPIDs consecutivos**: Mantiene el consecutivo UNP-{número}
    - **Campo automático**: Agrega `tipo_equipamiento: "Vías"` a todos los elementos
    - **Modo dry-run**: Simula la carga para validar datos sin escribir en BD
    
    ### 🔄 Comportamiento UPSERT:
    - **Si el documento existe**: Actualiza solo los campos que cambiaron (merge)
    - **Si el documento NO existe**: Crea un nuevo documento completo
    - **Beneficios**: 
      - No duplica datos
      - Preserva campos que no están en el GeoJSON
      - Actualiza solo lo necesario
      - Más eficiente que sobrescribir completo
    
    ### 📋 Estructura esperada del GeoJSON:
    ```json
    {
      "type": "FeatureCollection",
      "features": [
        {
          "type": "Feature",
          "geometry": {
            "type": "LineString|Point|Polygon",
            "coordinates": [[lng, lat], ...]
          },
          "properties": {
            "nombre_up": "Nombre del proyecto",
            "estado": "Finalizado|En Ejecución|etc.",
            "clase_up": "Obra Vial|etc.",
            "comuna_corregimiento": "COMUNA XX",
            "barrio_vereda": "Nombre del barrio",
            "presupuesto_base": "123456.78",
            "avance_obra": "100",
            "ano": "2024",
            "nombre_centro_gestor": "Secretaría de...",
            "bpin": "2023760010180",
            ...otros campos opcionales
          }
        }
      ]
    }
    ```
    
    ### 🔧 Parámetros de configuración:
    - **batch_size** (1-500): Número de documentos por lote (default: 500)
    - **override_existing**: [DEPRECADO] Ya no se usa, siempre hace UPSERT
    - **override_upid**: 
      - `false` (default): Usa UPIDs del GeoJSON si existen, genera consecutivos si no
      - `true`: Genera nuevos UPIDs consecutivos para todos
    - **dry_run**: 
      - `false` (default): Ejecuta la carga/actualización real
      - `true`: Solo simula y muestra estadísticas
    
    ### 📊 Procesamiento automático:
    - **UPID**: Genera `UNP-{número}` consecutivo si no existe
    - **tipo_equipamiento**: Agrega automáticamente valor "Vías"
    - **Geometría**: Detecta tipo (Point, LineString, Polygon, Multi*) y serializa como JSON string
    - **Validación de coordenadas**: Identifica coordenadas válidas vs placeholders [0,0]
    - **Conversión de tipos**: 
      - `presupuesto_base` → float
      - `avance_obra` → float (porcentaje)
      - `cantidad` → int
      - `bpin` → string limpia (sin prefijos '-')
    - **Limpieza de datos**: Elimina valores null, NaN, vacíos
    - **Timestamps**: Agrega `updated_at` y `loaded_at` automáticamente
    
    ### 📈 Respuesta incluye:
    - **Estadísticas detalladas**:
      - Total de features procesados
      - Documentos **creados** (nuevos)
      - Documentos **actualizados** (existentes modificados)
      - Documentos omitidos (solo en dry-run)
      - Errores encontrados
    - **Detalles de errores**: Lista de features que fallaron con razón
    - **Tasa de éxito**: Porcentaje de procesamiento exitoso
    
    ### 🎯 Casos de uso:
    - **Migración inicial**: Cargar datos históricos desde sistemas SIG
    - **Actualización masiva**: Importar nuevos proyectos desde herramientas externas
    - **Sincronización**: Mantener datos actualizados desde fuentes GeoJSON
    - **Backup/Restore**: Restaurar datos desde respaldos
    - **Integración**: Importar desde QGIS, ArcGIS, u otras plataformas SIG
    
    ### ⚠️ Consideraciones:
    - El archivo debe ser GeoJSON válido (RFC 7946)
    - Máximo 500 documentos por batch (limitación de Firestore)
    - Los UPIDs deben ser únicos en toda la colección
    - Para archivos muy grandes (>1000 features), considerar múltiples cargas
    - En modo dry-run, no se valida duplicidad de UPIDs
    
    ### 📝 Ejemplo de respuesta exitosa:
    ```json
    {
      "success": true,
      "message": "Carga completada: 646/646 features procesados",
      "stats": {
        "total_features": 646,
        "processed": 646,
        "created": 500,
        "updated": 0,
        "skipped": 146,
        "errors": 0,
        "error_details": []
      },
      "dry_run": false
    }
    ```
    """
    
    # Verificar disponibilidad de Firebase
    if not FIREBASE_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="Firebase no está disponible en este momento"
        )
    
    # Validar tipo de archivo
    if not geojson_file.filename.lower().endswith('.geojson') and not geojson_file.filename.lower().endswith('.json'):
        raise HTTPException(
            status_code=400,
            detail="Solo se permiten archivos .geojson o .json"
        )
    
    try:
        # Leer contenido del archivo
        print(f"📁 Leyendo archivo: {geojson_file.filename}")
        geojson_content = await geojson_file.read()
        
        # Decodificar como UTF-8
        try:
            geojson_text = geojson_content.decode('utf-8')
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=400,
                detail="El archivo debe estar codificado en UTF-8"
            )
        
        # Parsear JSON
        try:
            geojson_data = json.loads(geojson_text)
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Error parseando JSON: {str(e)}"
            )
        
        # Importar función de carga
        try:
            from api.scripts.unidades_proyecto_loader import load_geojson_to_firestore
        except ImportError as e:
            raise HTTPException(
                status_code=500,
                detail=f"Error importando módulo de carga: {str(e)}"
            )
        
        # Ejecutar carga
        print(f"🚀 Iniciando carga de GeoJSON...")
        print(f"   - Archivo: {geojson_file.filename}")
        print(f"   - Tamaño: {len(geojson_content)} bytes")
        print(f"   - Batch size: {batch_size}")
        print(f"   - Override existing: {override_existing}")
        print(f"   - Override UPID: {override_upid}")
        print(f"   - Dry run: {dry_run}")
        
        result = await load_geojson_to_firestore(
            geojson_data=geojson_data,
            batch_size=batch_size,
            override_existing=override_existing,
            override_upid=override_upid,
            dry_run=dry_run
        )
        
        if not result.get('success'):
            raise HTTPException(
                status_code=400,
                detail=result.get('error', 'Error desconocido durante la carga')
            )
        
        # Preparar respuesta
        response_data = {
            "success": True,
            "message": result.get('message'),
            "stats": result.get('stats', {}),
            "dry_run": dry_run,
            "file_info": {
                "filename": geojson_file.filename,
                "size_bytes": len(geojson_content),
                "processed_at": datetime.now().isoformat()
            }
        }
        
        # Agregar advertencias si hay
        if result.get('stats', {}).get('errors', 0) > 0:
            response_data['warnings'] = {
                "message": "Algunos features no pudieron ser procesados",
                "error_count": result['stats']['errors'],
                "error_details": result['stats'].get('error_details', [])[:10]  # Limitar a 10 primeros errores
            }
        
        return create_utf8_response(response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"❌ ERROR CRÍTICO: {str(e)}")
        print(traceback.format_exc())
        
        raise HTTPException(
            status_code=500,
            detail=f"Error interno procesando archivo GeoJSON: {str(e)}"
        )


@app.delete("/unidades-proyecto/delete-by-centro-gestor", tags=["Unidades de Proyecto"], summary="🔴 Eliminar por Centro Gestor")
async def delete_unidades_by_centro_gestor(
    nombre_centro_gestor: str = Query(..., description="Nombre del centro gestor cuyos proyectos serán eliminados"),
    confirm: bool = Query(False, description="Debe ser true para confirmar la eliminación")
):
    """
    ## 🔴 DELETE | Eliminar Unidades de Proyecto por Centro Gestor
    
    **Propósito**: Eliminar todos los documentos de la colección 'unidades_proyecto' que 
    correspondan a un centro gestor específico.
    
    ### ⚠️ ADVERTENCIA
    Esta operación es **IRREVERSIBLE**. Todos los documentos que coincidan con el filtro 
    serán eliminados permanentemente de Firebase.
    
    ### 🔧 Parámetros:
    - **nombre_centro_gestor** (requerido): Nombre exacto del centro gestor
    - **confirm** (requerido): Debe ser `true` para ejecutar la eliminación
    
    ### 📊 Proceso:
    1. Busca todos los documentos con el `nombre_centro_gestor` especificado
    2. Cuenta cuántos documentos serán eliminados
    3. Si `confirm=true`, elimina los documentos en batches de 500
    4. Retorna estadísticas de la operación
    
    ### 📝 Ejemplo de uso:
    ```
    DELETE /unidades-proyecto/delete-by-centro-gestor?nombre_centro_gestor=Secretaría de Infraestructura&confirm=true
    ```
    
    ### 📈 Respuesta exitosa:
    ```json
    {
      "success": true,
      "message": "15 documentos eliminados correctamente",
      "stats": {
        "deleted_count": 15,
        "nombre_centro_gestor": "Secretaría de Infraestructura"
      }
    }
    ```
    
    ### ⚠️ Seguridad:
    - Requiere `confirm=true` para ejecutar
    - Sin `confirm=true`, solo muestra cuántos documentos serían eliminados
    """
    
    # Verificar disponibilidad de Firebase
    if not FIREBASE_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="Firebase no está disponible en este momento"
        )
    
    try:
        from database.firebase_config import get_firestore_client
        
        db = get_firestore_client()
        if db is None:
            raise HTTPException(
                status_code=503,
                detail="No se pudo conectar a Firestore"
            )
        
        collection_ref = db.collection('unidades_proyecto')
        
        # Buscar documentos que coincidan con el filtro
        print(f"🔍 Buscando documentos con nombre_centro_gestor='{nombre_centro_gestor}'...")
        query = collection_ref.where('nombre_centro_gestor', '==', nombre_centro_gestor)
        docs = list(query.stream())
        
        total_docs = len(docs)
        
        if total_docs == 0:
            return create_utf8_response({
                "success": False,
                "message": f"No se encontraron documentos con nombre_centro_gestor='{nombre_centro_gestor}'",
                "stats": {
                    "deleted_count": 0,
                    "nombre_centro_gestor": nombre_centro_gestor
                }
            })
        
        # Si no hay confirmación, solo reportar cuántos se eliminarían
        if not confirm:
            return create_utf8_response({
                "success": False,
                "message": f"Se encontraron {total_docs} documentos. Use confirm=true para eliminarlos.",
                "warning": "La eliminación no se ejecutó porque confirm=false",
                "stats": {
                    "found_count": total_docs,
                    "nombre_centro_gestor": nombre_centro_gestor
                }
            })
        
        # Eliminar en batches de 500 (límite de Firestore)
        print(f"🗑️  Eliminando {total_docs} documentos...")
        batch_size = 500
        deleted_count = 0
        
        for i in range(0, total_docs, batch_size):
            batch = db.batch()
            batch_docs = docs[i:i + batch_size]
            
            for doc in batch_docs:
                batch.delete(doc.reference)
            
            batch.commit()
            deleted_count += len(batch_docs)
            print(f"   Eliminados {deleted_count}/{total_docs} documentos...")
        
        print(f"✅ Eliminación completada: {deleted_count} documentos")
        
        return create_utf8_response({
            "success": True,
            "message": f"{deleted_count} documentos eliminados correctamente",
            "stats": {
                "deleted_count": deleted_count,
                "nombre_centro_gestor": nombre_centro_gestor
            }
        })
        
    except Exception as e:
        import traceback
        print(f"❌ ERROR: {str(e)}")
        print(traceback.format_exc())
        
        raise HTTPException(
            status_code=500,
            detail=f"Error eliminando documentos: {str(e)}"
        )


@app.delete("/unidades-proyecto/delete-by-tipo-equipamiento", tags=["Unidades de Proyecto"], summary="🔴 Eliminar por Tipo de Equipamiento")
async def delete_unidades_by_tipo_equipamiento(
    tipo_equipamiento: str = Query(..., description="Tipo de equipamiento cuyos proyectos serán eliminados"),
    confirm: bool = Query(False, description="Debe ser true para confirmar la eliminación")
):
    """
    ## 🔴 DELETE | Eliminar Unidades de Proyecto por Tipo de Equipamiento
    
    **Propósito**: Eliminar todos los documentos de la colección 'unidades_proyecto' que 
    correspondan a un tipo de equipamiento específico.
    
    ### ⚠️ ADVERTENCIA
    Esta operación es **IRREVERSIBLE**. Todos los documentos que coincidan con el filtro 
    serán eliminados permanentemente de Firebase.
    
    ### 🔧 Parámetros:
    - **tipo_equipamiento** (requerido): Tipo de equipamiento exacto (ej: "Vías", "Parques y zonas verdes")
    - **confirm** (requerido): Debe ser `true` para ejecutar la eliminación
    
    ### 📊 Proceso:
    1. Busca todos los documentos con el `tipo_equipamiento` especificado
    2. Cuenta cuántos documentos serán eliminados
    3. Si `confirm=true`, elimina los documentos en batches de 500
    4. Retorna estadísticas de la operación
    
    ### 📝 Ejemplo de uso:
    ```
    DELETE /unidades-proyecto/delete-by-tipo-equipamiento?tipo_equipamiento=Vías&confirm=true
    ```
    
    ### 📈 Respuesta exitosa:
    ```json
    {
      "success": true,
      "message": "369 documentos eliminados correctamente",
      "stats": {
        "deleted_count": 369,
        "tipo_equipamiento": "Vías"
      }
    }
    ```
    
    ### ⚠️ Seguridad:
    - Requiere `confirm=true` para ejecutar
    - Sin `confirm=true`, solo muestra cuántos documentos serían eliminados
    """
    
    # Verificar disponibilidad de Firebase
    if not FIREBASE_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="Firebase no está disponible en este momento"
        )
    
    try:
        from database.firebase_config import get_firestore_client
        
        db = get_firestore_client()
        if db is None:
            raise HTTPException(
                status_code=503,
                detail="No se pudo conectar a Firestore"
            )
        
        collection_ref = db.collection('unidades_proyecto')
        
        # Buscar documentos que coincidan con el filtro
        print(f"🔍 Buscando documentos con tipo_equipamiento='{tipo_equipamiento}'...")
        query = collection_ref.where('tipo_equipamiento', '==', tipo_equipamiento)
        docs = list(query.stream())
        
        total_docs = len(docs)
        
        if total_docs == 0:
            return create_utf8_response({
                "success": False,
                "message": f"No se encontraron documentos con tipo_equipamiento='{tipo_equipamiento}'",
                "stats": {
                    "deleted_count": 0,
                    "tipo_equipamiento": tipo_equipamiento
                }
            })
        
        # Si no hay confirmación, solo reportar cuántos se eliminarían
        if not confirm:
            return create_utf8_response({
                "success": False,
                "message": f"Se encontraron {total_docs} documentos. Use confirm=true para eliminarlos.",
                "warning": "La eliminación no se ejecutó porque confirm=false",
                "stats": {
                    "found_count": total_docs,
                    "tipo_equipamiento": tipo_equipamiento
                }
            })
        
        # Eliminar en batches de 500 (límite de Firestore)
        print(f"🗑️  Eliminando {total_docs} documentos...")
        batch_size = 500
        deleted_count = 0
        
        for i in range(0, total_docs, batch_size):
            batch = db.batch()
            batch_docs = docs[i:i + batch_size]
            
            for doc in batch_docs:
                batch.delete(doc.reference)
            
            batch.commit()
            deleted_count += len(batch_docs)
            print(f"   Eliminados {deleted_count}/{total_docs} documentos...")
        
        print(f"✅ Eliminación completada: {deleted_count} documentos")
        
        return create_utf8_response({
            "success": True,
            "message": f"{deleted_count} documentos eliminados correctamente",
            "stats": {
                "deleted_count": deleted_count,
                "tipo_equipamiento": tipo_equipamiento
            }
        })
        
    except Exception as e:
        import traceback
        print(f"❌ ERROR: {str(e)}")
        print(traceback.format_exc())
        
        raise HTTPException(
            status_code=500,
            detail=f"Error eliminando documentos: {str(e)}"
        )


# ============================================================================
# ENDPOINTS DE INTEROPERABILIDAD CON ARTEFACTO DE SEGUIMIENTO
# ============================================================================

@app.get("/contratos/init_contratos_seguimiento", tags=["Interoperabilidad con Artefacto de Seguimiento"])
@async_cache(ttl_seconds=300)  # Cache de 5 minutos para contratos
async def init_contratos_seguimiento(
    referencia_contrato: Optional[str] = Query(None, description="Referencia del contrato (búsqueda parcial)"),
    nombre_centro_gestor: Optional[str] = Query(None, description="Centro gestor responsable (exacto)")
):
    """
    ## Inicialización de Contratos para Seguimiento
    
    Obtiene datos combinados desde las colecciones `contratos_emprestito`, `ordenes_compra_emprestito` 
    y `convenios_transferencias_emprestito` con filtros optimizados.
    
    **Colecciones incluidas**:
    - `contratos_emprestito`: Contratos de empréstito
    - `ordenes_compra_emprestito`: Órdenes de compra TVEC
    - `convenios_transferencias_emprestito`: Convenios de transferencia
    
    **Campos retornados**: bpin, banco, nombre_centro_gestor, estado_contrato, referencia_contrato, 
    referencia_proceso, nombre_resumido_proceso, objeto_contrato, modalidad_contratacion, fecha_inicio_contrato, fecha_firma, 
    fecha_fin_contrato, _source (indica la colección de origen)
    
    **Filtros**:
    - `referencia_contrato`: Textbox - búsqueda parcial
    - `nombre_centro_gestor`: Dropdown - selección exacta
    
    Sin filtros retorna todos los datos disponibles de las tres colecciones.
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        return {"success": False, "error": "Firebase no disponible", "data": [], "count": 0}
    
    try:
        filters = {}
        if referencia_contrato:
            filters["referencia_contrato"] = referencia_contrato
        if nombre_centro_gestor:
            filters["nombre_centro_gestor"] = nombre_centro_gestor
        
        result = await get_contratos_init_data(filters)
        
        if not result["success"]:
            raise HTTPException(status_code=500, detail=result.get('error', 'Error obteniendo contratos'))
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando contratos: {str(e)}")

@app.post("/reportes_contratos/", tags=["Interoperabilidad con Artefacto de Seguimiento"])
async def crear_reporte_contrato(
    # Información básica del reporte
    referencia_contrato: str = Form(..., min_length=1, description="Referencia del contrato"),
    observaciones: str = Form(..., description="Observaciones del reporte"),
    
    # Avances del proyecto (soporte para decimales)
    avance_fisico: float = Form(..., ge=0, le=100, description="Porcentaje de avance físico (0-100, decimales permitidos)"),
    avance_financiero: float = Form(..., ge=0, le=100, description="Porcentaje de avance financiero (0-100, decimales permitidos)"),
    
    # Información de alertas
    alertas_descripcion: str = Form(..., description="Descripción de la alerta"),
    alertas_es_alerta: bool = Form(..., description="Indica si es una alerta activa"),
    alertas_tipo_alerta: str = Form(default="", description="Tipos de alerta separados por coma"),
    
    # Archivos de evidencia (carga real de archivos)
    archivos_evidencia: List[UploadFile] = File(..., description="Archivos de evidencia (PDF, DOC, DOCX, XLS, XLSX, TXT, CSV, JPG, PNG, GIF)")
):
    """
    ## 📊 Crear Reporte de Contrato con Evidencias y Upload de Archivos
    
    **Propósito**: Endpoint unificado para crear reportes de seguimiento de contratos 
    con carga de archivos y estructura de carpetas organizada.
    
    ### ✅ IMPORTANTE - Google Drive:
    - **Estado actual**: PRODUCCIÓN - Subida real de archivos funcionando
    - **Configuración**: Google Drive API con Service Account y Shared Drive
    - **Archivos**: Se suben realmente y son accesibles desde Google Drive
    
    ### ✅ Características principales:
    - **Carga de archivos**: Upload directo de archivos de evidencia
    - **Estructura automática**: Carpetas organizadas por contrato y fecha  
    - **Firebase**: Almacenamiento en colección `reportes_contratos`
    - **Timestamp automático**: Fecha de reporte generada automáticamente
    - **Decimales**: Soporte para avances con decimales (ej: 75.5)
    
    ### 📋 Parámetros (Form Data):
    - **referencia_contrato**: Referencia del contrato (obligatorio)
    - **observaciones**: Descripción detallada del avance (obligatorio)
    - **avance_fisico**: Porcentaje de avance físico 0-100 con decimales (obligatorio)
    - **avance_financiero**: Porcentaje de avance financiero 0-100 con decimales (obligatorio)
    - **alertas_descripcion**: Descripción de la alerta (obligatorio)
    - **alertas_es_alerta**: Booleano si es alerta activa (obligatorio)
    - **alertas_tipo_alerta**: Tipos de alerta separados por coma (opcional)
    - **archivos_evidencia**: Archivos de evidencia para subir (obligatorio, múltiples archivos)
    
    ### 📁 Estructura de carpetas en Google Drive:
    ```
    📁 CONTRATOS_REPORTES/
      📁 {referencia_contrato}/
        📁 REPORTE_{YYYY-MM-DD}_{HH-MM-SS}_{UUID}/
          📄 evidencia1.pdf
          📄 evidencia2.jpg
          📄 ...
    ```
    
    ### 🔒 Validaciones aplicadas:
    - **Archivos**: Tipos permitidos (PDF, DOC, DOCX, XLS, XLSX, JPG, PNG, GIF)
    - **Tamaño**: Máximo 10MB por archivo
    - **Cantidad**: Al menos 1 archivo requerido
    - **Avances**: Rango 0-100 con decimales (ej: 75.5)
    - **Nombres**: Caracteres especiales manejados automáticamente
    
    ### 🚀 Proceso automático:
    1. Validar archivos subidos
    2. Crear/verificar carpeta del contrato
    3. Crear carpeta única para este reporte
    4. Subir archivos a Google Drive
    5. Guardar metadata en Firebase con timestamp actual
    6. Retornar URLs y confirmación
    
    ### � Ejemplo de uso con HTML Form:
    ```html
    <form method="POST" enctype="multipart/form-data">
        <input name="referencia_contrato" value="CONTRATO-2025-001" required>
        <textarea name="observaciones" required>Avance del proyecto...</textarea>
        <input name="avance_fisico" type="number" step="0.1" min="0" max="100" required>
        <input name="avance_financiero" type="number" step="0.1" min="0" max="100" required>
        <textarea name="alertas_descripcion" required>Descripción de alerta...</textarea>
        <input name="alertas_es_alerta" type="checkbox">
        <input name="alertas_tipo_alerta" value="logistica,cronograma">
        <input name="archivos_evidencia" type="file" multiple accept=".pdf,.doc,.docx,.xls,.xlsx,.jpg,.png,.gif" required>
        <button type="submit">Crear Reporte</button>
    </form>
    ```
    """
    # Verificar disponibilidad de servicios
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(
            status_code=503, 
            detail="Servicios no disponibles: Firebase o scripts requeridos"
        )
    
    if not REPORTES_CONTRATOS_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="Operaciones de reportes de contratos no disponibles"
        )
    
    try:
        # Validar archivos subidos
        if not archivos_evidencia:
            raise HTTPException(
                status_code=400,
                detail="Se requiere al menos un archivo de evidencia"
            )
        
        # Validar cada archivo subido
        archivos_validados = []
        tipos_permitidos = {
            'application/pdf': '.pdf',
            'application/msword': '.doc',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
            'application/vnd.ms-excel': '.xls',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
            'text/plain': '.txt',
            'text/csv': '.csv',
            'image/jpeg': '.jpg',
            'image/png': '.png',
            'image/gif': '.gif'
        }
        
        max_size = 10 * 1024 * 1024  # 10MB
        
        for archivo in archivos_evidencia:
            # Validar tamaño
            if archivo.size > max_size:
                raise HTTPException(
                    status_code=400,
                    detail=f"Archivo {archivo.filename} excede el tamaño máximo de 10MB"
                )
            
            # Validar tipo de archivo
            if archivo.content_type not in tipos_permitidos:
                raise HTTPException(
                    status_code=400,
                    detail=f"Tipo de archivo no permitido: {archivo.content_type} para {archivo.filename}"
                )
            
            # Leer contenido del archivo
            contenido = await archivo.read()
            await archivo.seek(0)  # Reset para lectura posterior si es necesario
            
            archivo_info = {
                "filename": archivo.filename,
                "content_type": archivo.content_type,
                "size": archivo.size,
                "content": contenido
            }
            archivos_validados.append(archivo_info)
        
        # Construir datos optimizados para Firebase
        reporte_dict = {
            "referencia_contrato": referencia_contrato.strip(),
            "observaciones": observaciones.strip(),
            "avance_fisico": float(avance_fisico),
            "avance_financiero": float(avance_financiero),
            "alertas": {
                "descripcion": alertas_descripcion.strip(),
                "es_alerta": alertas_es_alerta,
                "tipos": [tipo.strip() for tipo in alertas_tipo_alerta.split(",") if tipo.strip()] if alertas_tipo_alerta else []
            },
            "archivos_evidencia": archivos_validados
        }
        
        # Crear el reporte usando la función del script
        result = await create_reporte_contrato(reporte_dict)
        
        if not result["success"]:
            raise HTTPException(
                status_code=400,
                detail=f"Error creando reporte: {result.get('error', 'Error desconocido')}"
            )
        
        # Respuesta optimizada sin redundancias
        response_data = {
            "success": True,
            "message": result["message"],
            "doc_id": result["doc_id"],
            "url_carpeta_drive": result["url_carpeta_drive"],
            "archivos_count": len(archivos_validados)
        }
        

        
        return create_utf8_response(response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint crear_reporte_contrato: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor: {str(e)}"
        )

@app.get("/reportes_contratos/", tags=["Interoperabilidad con Artefacto de Seguimiento"])
async def obtener_reportes_contratos(request: Request):
    """
    ## 📋 Obtener Todos los Reportes de Contratos
    
    **Propósito**: Obtener listado completo de todos los reportes de contratos almacenados en Firebase.
    Muestra todos los registros de la colección `reportes_contratos` con `nombre_centro_gestor` 
    actualizado desde las colecciones de empréstito cuando sea necesario.
    
    ### 🔄 Integración con colecciones de empréstito:
    - Si un reporte no tiene `nombre_centro_gestor` o está vacío, se busca automáticamente 
      en las colecciones `contratos_emprestito`, `ordenes_compra_emprestito` y 
      `convenios_transferencias_emprestito` usando `referencia_contrato` como clave
    - Los reportes actualizados incluyen el campo `nombre_centro_gestor_source` indicando la colección de origen
    
    ### 📊 Ordenamiento:
    Los resultados se ordenan por `fecha_reporte` descendente (más recientes primero).
    
    ### 💡 Casos de uso:
    - Obtener listado completo para dashboard de seguimiento
    - Vista general de todos los reportes generados con datos completos
    - Administración y auditoría de reportes con información del centro gestor
    """
    # Verificar disponibilidad de servicios
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE or not REPORTES_CONTRATOS_AVAILABLE:
        return {
            "success": False, 
            "error": "Servicios no disponibles", 
            "data": [], 
            "count": 0
        }
    
    try:
        # Obtener todos los reportes (sin filtros)
        result = await get_reportes_contratos(None)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo reportes: {result.get('error', 'Error desconocido')}"
            )
        
        # Forzar respuesta sin compresión para evitar conflictos
        response = JSONResponse(
            content=result,
            status_code=200,
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "Content-Encoding": "identity",  # Sin compresión
                "Cache-Control": "no-transform"   # Prevenir transformaciones proxy
            }
        )
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta de reportes: {str(e)}"
        )

@app.get("/reportes_contratos/centro_gestor/{nombre_centro_gestor}", tags=["Interoperabilidad con Artefacto de Seguimiento"])
async def obtener_reportes_por_centro_gestor(nombre_centro_gestor: str):
    """
    ## � Obtener Reportes por Centro Gestor
    
    **Propósito**: Obtener reportes filtrados por nombre del centro gestor.
    Los resultados se ordenan por fecha de reporte descendente.
    
    ### 📋 Parámetros:
    - **nombre_centro_gestor**: Nombre del centro gestor para filtrar reportes
    
    ### � Ordenamiento:
    Los resultados se ordenan por `fecha_reporte` descendente (más recientes primero).
    
    ### 💡 Casos de uso:
    - Consultar reportes específicos de un centro gestor
    - Dashboard por centro de responsabilidad
    - Seguimiento por área organizacional
    """
    # Verificar disponibilidad de servicios
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE or not REPORTES_CONTRATOS_AVAILABLE:
        return {
            "success": False,
            "error": "Servicios no disponibles",
            "data": [],
            "count": 0
        }
    
    try:
        result = await get_reportes_by_centro_gestor(nombre_centro_gestor)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo reportes: {result.get('error', 'Error desconocido')}"
            )
        
        return create_utf8_response(result)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error obteniendo reportes por centro gestor: {str(e)}"
        )

@app.get("/reportes_contratos/referencia/{referencia_contrato}", tags=["Interoperabilidad con Artefacto de Seguimiento"])
async def obtener_reportes_por_referencia_contrato(referencia_contrato: str):
    """
    ## 📄 Obtener Reportes por Referencia de Contrato
    
    **Propósito**: Obtener reportes específicos de un contrato usando su referencia.
    Los resultados se ordenan por fecha de reporte descendente.
    
    ### 📋 Parámetros:
    - **referencia_contrato**: Referencia específica del contrato
    
    ### 📊 Ordenamiento:
    Los resultados se ordenan por `fecha_reporte` descendente (más recientes primero).
    
    ### 💡 Casos de uso:
    - Historial completo de reportes de un contrato específico
    - Seguimiento detallado por contrato
    - Auditoría de reportes por referencia
    """
    # Verificar disponibilidad de servicios
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE or not REPORTES_CONTRATOS_AVAILABLE:
        return {
            "success": False,
            "error": "Servicios no disponibles",
            "data": [],
            "count": 0
        }
    
    try:
        result = await get_reportes_by_referencia_contrato(referencia_contrato)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo reportes: {result.get('error', 'Error desconocido')}"
            )
        
        return create_utf8_response(result)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error obteniendo reportes por referencia: {str(e)}"
        )


# ============================================================================
# ENDPOINTS DE ADMINISTRACIÓN Y CONTROL DE ACCESOS
# ============================================================================

def check_user_management_availability():
    """✅ FUNCIONAL: Verificación simple sin lógica redundante"""
    if not (FIREBASE_AVAILABLE and USER_MANAGEMENT_AVAILABLE):
        raise HTTPException(
            status_code=503,
            detail={
                "error": "Servicios no disponibles",
                "code": "SERVICES_UNAVAILABLE"
            }
        )

@app.post("/auth/validate-session", tags=["Administración y Control de Accesos"])
async def validate_session(
    request: Request
):
    """
    ## 🔐 Validación de Sesión Activa para Next.js
    
    Valida si un token de ID de Firebase es válido y obtiene información completa del usuario.
    Optimizado para integración con Next.js y Firebase Auth SDK del frontend.
    
    ### ✅ Casos de uso:
    - Middleware de autenticación en Next.js
    - Verificación de permisos antes de acciones sensibles
    - Obtener datos actualizados del usuario
    - Validar sesiones activas desde el frontend
    
    ### 🔧 Proceso:
    1. Verifica token de Firebase desde Authorization header o body
    2. Valida estado del usuario (activo/deshabilitado)
    3. Obtiene datos completos de perfil desde Firestore
    4. Verifica permisos y roles
    
    ### 📝 Ejemplo de uso desde Next.js:
    ```javascript
    // En tu frontend NextJS
    import { getAuth } from 'firebase/auth';
    
    const auth = getAuth();
    const user = auth.currentUser;
    if (user) {
        const idToken = await user.getIdToken();
        const response = await fetch('/auth/validate-session', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'Authorization': `Bearer ${idToken}`
            },
            body: JSON.stringify({ id_token: idToken })
        });
        const data = await response.json();
        if (data.success) {
            console.log('Usuario autenticado:', data.user);
        }
    }
    ```
    """
    try:
        check_user_management_availability()
        
        # Obtener token del header Authorization o del body
        id_token = None
        
        # Primero intentar obtener del header Authorization
        auth_header = request.headers.get("authorization") or request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            id_token = auth_header.split(" ")[1]
        
        # Si no está en el header, intentar obtener del body
        if not id_token:
            try:
                body = await request.json()
                id_token = body.get("id_token")
            except:
                # Si no se puede parsear el JSON, intentar obtener como form data
                try:
                    form = await request.form()
                    id_token = form.get("id_token")
                except:
                    pass
        
        if not id_token:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "Token requerido",
                    "message": "Proporcione el token en el header Authorization o en el body como id_token",
                    "code": "TOKEN_REQUIRED"
                }
            )
        
        result = await validate_user_session(id_token)
        
        if not result["valid"]:
            raise HTTPException(
                status_code=401,
                detail={
                    "error": result["error"],
                    "code": result.get("code", "SESSION_INVALID")
                }
            )
        
        # Limpiar datos de Firebase antes de serializar
        clean_user_data = clean_firebase_data(result.get("user", {}))
        clean_token_data = clean_firebase_data(result.get("token_data", {}))
        
        return JSONResponse(
            content={
                "success": True,
                "session_valid": True,
                "user": clean_user_data,
                "token_info": clean_token_data,
                "verified_at": result.get("verified_at"),
                "message": "Sesión válida"
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Ocurrió un error inesperado durante la validación de sesión",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.post("/auth/login", tags=["Administración y Control de Accesos"])
async def login_user(login_data: UserLoginRequest):
    """
    ## 🔐 Autenticación de Usuario con Email y Contraseña
    
    Valida credenciales de usuario usando Firebase Authentication.
    Requiere email y contraseña válidos para permitir el acceso.
    
    ### Validaciones realizadas:
    - Formato de email válido
    - Contraseña correcta mediante Firebase Auth REST API
    - Usuario activo y no deshabilitado
    - Estado de cuenta en Firestore
    
    ### Respuesta exitosa:
    - Información completa del usuario
    - Tokens de Firebase para sesión
    - Datos adicionales de Firestore
    
    ### Errores comunes:
    - 401: Credenciales incorrectas
    - 403: Usuario deshabilitado o cuenta inactiva
    - 400: Formato de email inválido
    """
    try:
        check_user_management_availability()
        
        # Autenticación con validación real de credenciales
        result = await authenticate_email_password(login_data.email, login_data.password)
        
        # Verificar si la autenticación fue exitosa
        if result.get("success"):
            clean_user_data = clean_firebase_data(result.get("user", {}))
            
            # ✅ PREPARAR RESPUESTA CON CUSTOM TOKEN
            response_data = {
                "success": True,
                "user": clean_user_data,
                "auth_method": result.get("auth_method", "email_password"),
                "credentials_validated": result.get("credentials_validated", True),
                "message": result.get("message", "Autenticación exitosa"),
                "timestamp": datetime.now().isoformat()
            }
            
            # ✅ AGREGAR CUSTOM TOKEN SI ESTÁ DISPONIBLE
            if "custom_token" in result and result["custom_token"]:
                response_data["custom_token"] = result["custom_token"]
                response_data["token_usage"] = result.get("token_usage", "Use signInWithCustomToken() en Firebase Auth SDK")
            
            # Agregar información de autenticación alternativa si está disponible
            if "alternative_auth" in result:
                response_data["alternative_auth"] = result["alternative_auth"]
            
            # 🔍 LOG TEMPORAL PARA DEBUGGING
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"🔍 LOGIN RESPONSE KEYS: {list(response_data.keys())}")
            logger.info(f"⚠️  custom_token present: {'custom_token' in response_data}")
            if 'custom_token' in response_data:
                logger.info(f"✅ Token preview: {response_data['custom_token'][:50]}...")
            else:
                logger.warning(f"⚠️  No custom_token - Alternative auth available: {'alternative_auth' in response_data}")
            
            return JSONResponse(
                content=response_data,
                status_code=200,
                headers={"Content-Type": "application/json; charset=utf-8"}
            )
        else:
            # Autenticación fallida - mapear errores apropiados
            error_code = result.get("code", "AUTH_ERROR")
            
            # Mapear códigos de error a respuestas HTTP apropiadas
            if error_code in ["INVALID_CREDENTIALS", "USER_NOT_FOUND"]:
                raise HTTPException(
                    status_code=401,
                    detail={
                        "success": False,
                        "error": result["error"],
                        "code": error_code
                    }
                )
            elif error_code in ["USER_DISABLED", "ACCOUNT_INACTIVE"]:
                raise HTTPException(
                    status_code=403,
                    detail={
                        "success": False,
                        "error": result["error"],
                        "code": error_code
                    }
                )
            elif error_code in ["EMAIL_VALIDATION_ERROR", "INVALID_EMAIL_FORMAT"]:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "success": False,
                        "error": result["error"],
                        "code": error_code
                    }
                )
            else:
                # Cualquier otro error
                raise HTTPException(
                    status_code=500,
                    detail={
                        "success": False,
                        "error": result["error"],
                        "code": error_code
                    }
                )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error in login endpoint: {e}")
        return JSONResponse(
            content={
                "error": "Internal server error",
                "message": "An unexpected error occurred", 
                "fallback": True,
                "timestamp": datetime.now().isoformat()
            },
            status_code=500
        )

@app.get("/auth/register/health-check", tags=["Administración y Control de Accesos"])
async def register_health_check():
    """
    ## 🔍 Health Check para Registro de Usuario
    
    Verifica que todos los servicios necesarios para el registro estén disponibles.
    Útil para diagnosticar problemas en producción.
    """
    try:
        health_status = {
            "timestamp": datetime.now().isoformat(),
            "environment": os.getenv("ENVIRONMENT", "development"),
            "services": {}
        }
        
        # Verificar Firebase
        try:
            check_user_management_availability()
            health_status["services"]["user_management"] = {"status": "available", "error": None}
        except HTTPException as e:
            health_status["services"]["user_management"] = {
                "status": "unavailable", 
                "error": str(e.detail)
            }
        
        # Verificar importaciones
        health_status["services"]["imports"] = {
            "firebase_available": FIREBASE_AVAILABLE,
            "scripts_available": SCRIPTS_AVAILABLE,
            "user_management_available": USER_MANAGEMENT_AVAILABLE,
            "auth_operations_available": AUTH_OPERATIONS_AVAILABLE,
            "user_models_available": USER_MODELS_AVAILABLE
        }
        
        # Verificar configuración
        environment = os.getenv("ENVIRONMENT", "development")
        has_service_account = bool(os.getenv("FIREBASE_SERVICE_ACCOUNT_KEY"))
        
        health_status["configuration"] = {
            "project_id": PROJECT_ID,
            "environment": environment,
            "has_firebase_service_account": has_service_account,
            "firebase_available": FIREBASE_AVAILABLE,
            "auth_method": "Service Account Key" if has_service_account else "Workload Identity Federation",
            "authorized_domain": os.getenv("AUTHORIZED_EMAIL_DOMAIN", "@cali.gov.co"),
            "deployment_ready": FIREBASE_AVAILABLE  # Lo importante es que Firebase esté disponible
        }
        
        # Determinar estado general - soportar estructuras mixtas en 'services'
        def is_service_available(svc):
            """Evaluar si un servicio (o estructura) se considera disponible.

            - Si es dict y contiene 'status', se considera disponible cuando status == 'available'.
            - Si es dict con flags booleanos, se considera disponible cuando todos los flags booleanos son True.
            - Si es booleano, se usa su valor.
            - En cualquier otro caso se considera no disponible.
            """
            if isinstance(svc, dict):
                # Si tiene la clave 'status' respetarla
                if "status" in svc:
                    return svc.get("status") == "available"
                # Si es un diccionario de flags booleanas, todos deben ser True
                bool_flags = [v for v in svc.values() if isinstance(v, bool)]
                if bool_flags:
                    return all(bool_flags)
                # Fallback: consider available si el dict no está vacío
                return bool(svc)

            # Si es booleano, usar su valor
            if isinstance(svc, bool):
                return svc

            # Cualquier otro tipo se considera no disponible
            return False

        # Normalizar 'imports' a un campo 'status' legible para diagnósticos si procede
        imports_status = health_status["services"].get("imports")
        if isinstance(imports_status, dict) and "status" not in imports_status:
            health_status["services"]["imports"]["status"] = (
                "available" if is_service_available(imports_status) else "unavailable"
            )

        all_services_available = all(is_service_available(service) for service in health_status["services"].values())

        status_code = 200 if all_services_available else 503
        health_status["overall_status"] = "healthy" if all_services_available else "unhealthy"
        
        return JSONResponse(
            content=health_status,
            status_code=status_code,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except Exception as e:
        return JSONResponse(
            content={
                "overall_status": "error",
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            },
            status_code=500,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )

@app.post("/auth/register", tags=["Administración y Control de Accesos"], status_code=status.HTTP_201_CREATED)
async def register_user(registration_data: UserRegistrationRequest):
    """
    ✅ **REGISTRO DE USUARIO - VERSIÓN FUNCIONAL SIMPLIFICADA**
    
    **Fail Fast**: Si no hay Service Account configurado, falla inmediatamente
    **Sin Cache**: Cada request es independiente
    **Funcional**: Sin efectos colaterales entre registros
    """
    
    # � FAIL FAST: Verificar Service Account inmediatamente
    if not FIREBASE_AVAILABLE:
        environment = os.getenv("ENVIRONMENT", "development")
        if environment == "production":
            error_msg = "Firebase Service Account no configurado en producción"
            solution = "Configure FIREBASE_SERVICE_ACCOUNT_KEY en Railway"
        else:
            error_msg = "Firebase no disponible en desarrollo (requiere WIF o Service Account)"
            solution = "Configure Workload Identity Federation o FIREBASE_SERVICE_ACCOUNT_KEY"
        
        raise HTTPException(
            status_code=503,
            detail={
                "success": False,
                "error": error_msg,
                "code": "FIREBASE_UNAVAILABLE",
                "solution": solution,
                "environment": environment
            }
        )
    
    try:
        # ✅ PROGRAMACIÓN FUNCIONAL: Una sola responsabilidad
        result = await create_user_account(
            email=registration_data.email,
            password=registration_data.password,
            fullname=registration_data.name,
            cellphone=registration_data.cellphone,
            nombre_centro_gestor=registration_data.nombre_centro_gestor,
            send_email_verification=True
        )
        
        # ✅ FAIL FAST: Si hay error, fallar inmediatamente
        if not result.get("success", False):
            raise HTTPException(
                status_code=400,
                detail={
                    "success": False,
                    "error": result.get("error", "Error creando usuario"),
                    "code": result.get("code", "USER_CREATION_ERROR")
                }
            )
        
        # ✅ FUNCIONAL: Transformar datos sin mutación
        return {
            "success": True,
            "user": clean_firebase_data(result.get("user", {})),
            "message": "Usuario creado exitosamente",
            "timestamp": datetime.now().isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        # ✅ SIMPLE: Error handling directo
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "code": "INTERNAL_SERVER_ERROR",
                "debug": str(e) if os.getenv("ENVIRONMENT") == "development" else None
            }
        )

@app.post("/auth/change-password", tags=["Administración y Control de Accesos"])
async def change_password(
    uid: str = Form(..., description="ID del usuario"),
    new_password: str = Form(..., description="Nueva contraseña")
):
    """
    ## 🔒 Cambio de Contraseña
    
    Actualiza contraseñas de usuarios con validaciones de seguridad completas.
    
    ### ✅ Casos de uso:
    - Reset de contraseña por administrador
    - Cambio forzado por políticas de seguridad
    - Actualización por compromiso de cuenta
    
    ### 🔧 Validaciones:
    - Verificación de existencia del usuario
    - Validación de fortaleza de contraseña (8+ caracteres, mayúsculas, minúsculas, números, símbolos)
    - Actualización en Firebase Auth
    - Registro de timestamp en Firestore
    - Contador de cambios de contraseña
    
    ### 🛡️ Seguridad:
    - Solo administradores pueden cambiar contraseñas
    - Histórico de cambios para auditoría
    - Notificación automática al usuario
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const passwordData = {
      uid: "Zx9mK2pQ8RhV3nL7jM4uX1qW6tY0sA5e",
      new_password: "NuevaPassword123!"
    };
    const response = await fetch('/auth/change-password', {
      method: 'POST', 
      body: JSON.stringify(passwordData)
    });
    ```
    """
    try:
        check_user_management_availability()
        
        result = await update_user_password(uid, new_password)
        
        if not result.get("success", False):
            error_code = result.get("code", "PASSWORD_UPDATE_ERROR")
            error_message = result.get("error", "Error actualizando contraseña")
            
            if error_code == "USER_NOT_FOUND":
                raise HTTPException(
                    status_code=404, 
                    detail={
                        "success": False,
                        "error": error_message,
                        "code": error_code
                    }
                )
            else:
                raise HTTPException(
                    status_code=400, 
                    detail={
                        "success": False,
                        "error": error_message,
                        "code": error_code
                    }
                )
        
        return JSONResponse(
            content={
                "success": True,
                "message": result.get("message", "Contraseña actualizada exitosamente"),
                "updated_at": result.get("updated_at"),
                "timestamp": datetime.now().isoformat()
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Ocurrió un error inesperado durante el cambio de contraseña",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.get("/auth/config", tags=["Integración con el Frontend (NextJS)"])
async def get_firebase_config():
    """
    ## � Configuración Básica de Firebase para Frontend
    
    **ENDPOINT PÚBLICO** - Acceso directo desde frontend.
    
    Proporciona configuración mínima necesaria para Firebase Auth en frontend.
    
    ### �️ Seguridad:
    - Información pública solamente
    - Datos mínimos necesarios para SDK
    - Sin exposición de endpoints internos
    - Sin detalles de configuración sensibles
    
    ### � Información incluida:
    - Project ID de Firebase (público)
    - Auth Domain de Firebase (público)
    
    ### 🎯 Uso:
    - Inicialización de Firebase SDK en frontend
    - Configuración de autenticación client-side
    """
    # Solo información esencial para Firebase SDK
    return {
        "projectId": PROJECT_ID,
        "authDomain": f"{PROJECT_ID}.firebaseapp.com"
    }

# ENDPOINT REMOVIDO: /auth/integration-guide
# Razón: Documentación estática mejor manejada externamente
# Fecha: 2025-10-04
# La documentación de integración está disponible en README.md

@app.get("/auth/workload-identity/status", tags=["Administración y Control de Accesos"])
async def get_workload_identity_status():
    """
    ## 🔍 Estado de Autenticación con Google Cloud
    
    **ENDPOINT DE DIAGNÓSTICO** - Verifica el estado de autenticación con Google Cloud.
    
    ### 📊 Información incluida:
    - Estado de Service Account Key o Workload Identity
    - Validez de credenciales con Google Cloud
    - Configuración de Firebase
    - Nivel de seguridad actual
    
    ### 🛠️ Útil para:
    - Verificar configuración después de deployment en Railway
    - Diagnóstico de problemas de autenticación
    - Auditoría de seguridad
    - Monitoreo del sistema
    
    ### ⚠️ Nota:
    Este endpoint es principalmente para diagnóstico. En producción,
    considera eliminar o restringir acceso por seguridad.
    """
    try:
        from api.scripts.workload_identity_auth import get_workload_identity_status
        
        status = get_workload_identity_status()
        
        return {
            "success": True,
            "workload_identity_status": status,
            "system_ready": status.get("workload_identity", {}).get("initialized", False),
            "security_level": status.get("security_level", "unknown"),
            "timestamp": datetime.now().isoformat(),
            "message": "Estado de Workload Identity obtenido exitosamente"
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": "Error obteniendo estado de Workload Identity",
            "details": str(e),
            "fallback_available": True,
            "message": "Sistema puede funcionar en modo compatible"
        }

@app.post("/auth/google", tags=["Administración y Control de Accesos"])
async def google_auth_unified(
    google_token: str = Form(..., description="ID Token de Google Sign-In")
):
    """
    ## 🔐 Autenticación Google - ENDPOINT ÚNICO
    
    **EL ÚNICO ENDPOINT** que necesitas para autenticación Google completa.
    
    ### 🎯 **Funcionalidad Completa:**
    - ✅ Verifica token automáticamente con Workload Identity
    - ✅ Crea usuarios nuevos automáticamente
    - ✅ Actualiza usuarios existentes
    - ✅ Valida dominio @cali.gov.co
    - ✅ Retorna información completa del usuario
    - ✅ Máxima seguridad sin configuración manual
    
    ### � **Uso desde Frontend:**
    ```javascript
    // Después de Google Sign-In
    function handleGoogleAuth(response) {
        fetch('/auth/google', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ google_token: response.credential })
        })
        .then(res => res.json())
        .then(data => {
            if (data.success) {
                console.log('✅ Autenticado:', data.user);
                // Tu lógica aquí
            }
        });
    }
    ```
    
    ### 📱 **Compatible con:**
    - React, Vue, Angular, NextJS
    - Aplicaciones móviles
    - Progressive Web Apps
    - Cualquier framework que haga HTTP requests
    
    ### 🔒 **Seguridad:**
    - Workload Identity Federation
    - Sin credenciales en código
    - Verificación automática con Google
    - Auditoría completa de accesos
    """
    try:
        from api.scripts.workload_identity_auth import authenticate_with_workload_identity
        
        result = await authenticate_with_workload_identity(google_token)
        
        if not result["success"]:
            error_code = result.get("code", "GOOGLE_AUTH_ERROR")
            
            # Mapear errores específicos a códigos HTTP apropiados
            if error_code == "UNAUTHORIZED_DOMAIN":
                raise HTTPException(status_code=403, detail={
                    "error": "Dominio no autorizado",
                    "message": "Solo se permite autenticación con cuentas @cali.gov.co",
                    "code": "UNAUTHORIZED_DOMAIN"
                })
            elif error_code in ["INVALID_TOKEN", "TOKEN_VERIFICATION_ERROR"]:
                raise HTTPException(status_code=401, detail={
                    "error": "Token inválido",
                    "message": "El token de Google no es válido o ha expirado",
                    "code": "INVALID_TOKEN"
                })
            elif error_code == "WORKLOAD_IDENTITY_ERROR":
                raise HTTPException(status_code=503, detail={
                    "error": "Servicio no disponible",
                    "message": "Sistema de autenticación temporalmente no disponible",
                    "code": "SERVICE_UNAVAILABLE"
                })
            else:
                raise HTTPException(status_code=400, detail={
                    "error": "Error de autenticación",
                    "message": result.get("error", "Error desconocido"),
                    "code": error_code
                })
        
        # Limpiar datos de Firebase antes de serializar
        clean_user_data = clean_firebase_data(result["user"])
        
        return {
            "success": True,
            "user": clean_user_data,
            "auth_method": "workload_identity_google",
            "security_level": "high",
            "user_created": result.get("user_created", False),
            "message": result["message"],
            "timestamp": datetime.now().isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Unexpected error in Google auth: {e}")  # Usar print en lugar de logger
        raise HTTPException(status_code=500, detail={
            "error": "Error interno del servidor",
            "message": "Por favor, inténtelo de nuevo más tarde",
            "code": "INTERNAL_ERROR"
        })

# ============================================================================
# ENDPOINTS DE ELIMINACIÓN DE USUARIOS
# ============================================================================

@app.delete("/auth/user/{uid}", tags=["Administración y Control de Accesos"])
async def delete_user(uid: str, soft_delete: Optional[bool] = Query(default=None, description="Eliminación lógica (true) o física (false)")):
    """
    ## 🗑️ Eliminación de Usuario
    
    Elimina cuentas con opciones flexibles de soft delete (recomendado) o hard delete.
    
    ### ✅ Casos de uso:
    - Desvinculación de empleados (soft delete)
    - Limpieza de cuentas de prueba (hard delete)
    - Cumplimiento de políticas de retención de datos
    
    ### 🔧 Tipos de eliminación:
    - **Soft delete (predeterminado)**: Deshabilita usuario, mantiene datos para auditoría
    - **Hard delete**: Elimina completamente de Firebase Auth y Firestore
    
    ### 🛡️ Protecciones:
    - No permite eliminar el último administrador del sistema
    - Validación de permisos para hard delete
    - Registro de auditoría de eliminaciones
    
    ### 📝 Ejemplos de uso:
    ```javascript
    // Eliminación lógica (recomendada)
    const response = await fetch('/auth/user/Zx9mK2pQ8RhV3nL7jM4uX1qW6tY0sA5e?soft_delete=true', {
      method: 'DELETE'
    });
    
    // Eliminación física (permanente)
    const response = await fetch('/auth/user/Zx9mK2pQ8RhV3nL7jM4uX1qW6tY0sA5e?soft_delete=false', {
      method: 'DELETE'
    });
    ```
    """
    try:
        check_user_management_availability()
        
        result = await delete_user_account(uid, soft_delete if soft_delete is not None else True)
        
        if not result.get("success", False):
            error_code = result.get("code", "USER_DELETE_ERROR")
            error_message = result.get("error", "Error eliminando usuario")
            
            if error_code == "USER_NOT_FOUND":
                raise HTTPException(
                    status_code=404, 
                    detail={
                        "success": False,
                        "error": error_message,
                        "code": error_code
                    }
                )
            else:
                raise HTTPException(
                    status_code=400, 
                    detail={
                        "success": False,
                        "error": error_message,
                        "code": error_code
                    }
                )
        
        return JSONResponse(
            content={
                "success": True,
                "message": result.get("message", "Usuario eliminado exitosamente"),
                "deleted_at": result.get("deleted_at"),
                "soft_delete": result.get("soft_delete", True),
                "timestamp": datetime.now().isoformat()
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Ocurrió un error inesperado durante la eliminación",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

# ============================================================================
# ENDPOINTS ADMINISTRATIVOS DE USUARIOS
# ============================================================================

@app.get("/admin/users", tags=["Administración y Control de Accesos"])
async def list_system_users(
    limit: int = Query(default=100, ge=1, le=1000, description="Límite de resultados por página")
):
    """
    ## 📋 Listado de Usuarios desde Firestore
    
    Lee directamente la colección "users" de Firestore y devuelve todos los usuarios registrados.
    
    ### � Información incluida:
    - UID del usuario
    - Email y nombre completo
    - Teléfono y centro gestor
    - Fechas de creación y actualización
    - Estado de activación y verificación
    - Proveedores de autenticación
    - Estadísticas de login
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const response = await fetch('/admin/users?limit=50');
    const data = await response.json();
    console.log(`Encontrados ${data.count} usuarios`);
    ```
    """
    try:
        check_user_management_availability()
        
        from database.firebase_config import get_firestore_client
        
        firestore_client = get_firestore_client()
        
        # Consultar la colección "users" directamente
        users_ref = firestore_client.collection('users')
        query = users_ref.limit(limit)
        docs = query.get()
        
        users_list = []
        for doc in docs:
            if doc.exists:
                user_data = doc.to_dict()
                
                user_info = {
                    "uid": doc.id,
                    "email": user_data.get("email"),
                    "fullname": user_data.get("fullname"),
                    "cellphone": user_data.get("cellphone"),
                    "nombre_centro_gestor": user_data.get("nombre_centro_gestor"),
                    "created_at": user_data.get("created_at"),
                    "updated_at": user_data.get("updated_at"),
                    "is_active": user_data.get("is_active", True),
                    "email_verified": user_data.get("email_verified", False),
                    "can_use_google_auth": user_data.get("can_use_google_auth", False),
                    "auth_providers": user_data.get("auth_providers", []),
                    "last_login": user_data.get("last_login"),
                    "login_count": user_data.get("login_count", 0)
                }
                
                # Limpiar datos de Firebase antes de agregar a la lista
                user_info = clean_firebase_data(user_info)
                users_list.append(user_info)
        
        return JSONResponse(
            content={
                "success": True,
                "users": users_list,
                "count": len(users_list),
                "collection": "users",
                "timestamp": datetime.now().isoformat(),
                "message": f"Se obtuvieron {len(users_list)} usuarios de la colección 'users'"
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail={
                "success": False,
                "error": str(e),
                "message": "Error leyendo la colección 'users' de Firestore",
                "code": "FIRESTORE_READ_ERROR"
            }
        )

# ============================================================================
# ENDPOINTS DE GESTIÓN DE EMPRÉSTITO
# ============================================================================

# Verificar disponibilidad de operaciones de empréstito
try:
    from api.scripts import (
        procesar_emprestito_completo,
        verificar_proceso_existente,
        eliminar_proceso_emprestito,
        actualizar_proceso_emprestito,
        obtener_codigos_contratos,
        buscar_y_poblar_contratos_secop,
        obtener_contratos_desde_proceso_contractual,
        get_emprestito_operations_status,
        cargar_orden_compra_directa,
        cargar_convenio_transferencia,
        modificar_convenio_transferencia,
        cargar_rpc_emprestito,
        cargar_pago_emprestito,
        get_pagos_emprestito_all,
        get_rpc_contratos_emprestito_all,
        get_asignaciones_emprestito_banco_centro_gestor_all,
        get_convenios_transferencia_emprestito_all,
        obtener_ordenes_compra_tvec_enriquecidas,
        get_tvec_enrich_status,
        get_ordenes_compra_emprestito_all,
        get_ordenes_compra_emprestito_by_referencia,
        get_ordenes_compra_emprestito_by_centro_gestor,
        EMPRESTITO_OPERATIONS_AVAILABLE,
        TVEC_ENRICH_OPERATIONS_AVAILABLE,
        ORDENES_COMPRA_OPERATIONS_AVAILABLE
    )
    from api.models import (
        EmprestitoRequest, 
        EmprestitoResponse,
        PagoEmprestitoRequest,
        PagoEmprestitoResponse
    )
    print(f"✅ Empréstito imports successful - AVAILABLE: {EMPRESTITO_OPERATIONS_AVAILABLE}")
    print(f"✅ TVEC enrich imports successful - AVAILABLE: {TVEC_ENRICH_OPERATIONS_AVAILABLE}")
except ImportError as e:
    print(f"❌ Warning: Empréstito or TVEC imports failed: {e}")
    EMPRESTITO_OPERATIONS_AVAILABLE = False
    TVEC_ENRICH_OPERATIONS_AVAILABLE = False

def check_emprestito_availability():
    """Verificar disponibilidad de operaciones de empréstito"""
    if not EMPRESTITO_OPERATIONS_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail={
                "error": "Servicios de empréstito no disponibles",
                "message": "Firebase o dependencias no configuradas correctamente",
                "code": "EMPRESTITO_SERVICES_UNAVAILABLE"
            }
        )

@app.post("/emprestito/cargar-proceso", tags=["Gestión de Empréstito"], summary="🟢 Cargar Proceso de Empréstito")
async def cargar_proceso_emprestito(
    referencia_proceso: str = Form(..., description="Referencia del proceso (obligatorio)"),
    nombre_centro_gestor: str = Form(..., description="Centro gestor responsable (obligatorio)"),
    nombre_banco: str = Form(..., description="Nombre del banco (obligatorio)"),
    plataforma: str = Form(..., description="Plataforma (SECOP, TVEC) (obligatorio)"),
    bp: Optional[str] = Form(None, description="Código BP (opcional)"),
    nombre_resumido_proceso: Optional[str] = Form(None, description="Nombre resumido del proceso (opcional)"),
    id_paa: Optional[str] = Form(None, description="ID PAA (opcional)"),
    valor_proyectado: Optional[float] = Form(None, description="Valor proyectado (opcional)")
):
    """
    ## � POST | 📥 Carga de Datos | Cargar Proceso de Empréstito
    
    Endpoint unificado para carga de procesos de empréstito con detección automática 
    de plataforma (SECOP/TVEC) y validación de duplicados.
    
    ### ✅ Funcionalidades principales:
    - **Detección automática**: Identifica si es SECOP o TVEC basado en el campo `plataforma`
    - **Validación de duplicados**: Verifica existencia previa usando `referencia_proceso`
    - **Integración API**: Obtiene datos completos desde APIs externas (SECOP/TVEC)
    - **Almacenamiento inteligente**: Guarda en colección apropiada según plataforma
    
    ### 🔍 Detección de plataforma:
    **SECOP**: "SECOP", "SECOP II", "SECOP I", "SECOP 2", "SECOP 1" y variantes
    **TVEC**: "TVEC" y variantes
    
    ### 📊 Almacenamiento por plataforma:
    - **SECOP** → Colección: `procesos_emprestito`
    - **TVEC** → Colección: `ordenes_compra_emprestito`
    
    ### 🛡️ Validación de duplicados:
    Busca `referencia_proceso` en ambas colecciones antes de crear nuevo registro.
    
    ### ⚙️ Campos obligatorios:
    - `referencia_proceso`: Referencia del proceso
    - `nombre_centro_gestor`: Centro gestor responsable
    - `nombre_banco`: Nombre del banco
    - `plataforma`: Plataforma (SECOP/TVEC)
    
    ### 📝 Campos opcionales:
    - `bp`: Código BP
    - `nombre_resumido_proceso`: Nombre resumido
    - `id_paa`: ID PAA
    - `valor_proyectado`: Valor proyectado
    
    ### 🔗 Integración con APIs:
    **SECOP**: Obtiene datos desde API de datos abiertos (p6dx-8zbt)
    **TVEC**: Obtiene datos desde API TVEC (rgxm-mmea)
    
    ### 📋 Ejemplo de request:
    ```json
    {
        "referencia_proceso": "SCMGSU-CM-003-2024",
        "nombre_centro_gestor": "Secretaría de Salud",
        "nombre_banco": "Banco Mundial",
        "bp": "BP-2024-001",
        "plataforma": "SECOP II",
        "nombre_resumido_proceso": "Suministro equipos médicos",
        "id_paa": "PAA-2024-123",
        "valor_proyectado": 1500000000.0
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        # Crear diccionario con los datos del formulario
        datos_emprestito = {
            "referencia_proceso": referencia_proceso,
            "nombre_centro_gestor": nombre_centro_gestor,
            "nombre_banco": nombre_banco,
            "bp": bp,
            "plataforma": plataforma,
            "nombre_resumido_proceso": nombre_resumido_proceso,
            "id_paa": id_paa,
            "valor_proyectado": valor_proyectado
        }
        
        # Procesar empréstito completo con todas las validaciones
        resultado = await procesar_emprestito_completo(datos_emprestito)
        
        # Manejar respuesta según el resultado
        if not resultado.get("success"):
            # Manejar caso especial de duplicado
            if resultado.get("duplicate"):
                return JSONResponse(
                    content={
                        "success": False,
                        "error": resultado.get("error"),
                        "duplicate": True,
                        "existing_data": resultado.get("existing_data"),
                        "message": "Ya existe un proceso con esta referencia",
                        "timestamp": datetime.now().isoformat()
                    },
                    status_code=409,  # Conflict
                    headers={"Content-Type": "application/json; charset=utf-8"}
                )
            else:
                # Error general
                return JSONResponse(
                    content={
                        "success": False,
                        "error": resultado.get("error"),
                        "plataforma_detectada": resultado.get("plataforma_detectada"),
                        "message": "Error procesando proceso de empréstito",
                        "timestamp": datetime.now().isoformat()
                    },
                    status_code=400,
                    headers={"Content-Type": "application/json; charset=utf-8"}
                )
        
        # Éxito: proceso creado correctamente
        respuesta_base = {
            "success": True,
            "message": "Proceso de empréstito cargado exitosamente",
            "data": resultado.get("data"),
            "doc_id": resultado.get("doc_id"),
            "coleccion": resultado.get("coleccion"),
            "plataforma_detectada": resultado.get("plataforma_detectada"),
            "fuente_datos": resultado.get("fuente_datos"),
            "timestamp": datetime.now().isoformat()
        }
        
        # Si es un proceso SECOP, intentar actualizar con datos completos automáticamente
        if resultado.get("plataforma_detectada") == "SECOP" and resultado.get("coleccion") == "procesos_emprestito":
            try:
                logger.info(f"🔄 Actualizando automáticamente proceso SECOP: {referencia_proceso}")
                resultado_actualizacion = await actualizar_proceso_emprestito_completo(referencia_proceso)
                
                if resultado_actualizacion.get("success"):
                    respuesta_base["actualizacion_completa"] = {
                        "success": True,
                        "changes_count": resultado_actualizacion.get("changes_count", 0),
                        "changes_summary": resultado_actualizacion.get("changes_summary", [])[:5],  # Máximo 5 cambios en resumen
                        "message": f"Proceso actualizado automáticamente con {resultado_actualizacion.get('changes_count', 0)} campos adicionales"
                    }
                    logger.info(f"✅ Actualización automática exitosa: {resultado_actualizacion.get('changes_count', 0)} cambios")
                else:
                    respuesta_base["actualizacion_completa"] = {
                        "success": False,
                        "error": resultado_actualizacion.get("error", "Error desconocido"),
                        "message": "No se pudo actualizar automáticamente con datos completos"
                    }
                    logger.warning(f"⚠️ Actualización automática falló: {resultado_actualizacion.get('error')}")
                    
            except Exception as e:
                logger.warning(f"⚠️ Error en actualización automática: {e}")
                respuesta_base["actualizacion_completa"] = {
                    "success": False,
                    "error": str(e),
                    "message": "Error durante actualización automática (proceso principal creado exitosamente)"
                }
        
        return JSONResponse(
            content=respuesta_base,
            status_code=201,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de empréstito: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Por favor, inténtelo de nuevo más tarde",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.post("/emprestito/cargar-orden-compra", tags=["Gestión de Empréstito"], summary="🟢 Cargar Orden de Compra")
async def cargar_orden_compra_emprestito(
    numero_orden: str = Form(..., description="Número de la orden de compra (obligatorio)"),
    nombre_centro_gestor: str = Form(..., description="Centro gestor responsable (obligatorio)"),
    nombre_banco: str = Form(..., description="Nombre del banco (obligatorio)"),
    nombre_resumido_proceso: str = Form(..., description="Nombre resumido del proceso (obligatorio)"),
    valor_proyectado: float = Form(..., description="Valor proyectado (obligatorio)"),
    bp: Optional[str] = Form(None, description="Código BP (opcional)")
):
    """
    ## � POST | 📥 Carga de Datos | Cargar Orden de Compra de Empréstito
    
    Endpoint para carga directa de órdenes de compra de empréstito en la colección 
    `ordenes_compra_emprestito` sin procesamiento de APIs externas.
    
    ### ✅ Funcionalidades principales:
    - **Carga directa**: Registra directamente en `ordenes_compra_emprestito`
    - **Validación de duplicados**: Verifica existencia previa usando `numero_orden`
    - **Validación de campos**: Verifica que todos los campos obligatorios estén presentes
    - **Timestamps automáticos**: Agrega fecha de creación y actualización
    
    ### ⚙️ Campos obligatorios:
    - `numero_orden`: Número único de la orden de compra
    - `nombre_centro_gestor`: Centro gestor responsable
    - `nombre_banco`: Nombre del banco
    - `nombre_resumido_proceso`: Nombre resumido del proceso
    - `valor_proyectado`: Valor proyectado en pesos colombianos
    
    ### 📝 Campos opcionales:
    - `bp`: Código BP
    
    ### 🛡️ Validación de duplicados:
    Busca `numero_orden` en la colección `ordenes_compra_emprestito` antes de crear nuevo registro.
    
    ### 📊 Estructura de datos guardados:
    ```json
    {
        "numero_orden": "OC-2024-001",
        "nombre_centro_gestor": "Secretaría de Salud",
        "nombre_banco": "Banco Mundial",
        "nombre_resumido_proceso": "Suministro equipos médicos",
        "valor_proyectado": 1500000000.0,
        "bp": "BP-2024-001",
        "fecha_creacion": "2024-10-14T10:30:00",
        "fecha_actualizacion": "2024-10-14T10:30:00",
        "estado": "activo",
        "tipo": "orden_compra_manual"
    }
    ```
    
    ### 📋 Ejemplo de request:
    ```json
    {
        "numero_orden": "OC-SALUD-003-2024",
        "nombre_centro_gestor": "Secretaría de Salud",
        "nombre_banco": "Banco Mundial",
        "nombre_resumido_proceso": "Suministro equipos médicos",
        "valor_proyectado": 1500000000.0,
        "bp": "BP-2024-001"
    }
    ```
    
    ### ✅ Respuesta exitosa (201):
    ```json
    {
        "success": true,
        "message": "Orden de compra OC-SALUD-003-2024 guardada exitosamente",
        "doc_id": "abc123def456",
        "data": { ... },
        "coleccion": "ordenes_compra_emprestito"
    }
    ```
    
    ### ❌ Respuesta de duplicado (409):
    ```json
    {
        "success": false,
        "error": "Ya existe una orden de compra con número: OC-SALUD-003-2024",
        "duplicate": true,
        "existing_data": { ... }
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        # Crear diccionario con los datos del formulario
        datos_orden = {
            "numero_orden": numero_orden,
            "nombre_centro_gestor": nombre_centro_gestor,
            "nombre_banco": nombre_banco,
            "nombre_resumido_proceso": nombre_resumido_proceso,
            "valor_proyectado": valor_proyectado,
            "bp": bp
        }
        
        # Procesar orden de compra
        resultado = await cargar_orden_compra_directa(datos_orden)
        
        # Manejar respuesta según el resultado
        if not resultado.get("success"):
            # Manejar caso especial de duplicado
            if resultado.get("duplicate"):
                return JSONResponse(
                    content={
                        "success": False,
                        "error": resultado.get("error"),
                        "duplicate": True,
                        "existing_data": resultado.get("existing_data"),
                        "message": "Ya existe una orden de compra con este número",
                        "timestamp": datetime.now().isoformat()
                    },
                    status_code=409,  # Conflict
                    headers={"Content-Type": "application/json; charset=utf-8"}
                )
            else:
                # Error general
                return JSONResponse(
                    content={
                        "success": False,
                        "error": resultado.get("error"),
                        "message": "Error al procesar la orden de compra",
                        "timestamp": datetime.now().isoformat()
                    },
                    status_code=400,
                    headers={"Content-Type": "application/json; charset=utf-8"}
                )
        
        # Respuesta exitosa
        return JSONResponse(
            content={
                "success": True,
                "message": resultado.get("message"),
                "data": resultado.get("data"),
                "doc_id": resultado.get("doc_id"),
                "coleccion": resultado.get("coleccion"),
                "timestamp": datetime.now().isoformat()
            },
            status_code=201,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de orden de compra: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Por favor, inténtelo de nuevo más tarde",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.post("/emprestito/cargar-convenio-transferencia", tags=["Gestión de Empréstito"], summary="🟢 Cargar Convenio de Transferencia")
async def cargar_convenio_transferencia_emprestito(
    referencia_contrato: str = Form(..., description="Referencia del contrato/convenio (obligatorio)"),
    nombre_centro_gestor: str = Form(..., description="Centro gestor responsable (obligatorio)"),
    banco: str = Form(..., description="Nombre del banco (obligatorio)"),
    objeto_contrato: str = Form(..., description="Objeto del contrato (obligatorio)"),
    valor_contrato: float = Form(..., description="Valor del contrato (obligatorio)"),
    bp: Optional[str] = Form(None, description="Código BP (opcional)"),
    bpin: Optional[str] = Form(None, description="Código BPIN (opcional)"),
    valor_convenio: Optional[float] = Form(None, description="Valor del convenio (opcional)"),
    urlproceso: Optional[str] = Form(None, description="URL del proceso (opcional)"),
    fecha_inicio_contrato: Optional[str] = Form(None, description="Fecha de inicio del contrato (opcional)"),
    fecha_fin_contrato: Optional[str] = Form(None, description="Fecha de fin del contrato (opcional)"),
    modalidad_contrato: Optional[str] = Form(None, description="Modalidad del contrato (opcional)"),
    ordenador_gastor: Optional[str] = Form(None, description="Ordenador del gasto (opcional)"),
    tipo_contrato: Optional[str] = Form(None, description="Tipo de contrato (opcional)"),
    estado_contrato: Optional[str] = Form(None, description="Estado del contrato (opcional)"),
    sector: Optional[str] = Form(None, description="Sector (opcional)"),
    nombre_resumido_proceso: str = Form(..., description="Nombre resumido del proceso (obligatorio)")
):
    """
    ## 📝 POST | 📥 Carga de Datos | Cargar Convenio de Transferencia de Empréstito
    
    Endpoint para carga directa de convenios de transferencia de empréstito en la colección 
    `convenios_transferencias_emprestito` sin procesamiento de APIs externas.
    
    ### ✅ Funcionalidades principales:
    - **Carga directa**: Registra directamente en `convenios_transferencias_emprestito`
    - **Validación de duplicados**: Verifica existencia previa usando `referencia_contrato`
    - **Validación de campos**: Verifica que todos los campos obligatorios estén presentes
    - **Timestamps automáticos**: Agrega fecha de creación y actualización
    
    ### ⚙️ Campos obligatorios:
    - `referencia_contrato`: Referencia única del contrato/convenio
    - `nombre_centro_gestor`: Centro gestor responsable
    - `banco`: Nombre del banco
    - `objeto_contrato`: Descripción del objeto del contrato
    - `valor_contrato`: Valor del contrato en pesos colombianos
    
    ### 📝 Campos opcionales:
    - `bp`: Código BP
    - `bpin`: Código BPIN (Banco de Programas y Proyectos de Inversión Nacional)
    - `valor_convenio`: Valor específico del convenio
    - `urlproceso`: URL del proceso de contratación
    - `fecha_inicio_contrato`: Fecha de inicio del contrato
    - `fecha_fin_contrato`: Fecha de finalización del contrato
    - `modalidad_contrato`: Modalidad de contratación
    - `ordenador_gastor`: Ordenador del gasto
    - `tipo_contrato`: Tipo de contrato
    - `estado_contrato`: Estado actual del contrato
    - `sector`: Sector al que pertenece
    
    ### 🛡️ Validación de duplicados:
    Busca `referencia_contrato` en la colección `convenios_transferencias_emprestito` antes de crear nuevo registro.
    
    ### 📊 Estructura de datos guardados:
    ```json
    {
        "referencia_contrato": "CONV-2024-001",
        "nombre_centro_gestor": "Secretaría de Salud",
        "banco": "Banco Mundial",
        "objeto_contrato": "Convenio de transferencia para equipamiento médico",
        "valor_contrato": 1500000000.0,
        "valor_convenio": 1200000000.0,
        "bp": "BP-2024-001",
        "bpin": "2024000010001",
        "urlproceso": "https://...",
        "fecha_inicio_contrato": "2024-01-15",
        "fecha_fin_contrato": "2024-12-31",
        "modalidad_contrato": "Convenio de Transferencia",
        "ordenador_gastor": "Juan Pérez",
        "tipo_contrato": "Transferencia",
        "estado_contrato": "Activo",
        "sector": "Salud",
        "fecha_creacion": "2024-10-14T10:30:00",
        "fecha_actualizacion": "2024-10-14T10:30:00",
        "estado": "activo",
        "tipo": "convenio_transferencia_manual"
    }
    ```
    
    ### 📋 Ejemplo de request:
    ```json
    {
        "referencia_contrato": "CONV-SALUD-003-2024",
        "nombre_centro_gestor": "Secretaría de Salud",
        "banco": "Banco Mundial",
        "objeto_contrato": "Convenio de transferencia para equipamiento médico",
        "valor_contrato": 1500000000.0,
        "valor_convenio": 1200000000.0,
        "bp": "BP-2024-001",
        "modalidad_contrato": "Convenio de Transferencia",
        "estado_contrato": "Activo"
    }
    ```
    
    ### ✅ Respuesta exitosa (201):
    ```json
    {
        "success": true,
        "message": "Convenio de transferencia CONV-SALUD-003-2024 guardado exitosamente",
        "doc_id": "abc123def456",
        "data": { ... },
        "coleccion": "convenios_transferencias_emprestito"
    }
    ```
    
    ### ❌ Respuesta de duplicado (409):
    ```json
    {
        "success": false,
        "error": "Ya existe un convenio de transferencia con referencia: CONV-SALUD-003-2024",
        "duplicate": true,
        "existing_data": { ... }
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        # Crear diccionario con los datos del formulario
        datos_convenio = {
            "referencia_contrato": referencia_contrato,
            "nombre_centro_gestor": nombre_centro_gestor,
            "banco": banco,
            "objeto_contrato": objeto_contrato,
            "valor_contrato": valor_contrato,
            "bp": bp,
            "bpin": bpin,
            "valor_convenio": valor_convenio,
            "urlproceso": urlproceso,
            "fecha_inicio_contrato": fecha_inicio_contrato,
            "fecha_fin_contrato": fecha_fin_contrato,
            "modalidad_contrato": modalidad_contrato,
            "ordenador_gastor": ordenador_gastor,
            "tipo_contrato": tipo_contrato,
            "estado_contrato": estado_contrato,
            "sector": sector,
            "nombre_resumido_proceso": nombre_resumido_proceso
        }
        
        # Procesar convenio de transferencia
        resultado = await cargar_convenio_transferencia(datos_convenio)
        
        # Manejar respuesta según el resultado
        if not resultado.get("success"):
            # Manejar caso especial de duplicado
            if resultado.get("duplicate"):
                return JSONResponse(
                    content={
                        "success": False,
                        "error": resultado.get("error"),
                        "duplicate": True,
                        "existing_data": resultado.get("existing_data"),
                        "message": "Ya existe un convenio de transferencia con esta referencia",
                        "timestamp": datetime.now().isoformat()
                    },
                    status_code=409,  # Conflict
                    headers={"Content-Type": "application/json; charset=utf-8"}
                )
            else:
                # Error general
                return JSONResponse(
                    content={
                        "success": False,
                        "error": resultado.get("error"),
                        "message": "Error al procesar el convenio de transferencia",
                        "timestamp": datetime.now().isoformat()
                    },
                    status_code=400,
                    headers={"Content-Type": "application/json; charset=utf-8"}
                )
        
        # Respuesta exitosa
        return JSONResponse(
            content={
                "success": True,
                "message": resultado.get("message"),
                "data": resultado.get("data"),
                "doc_id": resultado.get("doc_id"),
                "coleccion": resultado.get("coleccion"),
                "timestamp": datetime.now().isoformat()
            },
            status_code=201,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de convenio de transferencia: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Por favor, inténtelo de nuevo más tarde",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.put("/emprestito/modificar-convenio-transferencia", tags=["Gestión de Empréstito"], summary="🟠 Modificar Convenio de Transferencia")
async def modificar_convenio_transferencia_emprestito(
    doc_id: str = Form(..., description="ID del documento a modificar (obligatorio)"),
    referencia_contrato: Optional[str] = Form(None, description="Referencia del contrato/convenio (opcional)"),
    nombre_centro_gestor: Optional[str] = Form(None, description="Centro gestor responsable (opcional)"),
    banco: Optional[str] = Form(None, description="Nombre del banco (opcional)"),
    objeto_contrato: Optional[str] = Form(None, description="Objeto del contrato (opcional)"),
    valor_contrato: Optional[float] = Form(None, description="Valor del contrato (opcional)"),
    bp: Optional[str] = Form(None, description="Código BP (opcional)"),
    bpin: Optional[str] = Form(None, description="Código BPIN (opcional)"),
    valor_convenio: Optional[float] = Form(None, description="Valor del convenio (opcional)"),
    urlproceso: Optional[str] = Form(None, description="URL del proceso (opcional)"),
    fecha_inicio_contrato: Optional[str] = Form(None, description="Fecha de inicio del contrato (opcional)"),
    fecha_fin_contrato: Optional[str] = Form(None, description="Fecha de fin del contrato (opcional)"),
    modalidad_contrato: Optional[str] = Form(None, description="Modalidad del contrato (opcional)"),
    ordenador_gastor: Optional[str] = Form(None, description="Ordenador del gasto (opcional)"),
    tipo_contrato: Optional[str] = Form(None, description="Tipo de contrato (opcional)"),
    estado_contrato: Optional[str] = Form(None, description="Estado del contrato (opcional)"),
    sector: Optional[str] = Form(None, description="Sector (opcional)"),
    nombre_resumido_proceso: Optional[str] = Form(None, description="Nombre resumido del proceso (opcional)")
):
    """
    ## 🟠 PUT | ✏️ Actualización | Modificar Convenio de Transferencia de Empréstito
    
    Endpoint para modificar cualquier campo de un convenio de transferencia existente 
    en la colección `convenios_transferencias_emprestito`.
    
    ### ✅ Funcionalidades principales:
    - **Actualización flexible**: Permite modificar cualquier campo del convenio
    - **Actualización parcial**: Solo se actualizan los campos proporcionados
    - **Validación de existencia**: Verifica que el documento exista antes de actualizar
    - **Timestamp automático**: Actualiza automáticamente `fecha_actualizacion`
    - **Preservación de datos**: Los campos no proporcionados mantienen sus valores originales
    
    ### ⚙️ Campo obligatorio:
    - `doc_id`: ID del documento de Firestore que se desea modificar
    
    ### 📝 Campos opcionales (todos):
    Cualquiera de estos campos puede ser actualizado:
    - `referencia_contrato`: Referencia del contrato/convenio
    - `nombre_centro_gestor`: Centro gestor responsable
    - `banco`: Nombre del banco
    - `objeto_contrato`: Objeto del contrato
    - `valor_contrato`: Valor del contrato
    - `bp`: Código BP
    - `bpin`: Código BPIN
    - `valor_convenio`: Valor del convenio
    - `urlproceso`: URL del proceso
    - `fecha_inicio_contrato`: Fecha de inicio
    - `fecha_fin_contrato`: Fecha de finalización
    - `modalidad_contrato`: Modalidad de contratación
    - `ordenador_gastor`: Ordenador del gasto
    - `tipo_contrato`: Tipo de contrato
    - `estado_contrato`: Estado actual
    - `sector`: Sector al que pertenece
    - `nombre_resumido_proceso`: Nombre resumido del proceso
    
    ### 📋 Ejemplo de request (actualización parcial):
    ```json
    {
        "doc_id": "abc123def456",
        "estado_contrato": "Finalizado",
        "fecha_fin_contrato": "2024-12-31"
    }
    ```
    
    ### ✅ Respuesta exitosa (200):
    ```json
    {
        "success": true,
        "message": "Convenio de transferencia actualizado exitosamente",
        "doc_id": "abc123def456",
        "campos_actualizados": ["estado_contrato", "fecha_fin_contrato"],
        "data": { ... },
        "timestamp": "2024-11-17T10:30:00"
    }
    ```
    
    ### ❌ Respuesta de error (404):
    ```json
    {
        "success": false,
        "error": "No se encontró el convenio de transferencia con ID: abc123",
        "doc_id": "abc123"
    }
    ```
    
    ### 🔗 Endpoints relacionados:
    - `POST /emprestito/cargar-convenio-transferencia` - Para crear nuevos convenios
    - `GET /convenios_transferencias_all` - Para consultar convenios existentes
    """
    try:
        check_emprestito_availability()
        
        # Crear diccionario con los campos a actualizar
        campos_actualizar = {}
        
        if referencia_contrato is not None:
            campos_actualizar["referencia_contrato"] = referencia_contrato
        if nombre_centro_gestor is not None:
            campos_actualizar["nombre_centro_gestor"] = nombre_centro_gestor
        if banco is not None:
            campos_actualizar["banco"] = banco
        if objeto_contrato is not None:
            campos_actualizar["objeto_contrato"] = objeto_contrato
        if valor_contrato is not None:
            campos_actualizar["valor_contrato"] = valor_contrato
        if bp is not None:
            campos_actualizar["bp"] = bp
        if bpin is not None:
            campos_actualizar["bpin"] = bpin
        if valor_convenio is not None:
            campos_actualizar["valor_convenio"] = valor_convenio
        if urlproceso is not None:
            campos_actualizar["urlproceso"] = urlproceso
        if fecha_inicio_contrato is not None:
            campos_actualizar["fecha_inicio_contrato"] = fecha_inicio_contrato
        if fecha_fin_contrato is not None:
            campos_actualizar["fecha_fin_contrato"] = fecha_fin_contrato
        if modalidad_contrato is not None:
            campos_actualizar["modalidad_contrato"] = modalidad_contrato
        if ordenador_gastor is not None:
            campos_actualizar["ordenador_gastor"] = ordenador_gastor
        if tipo_contrato is not None:
            campos_actualizar["tipo_contrato"] = tipo_contrato
        if estado_contrato is not None:
            campos_actualizar["estado_contrato"] = estado_contrato
        if sector is not None:
            campos_actualizar["sector"] = sector
        if nombre_resumido_proceso is not None:
            campos_actualizar["nombre_resumido_proceso"] = nombre_resumido_proceso
        
        # Validar que se proporcionó al menos un campo para actualizar
        if not campos_actualizar:
            return JSONResponse(
                content={
                    "success": False,
                    "error": "Debe proporcionar al menos un campo para actualizar",
                    "message": "No se proporcionaron campos para modificar",
                    "timestamp": datetime.now().isoformat()
                },
                status_code=400,
                headers={"Content-Type": "application/json; charset=utf-8"}
            )
        
        # Modificar convenio de transferencia
        resultado = await modificar_convenio_transferencia(doc_id, campos_actualizar)
        
        # Manejar respuesta según el resultado
        if not resultado.get("success"):
            status_code = 404 if "No se encontró" in resultado.get("error", "") else 400
            return JSONResponse(
                content={
                    "success": False,
                    "error": resultado.get("error"),
                    "doc_id": doc_id,
                    "message": "Error al modificar el convenio de transferencia",
                    "timestamp": datetime.now().isoformat()
                },
                status_code=status_code,
                headers={"Content-Type": "application/json; charset=utf-8"}
            )
        
        # Respuesta exitosa
        return JSONResponse(
            content={
                "success": True,
                "message": resultado.get("message"),
                "doc_id": resultado.get("doc_id"),
                "campos_actualizados": resultado.get("campos_actualizados"),
                "data": resultado.get("data"),
                "coleccion": resultado.get("coleccion"),
                "timestamp": datetime.now().isoformat()
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de modificación de convenio de transferencia: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Por favor, inténtelo de nuevo más tarde",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.post("/emprestito/cargar-rpc", tags=["Gestión de Empréstito"], summary="🟢 Cargar RPC de Empréstito")
async def cargar_rpc_emprestito_endpoint(
    numero_rpc: str = Form(..., description="Número del RPC (obligatorio)"),
    beneficiario_id: str = Form(..., description="ID del beneficiario (obligatorio)"),
    beneficiario_nombre: str = Form(..., description="Nombre del beneficiario (obligatorio)"),
    descripcion_rpc: str = Form(..., description="Descripción del RPC (obligatorio)"),
    fecha_contabilizacion: str = Form(..., description="Fecha de contabilización (obligatorio)"),
    fecha_impresion: str = Form(..., description="Fecha de impresión (obligatorio)"),
    estado_liberacion: str = Form(..., description="Estado de liberación (obligatorio)"),
    bp: str = Form(..., description="Código BP (obligatorio)"),
    valor_rpc: float = Form(..., description="Valor del RPC (obligatorio)"),
    nombre_centro_gestor: str = Form(..., description="Centro gestor responsable (obligatorio)"),
    referencia_contrato: str = Form(..., description="Referencia del contrato (obligatorio)"),
    cdp_asociados: Optional[str] = Form(None, description="CDPs asociados separados por comas o JSON array (opcional)"),
    programacion_pac: Optional[str] = Form(None, description="Programación PAC en formato JSON (opcional)"),
    documentos: List[UploadFile] = File(..., description="Documentos del RPC (PDF, DOC, DOCX, XLS, XLSX, JPG, PNG) - OBLIGATORIO")
):
    """
    ## 📝 POST | 📥 Carga de Datos | Cargar RPC (Registro Presupuestal de Compromiso) de Empréstito
    
    Endpoint para carga directa de RPC de empréstito en la colección 
    `rpc_contratos_emprestito` sin procesamiento de APIs externas.
    
    ### ✅ Funcionalidades principales:
    - **Carga directa**: Registra directamente en `rpc_contratos_emprestito`
    - **Validación de duplicados**: Verifica existencia previa usando `numero_rpc`
    - **Validación de campos**: Verifica que todos los campos obligatorios estén presentes
    - **Carga de documentos a S3**: Los documentos son OBLIGATORIOS y se suben a AWS S3
    - **Validación de tipos de archivo**: Valida formatos permitidos (PDF, DOC, DOCX, XLS, XLSX, JPG, PNG)
    - **Timestamps automáticos**: Agrega fecha de creación y actualización
    - **Programación PAC**: Soporte para objeto JSON con valores mensuales
    
    ### ⚙️ Campos obligatorios:
    - `numero_rpc`: Número único del RPC
    - `beneficiario_id`: Identificación del beneficiario
    - `beneficiario_nombre`: Nombre completo del beneficiario
    - `descripcion_rpc`: Descripción del compromiso
    - `fecha_contabilizacion`: Fecha de contabilización del RPC
    - `fecha_impresion`: Fecha de impresión del documento
    - `estado_liberacion`: Estado de liberación del RPC
    - `bp`: Código BP (Banco de Programas)
    - `valor_rpc`: Valor monetario del RPC
    - `nombre_centro_gestor`: Centro gestor responsable
    - `referencia_contrato`: Referencia del contrato asociado
    - `documentos`: Archivos del RPC (al menos 1 archivo requerido)
    
    ### 📝 Campos opcionales:
    - `cdp_asociados`: Lista de CDPs (Certificados de Disponibilidad Presupuestal) asociados
      - Puede enviarse como: `"CDP-001,CDP-002,CDP-003"` (separados por comas)
      - O como JSON array: `["CDP-001", "CDP-002", "CDP-003"]`
      - Si se deja vacío, se guardará como lista vacía `[]`
    - `programacion_pac`: Objeto JSON con programación mensual del PAC (Plan Anual de Caja)
      - Formato: `{"enero-2024": "1000000", "febrero-2024": "500000"}`
      - **IMPORTANTE**: Debe ser un objeto JSON válido si se proporciona
      - Si no es JSON válido, se ignorará y se guardará como objeto vacío `{}`
    
    ### 🛡️ Validación de duplicados:
    Busca `numero_rpc` en la colección `rpc_contratos_emprestito` antes de crear nuevo registro.
    
    ### 📊 Estructura de datos guardados:
    ```json
    {
        "numero_rpc": "RPC-2024-001",
        "beneficiario_id": "890123456",
        "beneficiario_nombre": "Proveedor XYZ S.A.S.",
        "descripcion_rpc": "Suministro de equipos médicos",
        "fecha_contabilizacion": "2024-10-15",
        "fecha_impresion": "2024-10-16",
        "estado_liberacion": "Liberado",
        "bp": "BP-2024-001",
        "valor_rpc": 50000000.0,
        "cdp_asociados": ["CDP-2024-100", "CDP-2024-101", "CDP-2024-102"],
        "programacion_pac": {
            "enero-2024": "10000000",
            "febrero-2024": "20000000",
            "marzo-2024": "20000000"
        },
        "nombre_centro_gestor": "Secretaría de Salud",
        "referencia_contrato": "CONT-SALUD-003-2024",
        "fecha_creacion": "2024-10-14T10:30:00",
        "fecha_actualizacion": "2024-10-14T10:30:00",
        "estado": "activo",
        "tipo": "rpc_manual"
    }
    ```
    
    ### 📋 Ejemplo de request:
    ```json
    {
        "numero_rpc": "RPC-SALUD-003-2024",
        "beneficiario_id": "890123456",
        "beneficiario_nombre": "Proveedor XYZ S.A.S.",
        "descripcion_rpc": "Suministro de equipos médicos",
        "fecha_contabilizacion": "2024-10-15",
        "fecha_impresion": "2024-10-16",
        "estado_liberacion": "Liberado",
        "bp": "BP-2024-001",
        "valor_rpc": 50000000.0,
        "nombre_centro_gestor": "Secretaría de Salud",
        "referencia_contrato": "CONT-SALUD-003-2024",
        "cdp_asociados": "CDP-2024-100",
        "programacion_pac": "{\\"enero-2024\\": \\"10000000\\", \\"febrero-2024\\": \\"20000000\\"}"
    }
    ```
    
    ### ✅ Respuesta exitosa (201):
    ```json
    {
        "success": true,
        "message": "RPC RPC-SALUD-003-2024 guardado exitosamente",
        "doc_id": "abc123def456",
        "data": { ... },
        "coleccion": "rpc_contratos_emprestito"
    }
    ```
    
    ### ❌ Respuesta de duplicado (409):
    ```json
    {
        "success": false,
        "error": "Ya existe un RPC con número: RPC-SALUD-003-2024",
        "duplicate": true,
        "existing_data": { ... }
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        logger.info(f"📥 Recibiendo RPC: {numero_rpc}")
        logger.info(f"📎 Documentos recibidos: {len(documentos)}")
        
        # Validar que se hayan proporcionado documentos
        if not documentos or len(documentos) == 0:
            return JSONResponse(
                content={
                    "success": False,
                    "error": "Se requiere al menos un documento para cargar el RPC",
                    "message": "Debe proporcionar al menos un archivo PDF, DOC, DOCX, XLS, XLSX, JPG o PNG",
                    "timestamp": datetime.now().isoformat()
                },
                status_code=400,
                headers={"Content-Type": "application/json; charset=utf-8"}
            )
        
        # Validar tipos de archivo permitidos
        allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.jpg', '.jpeg', '.png']
        for doc in documentos:
            filename_lower = doc.filename.lower()
            if not any(filename_lower.endswith(ext) for ext in allowed_extensions):
                return JSONResponse(
                    content={
                        "success": False,
                        "error": f"Tipo de archivo no permitido: {doc.filename}",
                        "message": "Solo se permiten archivos PDF, DOC, DOCX, XLS, XLSX, JPG y PNG",
                        "allowed_types": allowed_extensions,
                        "timestamp": datetime.now().isoformat()
                    },
                    status_code=400,
                    headers={"Content-Type": "application/json; charset=utf-8"}
                )
            logger.info(f"   - {doc.filename} ({doc.content_type})")
        
        # Procesar cdp_asociados: puede venir como string separado por comas o como JSON array
        cdp_asociados_processed = None
        if cdp_asociados and cdp_asociados.strip():
            # Si parece JSON array, intentar parsear
            if cdp_asociados.strip().startswith('['):
                try:
                    cdp_parsed = json.loads(cdp_asociados)
                    if isinstance(cdp_parsed, list):
                        cdp_asociados_processed = cdp_parsed
                    else:
                        # Si no es lista, usar como string
                        cdp_asociados_processed = cdp_asociados
                except json.JSONDecodeError:
                    # Si falla el parseo, usar como string
                    cdp_asociados_processed = cdp_asociados
            else:
                # Si no parece JSON, asumir que es string separado por comas o simple
                cdp_asociados_processed = cdp_asociados
        
        # Procesar programacion_pac si viene como string JSON
        programacion_pac_dict = {}
        if programacion_pac and programacion_pac.strip():
            # Solo intentar parsear si parece ser JSON (empieza con { o [)
            if programacion_pac.strip().startswith('{') or programacion_pac.strip().startswith('['):
                try:
                    programacion_pac_dict = json.loads(programacion_pac)
                    if not isinstance(programacion_pac_dict, dict):
                        return JSONResponse(
                            content={
                                "success": False,
                                "error": "programacion_pac debe ser un objeto JSON (diccionario)",
                                "message": "El formato de programacion_pac debe ser un objeto JSON como {\"enero-2024\": \"1000000\"}",
                                "timestamp": datetime.now().isoformat()
                            },
                            status_code=400,
                            headers={"Content-Type": "application/json; charset=utf-8"}
                        )
                except json.JSONDecodeError as e:
                    return JSONResponse(
                        content={
                            "success": False,
                            "error": f"programacion_pac tiene formato JSON inválido: {str(e)}",
                            "message": "El formato de programacion_pac no es un JSON válido. Debe ser un objeto como {\"enero-2024\": \"1000000\"}",
                            "timestamp": datetime.now().isoformat()
                        },
                        status_code=400,
                        headers={"Content-Type": "application/json; charset=utf-8"}
                    )
            else:
                # Si no parece JSON, ignorar el campo con un warning
                logger.warning(f"programacion_pac no parece ser JSON, ignorando valor: {programacion_pac[:50]}")
                programacion_pac_dict = {}
        
        # Procesar documentos si se proporcionan
        documentos_procesados = []
        if documentos:
            for doc in documentos:
                # Leer contenido del archivo
                contenido = await doc.read()
                documentos_procesados.append({
                    'content': contenido,
                    'filename': doc.filename,
                    'content_type': doc.content_type,
                    'size': len(contenido)
                })
            logger.info(f"📄 Procesando {len(documentos_procesados)} documentos para RPC {numero_rpc}")
        
        # Crear diccionario con los datos del formulario
        datos_rpc = {
            "numero_rpc": numero_rpc,
            "beneficiario_id": beneficiario_id,
            "beneficiario_nombre": beneficiario_nombre,
            "descripcion_rpc": descripcion_rpc,
            "fecha_contabilizacion": fecha_contabilizacion,
            "fecha_impresion": fecha_impresion,
            "estado_liberacion": estado_liberacion,
            "bp": bp,
            "valor_rpc": valor_rpc,
            "cdp_asociados": cdp_asociados_processed,
            "programacion_pac": programacion_pac_dict,
            "nombre_centro_gestor": nombre_centro_gestor,
            "referencia_contrato": referencia_contrato
        }
        
        # Procesar RPC (función síncrona) con documentos
        logger.info(f"💾 Procesando RPC {numero_rpc} con {len(documentos_procesados)} documentos")
        resultado = cargar_rpc_emprestito(datos_rpc, documentos=documentos_procesados if documentos_procesados else None)
        
        # Log del resultado
        if resultado.get("success"):
            logger.info(f"✅ RPC {numero_rpc} procesado exitosamente")
        else:
            logger.error(f"❌ Error procesando RPC {numero_rpc}: {resultado.get('error')}")
        
        # Manejar respuesta según el resultado
        if not resultado.get("success"):
            # Manejar caso especial de duplicado
            if resultado.get("duplicate"):
                return JSONResponse(
                    content={
                        "success": False,
                        "error": resultado.get("error"),
                        "duplicate": True,
                        "existing_data": resultado.get("existing_data"),
                        "message": "Ya existe un RPC con este número",
                        "timestamp": datetime.now().isoformat()
                    },
                    status_code=409,  # Conflict
                    headers={"Content-Type": "application/json; charset=utf-8"}
                )
            else:
                # Error general
                return JSONResponse(
                    content={
                        "success": False,
                        "error": resultado.get("error"),
                        "message": "Error al procesar el RPC",
                        "timestamp": datetime.now().isoformat()
                    },
                    status_code=400,
                    headers={"Content-Type": "application/json; charset=utf-8"}
                )
        
        # Respuesta exitosa
        # Extraer URLs de documentos del resultado
        documentos_urls = []
        if resultado.get("data") and resultado.get("data").get("documentos_s3"):
            documentos_urls = [doc.get("url") for doc in resultado.get("data").get("documentos_s3") if doc.get("url")]
        
        return JSONResponse(
            content={
                "success": True,
                "message": resultado.get("message"),
                "data": {
                    "numero_rpc": numero_rpc,
                    "doc_id": resultado.get("doc_id"),
                    "documentos_urls": documentos_urls,
                    "total_documentos": resultado.get("documentos_count", 0),
                    "detalles_completos": resultado.get("data")
                },
                "coleccion": resultado.get("coleccion"),
                "timestamp": datetime.now().isoformat()
            },
            status_code=201,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de RPC: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Por favor, inténtelo de nuevo más tarde",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.post("/emprestito/cargar-pago", tags=["Gestión de Empréstito"], summary="🟢 Cargar Pago de Empréstito")
async def cargar_pago_emprestito_endpoint(
    numero_rpc: str = Form(..., description="Número del RPC (obligatorio)"),
    valor_pago: float = Form(..., description="Valor del pago (obligatorio, debe ser mayor a 0)"),
    fecha_transaccion: str = Form(..., description="Fecha de la transacción (obligatorio)"),
    referencia_contrato: str = Form(..., description="Referencia del contrato (obligatorio)"),
    nombre_centro_gestor: str = Form(..., description="Centro gestor responsable (obligatorio)"),
    documentos: List[UploadFile] = File(None, description="Documentos del pago (PDF, DOC, DOCX, XLS, XLSX, JPG, PNG) - OPCIONAL")
):
    """
    ## 📝 POST | 📥 Carga de Datos | Cargar Pago de Empréstito
    
    Endpoint para registrar un pago de empréstito en la colección `pagos_emprestito`.
    El campo `fecha_registro` se genera automáticamente con la hora actual del sistema como timestamp.
    
    ### ✅ Funcionalidades principales:
    - **Registro de pagos**: Guarda información de pagos realizados
    - **Carga de documentos a S3**: Los documentos son OPCIONALES y se suben a AWS S3 si se proporcionan
    - **Validación de tipos de archivo**: Valida formatos permitidos (PDF, DOC, DOCX, XLS, XLSX, JPG, PNG)
    - **Timestamp automático**: `fecha_registro` se genera automáticamente con la hora del sistema
    - **Validación de campos**: Verifica que todos los campos obligatorios estén presentes
    - **Validación de valores**: Verifica que el valor del pago sea positivo
    - **Trazabilidad**: Registra fecha de creación y actualización
    
    ### ⚙️ Campos obligatorios:
    - `numero_rpc`: Número del RPC asociado al pago
    - `valor_pago`: Valor monetario del pago (debe ser mayor a 0)
    - `fecha_transaccion`: Fecha en que se realizó la transacción
    - `referencia_contrato`: Referencia del contrato asociado
    - `nombre_centro_gestor`: Centro gestor responsable del pago
    
    ### ⚙️ Campos opcionales:
    - `documentos`: Archivos del pago (PDF, DOC, DOCX, XLS, XLSX, JPG, PNG)
    
    ### 🤖 Campos automáticos:
    - `fecha_registro`: Timestamp automático del momento de registro (NO se envía por el usuario)
    - `fecha_creacion`: Timestamp de creación del registro
    - `fecha_actualizacion`: Timestamp de última actualización
    - `estado`: "registrado" (valor por defecto)
    - `tipo`: "pago_manual" (valor por defecto)
    
    ### 📊 Estructura de datos guardados:
    ```json
    {
        "numero_rpc": "RPC-2024-001",
        "valor_pago": 10000000.0,
        "fecha_transaccion": "2024-11-11",
        "referencia_contrato": "CONT-SALUD-003-2024",
        "nombre_centro_gestor": "Secretaría de Salud",
        "fecha_registro": "2024-11-11T14:30:45.123456",
        "fecha_creacion": "2024-11-11T14:30:45.123456",
        "fecha_actualizacion": "2024-11-11T14:30:45.123456",
        "estado": "registrado",
        "tipo": "pago_manual"
    }
    ```
    
    ### 📋 Ejemplo de request:
    ```json
    {
        "numero_rpc": "RPC-SALUD-003-2024",
        "valor_pago": 10000000.0,
        "fecha_transaccion": "2024-11-11",
        "referencia_contrato": "CONT-SALUD-003-2024",
        "nombre_centro_gestor": "Secretaría de Salud"
    }
    ```
    
    ### ✅ Respuesta exitosa (201):
    ```json
    {
        "success": true,
        "message": "Pago registrado exitosamente para RPC RPC-SALUD-003-2024",
        "data": { ... },
        "doc_id": "abc123def456",
        "coleccion": "pagos_emprestito",
        "timestamp": "2024-11-11T14:30:45.123456"
    }
    ```
    
    ### ❌ Respuesta de error (400):
    ```json
    {
        "success": false,
        "error": "El campo 'numero_rpc' es obligatorio",
        "message": "Error al procesar el pago",
        "timestamp": "2024-11-11T14:30:45.123456"
    }
    ```
    
    ### 💡 Notas importantes:
    - El campo `fecha_registro` NO debe ser enviado por el usuario
    - Se genera automáticamente con la hora exacta del servidor
    - El `valor_pago` debe ser un número positivo mayor a 0
    - Todos los campos de texto se limpian de espacios en blanco
    """
    try:
        check_emprestito_availability()
        
        logger.info(f"📥 Recibiendo pago para RPC: {numero_rpc}")
        logger.info(f"📎 Documentos recibidos: {len(documentos) if documentos else 0}")
        logger.info(f"💰 Valor del pago: {valor_pago}")
        
        # Validar tipos de archivo permitidos solo si se proporcionaron documentos
        allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.jpg', '.jpeg', '.png']
        if documentos:
            for doc in documentos:
                filename_lower = doc.filename.lower()
                if not any(filename_lower.endswith(ext) for ext in allowed_extensions):
                    return JSONResponse(
                        content={
                            "success": False,
                            "error": f"Tipo de archivo no permitido: {doc.filename}",
                            "message": "Solo se permiten archivos PDF, DOC, DOCX, XLS, XLSX, JPG y PNG",
                            "allowed_types": allowed_extensions,
                            "timestamp": datetime.now().isoformat()
                        },
                        status_code=400,
                        headers={"Content-Type": "application/json; charset=utf-8"}
                    )
                logger.info(f"   - {doc.filename} ({doc.content_type})")
        
        # Procesar documentos si se proporcionan
        documentos_procesados = []
        if documentos:
            for doc in documentos:
                # Leer contenido del archivo
                contenido = await doc.read()
                documentos_procesados.append({
                    'content': contenido,
                    'filename': doc.filename,
                    'content_type': doc.content_type,
                    'size': len(contenido)
                })
            logger.info(f"📄 Procesando {len(documentos_procesados)} documentos para pago de RPC {numero_rpc}")
        
        # Preparar datos para procesar
        datos_pago = {
            "numero_rpc": numero_rpc,
            "valor_pago": valor_pago,
            "fecha_transaccion": fecha_transaccion,
            "referencia_contrato": referencia_contrato,
            "nombre_centro_gestor": nombre_centro_gestor
        }
        
        # Procesar pago (función síncrona) con documentos
        logger.info(f"💾 Procesando pago para RPC {numero_rpc} con {len(documentos_procesados)} documentos")
        resultado = cargar_pago_emprestito(datos_pago, documentos=documentos_procesados if documentos_procesados else None)
        
        # Log del resultado
        if resultado.get("success"):
            logger.info(f"✅ Pago para RPC {numero_rpc} procesado exitosamente")
        else:
            logger.error(f"❌ Error procesando pago para RPC {numero_rpc}: {resultado.get('error')}")
        
        # Manejar respuesta según el resultado
        if not resultado.get("success"):
            return JSONResponse(
                content={
                    "success": False,
                    "error": resultado.get("error"),
                    "message": "Error al procesar el pago",
                    "timestamp": datetime.now().isoformat()
                },
                status_code=400,
                headers={"Content-Type": "application/json; charset=utf-8"}
            )
        
        # Respuesta exitosa
        # Extraer URLs de documentos del resultado
        documentos_urls = []
        if resultado.get("data") and resultado.get("data").get("documentos_s3"):
            documentos_urls = [doc.get("url") for doc in resultado.get("data").get("documentos_s3") if doc.get("url")]
        
        return JSONResponse(
            content={
                "success": True,
                "message": resultado.get("message"),
                "data": {
                    "numero_rpc": numero_rpc,
                    "doc_id": resultado.get("doc_id"),
                    "valor_pago": valor_pago,
                    "documentos_urls": documentos_urls,
                    "total_documentos": resultado.get("documentos_count", 0),
                    "detalles_completos": resultado.get("data")
                },
                "coleccion": resultado.get("coleccion"),
                "timestamp": resultado.get("timestamp")
            },
            status_code=201,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de pago de empréstito: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Por favor, inténtelo de nuevo más tarde",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.get("/contratos_pagos_all", tags=["Gestión de Empréstito"], summary="🔵 Obtener Todos los Pagos")
async def get_all_pagos_emprestito():
    """
    ## 🔵 GET | 📋 Consultas | Obtener Todos los Pagos de Empréstito
    
    Endpoint para obtener todos los pagos de empréstito registrados en la colección `pagos_emprestito`.
    
    ### ✅ Funcionalidades principales:
    - **Listado completo**: Retorna todos los pagos registrados
    - **Datos completos**: Incluye todos los campos de cada pago
    - **Detección de documentos soporte**: Verifica si cada pago tiene documentos en S3
    - **Metadatos**: Incluye ID del documento, conteo total y timestamp
    - **Serialización JSON**: Fechas y objetos datetime convertidos correctamente
    - **Trazabilidad**: Información completa de cada transacción registrada
    
    ### 📊 Información incluida:
    - Todos los campos del pago
    - ID del documento para referencia
    - Campo `tiene_documentos_soporte`: indica si el pago tiene documentos en S3 (true/false)
    - Conteo total de registros
    - Timestamp de la consulta
    - Datos serializados correctamente para JSON
    
    ### 🗄️ Campos principales esperados:
    - **numero_rpc**: Número del RPC asociado al pago
    - **valor_pago**: Valor monetario del pago realizado
    - **fecha_transaccion**: Fecha en que se realizó la transacción
    - **referencia_contrato**: Referencia del contrato asociado
    - **nombre_centro_gestor**: Centro gestor responsable
    - **fecha_registro**: Timestamp automático del momento del registro
    - **fecha_creacion**: Fecha de creación del registro
    - **fecha_actualizacion**: Última actualización del registro
    - **estado**: Estado del pago (registrado, procesado, etc.)
    - **tipo**: Tipo de registro (pago_manual)
    - **tiene_documentos_soporte**: Boolean que indica si el pago tiene documentos en S3
    - **documentos_s3**: Array con información de documentos en S3 (si existen)
    
    ### 💡 Casos de uso:
    - Obtener historial completo de pagos de empréstito
    - Consulta de pagos para reportes financieros
    - Análisis de flujo de caja y ejecución presupuestal
    - Seguimiento de transacciones por RPC
    - Dashboard de pagos realizados
    - Exportación de datos para auditorías
    - Integración con sistemas contables
    - Reportes de ejecución por centro gestor
    
    ### 📈 Análisis posibles:
    - Total de pagos realizados
    - Suma de valores pagados
    - Pagos por centro gestor
    - Pagos por contrato
    - Pagos por RPC
    - Histórico de transacciones
    
    ### ✅ Respuesta exitosa (200):
    ```json
    {
        "success": true,
        "data": [
            {
                "id": "xyz789",
                "numero_rpc": "RPC-2024-001",
                "valor_pago": 10000000.0,
                "fecha_transaccion": "2024-11-11",
                "referencia_contrato": "CONT-SALUD-003-2024",
                "nombre_centro_gestor": "Secretaría de Salud",
                "fecha_registro": "2024-11-11T14:30:45.123456",
                "fecha_creacion": "2024-11-11T14:30:45.123456",
                "fecha_actualizacion": "2024-11-11T14:30:45.123456",
                "estado": "registrado",
                "tipo": "pago_manual",
                "tiene_documentos_soporte": true,
                "documentos_s3": [
                    {
                        "filename": "pago_001.pdf",
                        "s3_url": "https://contratos-emprestito.s3.us-east-1.amazonaws.com/...",
                        "upload_date": "2024-11-11T14:30:45.123456"
                    }
                ]
            },
            {
                "id": "abc456",
                "numero_rpc": "RPC-2024-002",
                "valor_pago": 5000000.0,
                "fecha_transaccion": "2024-11-10",
                "referencia_contrato": "CONT-INFRA-001-2024",
                "nombre_centro_gestor": "Secretaría de Infraestructura",
                "fecha_registro": "2024-11-10T10:15:30.654321",
                "fecha_creacion": "2024-11-10T10:15:30.654321",
                "fecha_actualizacion": "2024-11-10T10:15:30.654321",
                "estado": "registrado",
                "tipo": "pago_manual",
                "tiene_documentos_soporte": false,
                "documentos_s3": []
            }
        ],
        "count": 15,
        "collection": "pagos_emprestito",
        "timestamp": "2024-11-11T15:00:00.000000",
        "message": "Se obtuvieron 15 pagos exitosamente"
    }
    ```
    
    ### ❌ Respuesta de error (500):
    ```json
    {
        "success": false,
        "error": "Error obteniendo pagos de empréstito: [detalles del error]",
        "data": [],
        "count": 0
    }
    ```
    
    ### 📝 Notas:
    - Los campos de tipo datetime se serializan en formato ISO 8601
    - El campo `id` corresponde al ID del documento en Firestore
    - Los datos se retornan en el orden en que fueron insertados en Firestore
    - Para consultas filtradas, considere crear endpoints específicos adicionales
    """
    try:
        check_emprestito_availability()
        
        # Obtener todos los pagos
        resultado = await get_pagos_emprestito_all()
        
        if not resultado.get("success"):
            raise HTTPException(
                status_code=500,
                detail={
                    "success": False,
                    "error": resultado.get("error", "Error desconocido"),
                    "message": "Error al obtener los pagos de empréstito"
                }
            )
        
        # Respuesta exitosa
        return JSONResponse(
            content={
                "success": True,
                "data": resultado.get("data", []),
                "count": resultado.get("count", 0),
                "collection": resultado.get("collection", "pagos_emprestito"),
                "timestamp": resultado.get("timestamp"),
                "message": f"Se obtuvieron {resultado.get('count', 0)} pagos exitosamente"
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de consulta de pagos: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Por favor, inténtelo de nuevo más tarde",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.get("/rpc_all", tags=["Gestión de Empréstito"], summary="🔵 Obtener Todos los RPCs")
async def get_all_rpc_contratos_emprestito():
    """
    ## 🔵 GET | 📋 Consultas | Obtener Todos los RPCs de Empréstito
    
    Endpoint para obtener todos los RPC (Registros Presupuestales de Compromiso) de empréstito 
    almacenados en la colección `rpc_contratos_emprestito`.
    
    ### ✅ Funcionalidades principales:
    - **Listado completo**: Retorna todos los RPCs registrados
    - **Datos completos**: Incluye todos los campos de cada RPC
    - **Metadatos**: Incluye ID del documento, conteo total y timestamp
    - **Serialización JSON**: Fechas y objetos convertidos correctamente
    
    ### 📊 Información incluida:
    - Todos los campos del RPC
    - ID del documento para referencia
    - Conteo total de registros
    - Timestamp de la consulta
    - Datos serializados correctamente para JSON
    
    ### 🗄️ Campos principales esperados:
    - **numero_rpc**: Número único del RPC
    - **beneficiario_id**: Identificación del beneficiario
    - **beneficiario_nombre**: Nombre del beneficiario
    - **descripcion_rpc**: Descripción del compromiso
    - **fecha_contabilizacion**: Fecha de contabilización
    - **fecha_impresion**: Fecha de impresión del documento
    - **estado_liberacion**: Estado de liberación del RPC
    - **bp**: Código BP (Banco de Programas)
    - **valor_rpc**: Valor monetario del RPC
    - **cdp_asociados**: Lista de CDPs asociados
    - **programacion_pac**: Objeto con programación mensual del PAC
    - **nombre_centro_gestor**: Centro gestor responsable
    - **referencia_contrato**: Referencia del contrato asociado
    - **fecha_creacion**: Fecha de creación del registro
    - **fecha_actualizacion**: Última actualización
    - **estado**: Estado del registro (activo/inactivo)
    - **tipo**: Tipo de registro (rpc_manual)
    
    ### 💡 Casos de uso:
    - Obtener listado completo de RPCs de empréstito
    - Exportación de datos para análisis
    - Integración con sistemas externos
    - Reportes y dashboards de seguimiento presupuestal
    - Monitoreo de compromisos presupuestales
    - Análisis de ejecución presupuestal por contrato
    
    ### ✅ Respuesta exitosa (200):
    ```json
    {
        "success": true,
        "data": [
            {
                "id": "abc123",
                "numero_rpc": "RPC-2024-001",
                "beneficiario_id": "890123456",
                "beneficiario_nombre": "Proveedor XYZ S.A.S.",
                "descripcion_rpc": "Suministro de equipos médicos",
                "fecha_contabilizacion": "2024-10-15",
                "fecha_impresion": "2024-10-16",
                "estado_liberacion": "Liberado",
                "bp": "BP-2024-001",
                "valor_rpc": 50000000.0,
                "cdp_asociados": ["CDP-2024-100", "CDP-2024-101"],
                "programacion_pac": {
                    "enero-2024": "10000000",
                    "febrero-2024": "20000000"
                },
                "nombre_centro_gestor": "Secretaría de Salud",
                "referencia_contrato": "CONT-SALUD-003-2024",
                "fecha_creacion": "2024-10-14T10:30:00",
                "fecha_actualizacion": "2024-10-14T10:30:00",
                "estado": "activo",
                "tipo": "rpc_manual"
            }
        ],
        "count": 25,
        "collection": "rpc_contratos_emprestito",
        "timestamp": "2024-11-11T...",
        "message": "Se obtuvieron 25 RPCs exitosamente"
    }
    ```
    
    ### ❌ Respuesta de error (500):
    ```json
    {
        "success": false,
        "error": "Error obteniendo RPCs: ...",
        "data": [],
        "count": 0
    }
    ```
    
    ### 🔗 Endpoints relacionados:
    - `POST /emprestito/cargar-rpc` - Para crear nuevos RPCs
    - `GET /convenios_transferencias_all` - Para consultar convenios de transferencia
    """
    try:
        check_emprestito_availability()
        
        # Obtener todos los RPCs
        result = await get_rpc_contratos_emprestito_all()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo RPCs: {result.get('error', 'Error desconocido')}"
            )
        
        return JSONResponse(
            content={
                "success": True,
                "data": result["data"],
                "count": result["count"],
                "collection": result["collection"],
                "timestamp": result["timestamp"],
                "message": f"Se obtuvieron {result['count']} RPCs exitosamente"
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de RPCs: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Por favor, inténtelo de nuevo más tarde",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.get("/convenios_transferencias_all", tags=["Gestión de Empréstito"], summary="🔵 Obtener Todos los Convenios de Transferencia")
async def get_all_convenios_transferencia_emprestito():
    """
    ## 🔵 GET | 📋 Consultas | Obtener Todos los Convenios de Transferencia
    
    Endpoint para obtener todos los convenios de transferencia de empréstito 
    almacenados en la colección `convenios_transferencias_emprestito`.
    
    ### ✅ Funcionalidades principales:
    - **Listado completo**: Retorna todos los convenios registrados
    - **Ordenamiento**: Por fecha de creación (más recientes primero)
    - **Datos completos**: Incluye todos los campos de cada convenio
    - **Metadatos**: Incluye ID del documento, conteo total y timestamp
    
    ### 📊 Información incluida:
    - Todos los campos del convenio
    - ID del documento para referencia
    - Conteo total de registros
    - Timestamp de la consulta
    - Datos serializados correctamente para JSON
    
    ### 🗄️ Campos principales esperados:
    - **referencia_contrato**: Referencia única del contrato/convenio
    - **nombre_centro_gestor**: Centro gestor responsable
    - **banco**: Nombre del banco
    - **bp**: Código BP
    - **bpin**: Código BPIN
    - **objeto_contrato**: Descripción del objeto del contrato
    - **valor_contrato**: Valor del contrato
    - **valor_convenio**: Valor específico del convenio
    - **fecha_inicio_contrato**: Fecha de inicio
    - **fecha_fin_contrato**: Fecha de finalización
    - **modalidad_contrato**: Modalidad de contratación
    - **ordenador_gastor**: Ordenador del gasto
    - **tipo_contrato**: Tipo de contrato
    - **estado_contrato**: Estado actual
    - **sector**: Sector al que pertenece
    - **nombre_resumido_proceso**: Nombre resumido del proceso
    - **fecha_creacion**: Fecha de creación del registro
    - **fecha_actualizacion**: Última actualización
    - **estado**: Estado del registro (activo/inactivo)
    - **tipo**: Tipo de registro
    
    ### 💡 Casos de uso:
    - Obtener listado completo de convenios de transferencia
    - Exportación de datos para análisis
    - Integración con sistemas externos
    - Reportes y dashboards
    - Monitoreo del estado de convenios
    
    ### ✅ Respuesta exitosa (200):
    ```json
    {
        "success": true,
        "data": [
            {
                "id": "abc123",
                "referencia_contrato": "CONV-2024-001",
                "nombre_centro_gestor": "Secretaría de Salud",
                "banco": "Banco Mundial",
                "objeto_contrato": "Convenio de transferencia...",
                "valor_contrato": 1500000000.0,
                "bpin": "2024000010001",
                ...
            }
        ],
        "count": 15,
        "collection": "convenios_transferencias_emprestito",
        "timestamp": "2024-11-09T...",
        "message": "Se obtuvieron 15 convenios de transferencia exitosamente"
    }
    ```
    
    ### ❌ Respuesta de error (500):
    ```json
    {
        "success": false,
        "error": "Error obteniendo convenios de transferencia: ...",
        "data": [],
        "count": 0
    }
    ```
    
    ### 🔗 Endpoints relacionados:
    - `POST /emprestito/cargar-convenio-transferencia` - Para crear nuevos convenios
    - `GET /bancos_emprestito_all` - Para consultar bancos disponibles
    """
    try:
        check_emprestito_availability()
        
        # Obtener todos los convenios de transferencia
        result = await get_convenios_transferencia_emprestito_all()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo convenios de transferencia: {result.get('error', 'Error desconocido')}"
            )
        
        return JSONResponse(
            content={
                "success": True,
                "data": result["data"],
                "count": result["count"],
                "collection": result["collection"],
                "timestamp": result["timestamp"],
                "message": result["message"],
                "metadata": {
                    "sorted_by": "fecha_creacion",
                    "order": "desc",
                    "utf8_enabled": True,
                    "spanish_support": True,
                    "purpose": "Lista completa de convenios de transferencia de empréstito"
                }
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de convenios de transferencia: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error al obtener convenios de transferencia",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.get("/pagos_emprestito_all", tags=["Gestión de Empréstito"], summary="🔵 Obtener Todos los Pagos de Empréstito")
async def get_all_pagos_emprestito():
    """
    ## 🔵 GET | 📋 Consultas | Obtener Todos los Pagos de Empréstito
    
    Endpoint para obtener todos los pagos de empréstito almacenados en la colección `pagos_emprestito`.
    
    ### ✅ Funcionalidades principales:
    - **Listado completo**: Retorna todos los pagos registrados
    - **Datos completos**: Incluye todos los campos de cada pago
    - **Metadatos**: Incluye ID del documento, conteo total y timestamp
    
    ### 📊 Información incluida:
    - Todos los campos del pago
    - ID del documento para referencia
    - Conteo total de registros
    - Timestamp de la consulta
    
    ### ✅ Respuesta exitosa (200):
    ```json
    {
        "success": true,
        "data": [...],
        "count": 10,
        "collection": "pagos_emprestito",
        "timestamp": "2024-11-17T..."
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        result = await get_pagos_emprestito_all()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo pagos de empréstito: {result.get('error', 'Error desconocido')}"
            )
        
        return JSONResponse(
            content={
                "success": True,
                "data": result["data"],
                "count": result["count"],
                "collection": result["collection"],
                "timestamp": result["timestamp"],
                "message": f"Se obtuvieron {result['count']} pagos de empréstito exitosamente"
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de pagos de empréstito: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error al obtener pagos de empréstito",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.get("/rpc_contratos_emprestito_all", tags=["Gestión de Empréstito"], summary="🔵 Obtener Todos los RPCs de Empréstito")
async def get_all_rpc_contratos_emprestito():
    """
    ## 🔵 GET | 📋 Consultas | Obtener Todos los RPCs de Empréstito
    
    Endpoint para obtener todos los Registros Presupuestales de Compromiso (RPC) 
    de empréstito almacenados en la colección `rpc_contratos_emprestito`.
    
    ### ✅ Funcionalidades principales:
    - **Listado completo**: Retorna todos los RPCs registrados
    - **Datos completos**: Incluye todos los campos de cada RPC
    - **Metadatos**: Incluye ID del documento, conteo total y timestamp
    
    ### 📊 Información incluida:
    - Todos los campos del RPC
    - ID del documento para referencia
    - Conteo total de registros
    - Timestamp de la consulta
    
    ### 🗄️ Campos principales esperados:
    - **numero_rpc**: Número único del RPC
    - **beneficiario_id**: Identificación del beneficiario
    - **beneficiario_nombre**: Nombre del beneficiario
    - **descripcion_rpc**: Descripción del compromiso
    - **fecha_contabilizacion**: Fecha de contabilización
    - **fecha_impresion**: Fecha de impresión
    - **estado_liberacion**: Estado de liberación
    - **bp**: Código BP
    - **valor_rpc**: Valor monetario del RPC
    - **nombre_centro_gestor**: Centro gestor responsable
    - **referencia_contrato**: Referencia del contrato asociado
    - **cdp_asociados**: CDPs asociados
    - **programacion_pac**: Programación PAC
    
    ### ✅ Respuesta exitosa (200):
    ```json
    {
        "success": true,
        "data": [...],
        "count": 15,
        "collection": "rpc_contratos_emprestito",
        "timestamp": "2024-11-17T..."
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        result = await get_rpc_contratos_emprestito_all()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo RPCs de empréstito: {result.get('error', 'Error desconocido')}"
            )
        
        return JSONResponse(
            content={
                "success": True,
                "data": result["data"],
                "count": result["count"],
                "collection": result["collection"],
                "timestamp": result["timestamp"],
                "message": f"Se obtuvieron {result['count']} RPCs de empréstito exitosamente"
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de RPCs de empréstito: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error al obtener RPCs de empréstito",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )

@app.get("/emprestito/proceso/{referencia_proceso}", tags=["Gestión de Empréstito"], summary="🔵 Verificar Proceso Existente")
async def verificar_proceso_existente_endpoint(referencia_proceso: str):
    """
    ## � GET | �🔍 Consultas | Verificar Proceso Existente
    
    Verifica si ya existe un proceso con la referencia especificada en cualquiera 
    de las colecciones de empréstito.
    
    ### ✅ Funcionalidades:
    - Búsqueda en `procesos_emprestito` (SECOP)
    - Búsqueda en `ordenes_compra_emprestito` (TVEC)
    - Información detallada del proceso encontrado
    
    ### 📊 Respuesta si existe:
    - Datos completos del proceso
    - Colección donde se encontró
    - ID del documento
    
    ### 💡 Casos de uso:
    - Validación previa antes de crear proceso
    - Búsqueda de procesos existentes
    - Prevención de duplicados
    
    ### 📝 Ejemplo de respuesta (proceso existente):
    ```json
    {
        "existe": true,
        "coleccion": "procesos_emprestito",
        "documento": { ... },
        "doc_id": "xyz123",
        "timestamp": "2025-10-06T..."
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        resultado = await verificar_proceso_existente(referencia_proceso)
        
        return JSONResponse(
            content={
                **resultado,
                "referencia_proceso": referencia_proceso,
                "timestamp": datetime.now().isoformat()
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error verificando proceso: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error verificando proceso existente"
            }
        )


@app.delete("/emprestito/proceso/{referencia_proceso}", tags=["Gestión de Empréstito"], summary="🔴 Eliminar Proceso")
async def eliminar_proceso_emprestito_endpoint(referencia_proceso: str):
    """
    ## � DELETE | �🗑️ Eliminación | Eliminar Proceso de Empréstito
    
    Elimina un proceso de empréstito específico basado en su referencia_proceso.
    Busca automáticamente en ambas colecciones (SECOP y TVEC) y elimina el proceso encontrado.
    
    ### ✅ Funcionalidades principales:
    - **Búsqueda automática**: Localiza el proceso en ambas colecciones
    - **Eliminación segura**: Elimina únicamente el proceso especificado
    - **Información completa**: Retorna detalles del proceso eliminado
    - **Validación previa**: Verifica existencia antes de intentar eliminar
    
    ### 🔍 Colecciones de búsqueda:
    - **procesos_emprestito** (SECOP)
    - **ordenes_compra_emprestito** (TVEC)
    
    ### ⚠️ Consideraciones importantes:
    - La eliminación es **irreversible**
    - Solo se elimina un proceso por referencia_proceso
    - Se requiere coincidencia exacta en referencia_proceso
    
    ### 📋 Respuesta exitosa:
    ```json
    {
        "success": true,
        "message": "Proceso eliminado exitosamente",
        "referencia_proceso": "SCMGSU-CM-003-2024",
        "coleccion": "procesos_emprestito",
        "documento_id": "xyz123",
        "proceso_eliminado": {
            "referencia_proceso": "SCMGSU-CM-003-2024",
            "nombre_centro_gestor": "Secretaría de Salud",
            "nombre_banco": "Banco Mundial",
            "plataforma": "SECOP II",
            "fecha_creacion": "2025-10-06T..."
        },
        "timestamp": "2025-10-06T..."
    }
    ```
    
    ### 📋 Respuesta si no existe:
    ```json
    {
        "success": false,
        "error": "No se encontró ningún proceso con referencia_proceso: REFERENCIA",
        "referencia_proceso": "REFERENCIA",
        "colecciones_buscadas": ["procesos_emprestito", "ordenes_compra_emprestito"]
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        # Validar parámetro
        if not referencia_proceso or not referencia_proceso.strip():
            raise HTTPException(
                status_code=400,
                detail={
                    "success": False,
                    "error": "referencia_proceso es requerida",
                    "message": "Debe proporcionar una referencia_proceso válida"
                }
            )
        
        # Eliminar proceso
        resultado = await eliminar_proceso_emprestito(referencia_proceso.strip())
        
        # Manejar respuesta según el resultado
        if not resultado.get("success"):
            # Si no se encontró el proceso
            if "No se encontró" in resultado.get("error", ""):
                raise HTTPException(
                    status_code=404,
                    detail=resultado
                )
            else:
                # Otros errores
                raise HTTPException(
                    status_code=500,
                    detail=resultado
                )
        
        # Respuesta exitosa
        return JSONResponse(
            content=resultado,
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint eliminar proceso: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error eliminando proceso de empréstito",
                "referencia_proceso": referencia_proceso
            }
        )


@app.put("/emprestito/proceso/{referencia_proceso}", tags=["Gestión de Empréstito"], summary="🟡 Actualizar Proceso")
async def actualizar_proceso_emprestito_endpoint(
    referencia_proceso: str,
    bp: Optional[str] = Form(None, description="Código BP (opcional)"),
    nombre_resumido_proceso: Optional[str] = Form(None, description="Nombre resumido del proceso (opcional)"),
    id_paa: Optional[str] = Form(None, description="ID PAA (opcional)"),
    valor_proyectado: Optional[float] = Form(None, description="Valor proyectado (opcional)")
):
    """
    ## 🟡 PUT | ✏️ Actualización | Actualizar Proceso de Empréstito
    
    Actualiza campos específicos de un proceso de empréstito existente sin crear registros nuevos.
    Solo se actualizan los campos proporcionados, manteniendo los demás valores sin cambios.
    
    ### ✅ Funcionalidades principales:
    - **Búsqueda automática**: Localiza el proceso en ambas colecciones
    - **Actualización selectiva**: Solo modifica los campos proporcionados
    - **Preservación de datos**: Mantiene los campos no especificados
    - **Historial de cambios**: Muestra valores anteriores y nuevos
    
    ### 🔍 Colecciones de búsqueda:
    - **procesos_emprestito** (SECOP)
    - **ordenes_compra_emprestito** (TVEC)
    
    ### 📝 Campos actualizables:
    - `bp`: Código BP
    - `nombre_resumido_proceso`: Nombre resumido del proceso
    - `id_paa`: ID PAA
    - `valor_proyectado`: Valor proyectado (numérico)
    
    ### ⚙️ Comportamiento:
    - **Campos vacíos**: Se ignoran (no se actualizan)
    - **Campos con valor**: Se actualizan en la base de datos
    - **Timestamp**: Se actualiza automáticamente `fecha_actualizacion`
    - **Validación previa**: Verifica que el proceso existe
    
    ### 📋 Respuesta exitosa:
    ```json
    {
        "success": true,
        "message": "Proceso actualizado exitosamente",
        "referencia_proceso": "SCMGSU-CM-003-2024",
        "coleccion": "procesos_emprestito",
        "documento_id": "xyz123",
        "campos_modificados": ["bp", "valor_proyectado"],
        "valores_anteriores": {
            "bp": "BP-OLD-001",
            "valor_proyectado": 1000000.0
        },
        "valores_nuevos": {
            "bp": "BP-NEW-001",
            "valor_proyectado": 1500000.0
        },
        "proceso_actualizado": { ... },
        "timestamp": "2025-10-06T..."
    }
    ```
    
    ### 📋 Respuesta si no existe:
    ```json
    {
        "success": false,
        "error": "No se encontró ningún proceso con referencia_proceso: REFERENCIA",
        "referencia_proceso": "REFERENCIA",
        "colecciones_buscadas": ["procesos_emprestito", "ordenes_compra_emprestito"]
    }
    ```
    
    ### 📋 Respuesta sin campos:
    ```json
    {
        "success": false,
        "error": "No se proporcionaron campos para actualizar",
        "campos_disponibles": ["bp", "nombre_resumido_proceso", "id_paa", "valor_proyectado"]
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        # Validar parámetro
        if not referencia_proceso or not referencia_proceso.strip():
            raise HTTPException(
                status_code=400,
                detail={
                    "success": False,
                    "error": "referencia_proceso es requerida",
                    "message": "Debe proporcionar una referencia_proceso válida"
                }
            )
        
        # Actualizar proceso
        resultado = await actualizar_proceso_emprestito(
            referencia_proceso=referencia_proceso.strip(),
            bp=bp,
            nombre_resumido_proceso=nombre_resumido_proceso,
            id_paa=id_paa,
            valor_proyectado=valor_proyectado
        )
        
        # Manejar respuesta según el resultado
        if not resultado.get("success"):
            # Si no se encontró el proceso
            if "No se encontró" in resultado.get("error", ""):
                raise HTTPException(
                    status_code=404,
                    detail=resultado
                )
            # Si no se proporcionaron campos para actualizar
            elif "No se proporcionaron campos" in resultado.get("error", ""):
                raise HTTPException(
                    status_code=400,
                    detail=resultado
                )
            else:
                # Otros errores
                raise HTTPException(
                    status_code=500,
                    detail=resultado
                )
        
        # Respuesta exitosa
        return JSONResponse(
            content=resultado,
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint actualizar proceso: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error actualizando proceso de empréstito",
                "referencia_proceso": referencia_proceso
            }
        )


@app.post("/emprestito/obtener-contratos-secop", tags=["Gestión de Empréstito"], summary="🟢 Obtener Contratos SECOP")
async def obtener_contratos_secop_endpoint(offset: int = 0, limit: int = 10):
    """
    ## � POST | 🔄 Procesamiento por Lotes | Obtener Contratos de SECOP desde Procesos
    
    Procesa registros de la colección 'procesos_emprestito' en lotes, busca contratos en SECOP 
    para cada proceso y guarda los resultados en la nueva colección 'contratos_emprestito'.
    
    ### 📝 Parámetros opcionales:
    - **offset**: Índice inicial para procesar (default: 0)
    - **limit**: Cantidad de registros a procesar (default: 10, máximo: 50)
    
    ### 📤 Envío:
    ```http
    POST /emprestito/obtener-contratos-secop?offset=0&limit=10
    ```
    
    ### 🔄 Proceso:
    1. Leer registros de 'procesos_emprestito' desde offset hasta offset+limit
    2. Para cada proceso, extraer referencia_proceso y proceso_contractual
    3. Conectar con la API de SECOP (www.datos.gov.co) para cada proceso
    4. Buscar contratos que contengan el proceso_contractual y NIT = 890399011
    5. Transformar los datos al esquema de la colección 'contratos_emprestito'
    6. Verificar duplicados y actualizar/crear registros en Firebase
    7. Retornar resumen del lote procesado con información de paginación
    
    ### ✅ Respuesta exitosa:
    ```json
    {
        "success": true,
        "message": "Lote procesado: 10 procesos (offset 0-10)",
        "resumen_procesamiento": {
            "offset": 0,
            "limit": 10,
            "total_procesos_coleccion": 50,
            "procesos_en_lote": 10,
            "procesos_procesados": 9,
            "procesos_sin_contratos": 1,
            "procesos_con_errores": 0,
            "mas_registros": true,
            "siguiente_offset": 10
        },
        "criterios_busqueda": {
            "coleccion_origen": "procesos_emprestito",
            "filtro_secop": "nit_entidad = '890399011'"
        },
        "resultados_secop": {
            "total_contratos_encontrados": 12,
            "total_contratos_procesados": 12
        },
        "firebase_operacion": {
            "documentos_nuevos": 8,
            "documentos_actualizados": 3,
            "duplicados_ignorados": 1
        },
        "contratos_guardados": [
            {
                "referencia_proceso": "4151.010.32.1.0575-2025",
                "proceso_contractual": "CO1.REQ.8485621",
                "sector": "Educación",
                "referencia_contrato": "CONT-001-2025",
                "descripcion_proceso": "Descripción detallada del proceso contractual",
                "estado_contrato": "Activo",
                "valor_contrato": 150000000,
                "valor_pagado": "75000000",
                "representante_legal": "Juan Pérez García",
                "ordenador_gasto": "María López Silva",
                "supervisor": "Carlos Rodríguez Mesa",
                "fecha_firma_contrato": "2025-01-15",
                "entidad_contratante": "MUNICIPIO DE SANTIAGO DE CALI",
                "nombre_contratista": "EMPRESA XYZ LTDA",
                "nit_entidad": "890399011",
                "fuente_datos": "SECOP_API",
                "fecha_guardado": "2025-10-09T..."
            }
        ],
        "procesos_sin_contratos": [],
        "procesos_con_errores": [],
        "timestamp": "2025-10-09T..."
    }
    ```
    
    ### 📋 Respuesta sin procesos:
    ```json
    {
        "success": false,
        "error": "No se encontraron procesos en la colección procesos_emprestito",
        "timestamp": "2025-10-09T..."
    }
    ```
    
    ### 🗄️ Esquema de la colección 'contratos_emprestito':
    **🔄 Campos heredados desde procesos_emprestito:**
    - **referencia_proceso**: Heredado desde procesos_emprestito
    - **banco**: Heredado desde 'nombre_banco' de procesos_emprestito
    - **bp**: Heredado desde procesos_emprestito
    - **nombre_centro_gestor**: Heredado desde procesos_emprestito
    
    **📊 Campos desde SECOP API:**
    - **referencia_contrato**: referencia_del_contrato desde SECOP
    - **id_contrato**: Desde SECOP
    - **proceso_contractual**: Mapeado desde 'proceso_de_compra' de SECOP (sobrescribe el heredado)
    - **sector**: Desde SECOP
    - **nombre_procedimiento**: Mapeado desde 'nombre_del_procedimiento' de SECOP
    - **descripcion_proceso**: Mapeado desde 'descripcion_del_proceso' de SECOP
    - **estado_contrato**: Mapeado desde 'estado_contrato' de SECOP
    - **valor_contrato**: Desde SECOP (campo único, sin duplicados)
    - **valor_pagado**: Desde SECOP
    - **representante_legal**: Mapeado desde 'nombre_representante_legal' de SECOP
    - **ordenador_gasto**: Mapeado desde 'nombre_ordenador_del_gasto' de SECOP
    - **supervisor**: Mapeado desde 'nombre_supervisor' de SECOP
    - **bpin**: Mapeado desde 'c_digo_bpin' de SECOP
    - **fecha_firma_contrato**: Desde SECOP
    - **objeto_contrato**: Desde SECOP
    - **modalidad_contratacion**: Desde SECOP
    - **entidad_contratante**: Desde SECOP
    - **nombre_contratista**: Mapeado desde 'nombre_del_contratista' de SECOP
    - **nit_entidad**: Desde SECOP (filtrado por 890399011)
    - **nit_contratista**: Desde SECOP
    
    **🔧 Metadatos:**
    - **fecha_guardado**: Timestamp de cuando se guardó en Firebase
    - **fuente_datos**: "SECOP_API"
    - **version_esquema**: "1.1"
    
    ### 🔗 Integración SECOP:
    - **API**: www.datos.gov.co
    - **Dataset**: jbjy-vk9h (Contratos)
    - **Filtros**: proceso_de_compra LIKE '%{proceso_contractual}%' AND nit_entidad = '890399011'
    - **Mapeo**: proceso_de_compra → proceso_contractual (sobrescribe valor heredado)
    - **Nuevos campos**: sector desde SECOP
    - **Límite**: 2000 registros por consulta
    """
    try:
        check_emprestito_availability()
        
        # Validar límites
        if limit > 50:
            limit = 50
        if limit < 1:
            limit = 10
        if offset < 0:
            offset = 0
        
        # Ejecutar procesamiento por lotes
        resultado = await obtener_contratos_desde_proceso_contractual(offset=offset, limit=limit)
        
        # Retornar resultado
        return JSONResponse(
            content=resultado,
            status_code=200 if resultado.get("success") else 404,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint obtener contratos SECOP: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error obteniendo contratos de SECOP",
                "detalles": str(e)
            }
        )

@app.get("/contratos_emprestito_all", tags=["Gestión de Empréstito"], summary="🔵 Todos los Contratos Empréstito")
@optional_rate_limit("50/minute")  # Máximo 50 requests por minuto
@async_cache(ttl_seconds=300)  # Cache de 5 minutos
async def obtener_todos_contratos_emprestito(request: Request):
    """
    ## 🔵 GET | 📋 Listados | Obtener Todos los Contratos de Empréstito
    
    **Propósito**: Retorna todos los registros de las colecciones "contratos_emprestito", "ordenes_compra_emprestito" y "convenios_transferencias_emprestito".
    
    ### ✅ Casos de uso:
    - Obtener listado completo de contratos de empréstito
    - Exportación de datos para análisis
    - Integración con sistemas externos
    - Reportes y dashboards de contratos
    
    ### 📊 Información incluida:
    - Todos los campos disponibles en las tres colecciones
    - ID del documento para referencia
    - Conteo total de registros y por tipo
    - Timestamp de la consulta
    
    ### 🗄️ Colecciones incluidas:
    1. **contratos_emprestito**: Contratos principales
    2. **ordenes_compra_emprestito**: Órdenes de compra
    3. **convenios_transferencias_emprestito**: Convenios de transferencia
    
    ### 🗄️ Campos principales:
    - **referencia_contrato**: Referencia del contrato
    - **referencia_proceso**: Proceso de origen
    - **nombre_centro_gestor**: Entidad responsable
    - **banco**: Entidad bancaria
    - **estado_contrato**: Estado actual del contrato
    - **valor_contrato**: Valor del contrato
    - **fecha_firma_contrato**: Fecha de firma
    - **objeto_contrato**: Descripción del objeto
    - **modalidad_contratacion**: Modalidad de contratación
    - **entidad_contratante**: Entidad que contrata
    - **contratista**: Empresa contratista
    - **nombre_resumido_proceso**: 🔄 Heredado desde procesos_emprestito
    - **tipo_registro**: Identificador del tipo de registro (convenio_transferencia, contrato, orden)
    
    ### 🔄 Campos heredados desde procesos_emprestito:
    - **nombre_resumido_proceso**: Nombre resumido del proceso obtenido automáticamente usando referencia_proceso
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const response = await fetch('/contratos_emprestito_all');
    const data = await response.json();
    if (data.success) {
        console.log('Total de registros:', data.count);
        console.log('Contratos:', data.contratos_count);
        console.log('Órdenes:', data.ordenes_count);
        console.log('Convenios:', data.convenios_count);
        console.log('Datos:', data.data);
    }
    ```
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        result = await get_contratos_emprestito_all()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo contratos de empréstito: {result.get('error', 'Error desconocido')}"
            )
        
        return create_utf8_response({
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "contratos_count": result["contratos_count"],
            "ordenes_count": result["ordenes_count"],
            "convenios_count": result.get("convenios_count", 0),
            "collections": result["collections"],
            "timestamp": datetime.now().isoformat(),
            "last_updated": "2025-10-10T00:00:00Z",
            "message": result["message"]
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando contratos de empréstito: {str(e)}"
        )

@app.get("/contratos_emprestito/referencia/{referencia_contrato}", tags=["Gestión de Empréstito"], summary="🔵 Contratos por Referencia")
async def obtener_contratos_por_referencia(referencia_contrato: str):
    """
    ## � GET | �🔍 Consultas | Obtener Contratos por Referencia
    
    **Propósito**: Retorna contratos de empréstito filtrados por referencia_contrato específica.
    
    ### ✅ Casos de uso:
    - Búsqueda de contratos por referencia específica
    - Consulta de detalles de contrato individual
    - Validación de existencia de contrato
    - Integración con sistemas de seguimiento contractual
    
    ### 🔍 Filtrado:
    - **Campo**: `referencia_contrato` (coincidencia exacta)
    - **Tipo**: String - Referencia única del contrato
    - **Sensible a mayúsculas**: Sí
    
    ### 📊 Información incluida:
    - Todos los campos del contrato que coincida con la referencia
    - ID del documento para referencia
    - Conteo de registros encontrados
    - Información del filtro aplicado
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const referencia = "CONT-001-2025";
    const response = await fetch(`/contratos_emprestito/${referencia}`);
    const data = await response.json();
    if (data.success && data.count > 0) {
        console.log('Contrato encontrado:', data.data[0]);
    } else {
        console.log('No se encontró contrato con referencia:', referencia);
    }
    ```
    
    ### 💡 Notas:
    - Si no se encuentra ningún contrato, retorna array vacío
    - La referencia debe ser exacta (sin espacios adicionales)
    - Puede retornar múltiples contratos si hay duplicados
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        result = await get_contratos_emprestito_by_referencia(referencia_contrato)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo contratos por referencia: {result.get('error', 'Error desconocido')}"
            )
        
        return create_utf8_response({
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "filter": result["filter"],
            "timestamp": datetime.now().isoformat(),
            "last_updated": "2025-10-10T00:00:00Z",
            "message": result["message"]
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta por referencia de contrato: {str(e)}"
        )

@app.get("/contratos_emprestito/centro-gestor/{nombre_centro_gestor}", tags=["Gestión de Empréstito"])
async def obtener_contratos_por_centro_gestor(nombre_centro_gestor: str):
    """
    ## 🏢 Obtener Contratos de Empréstito por Centro Gestor
    
    **Propósito**: Retorna contratos de empréstito filtrados por nombre del centro gestor específico.
    
    ### ✅ Casos de uso:
    - Consulta de contratos por dependencia responsable
    - Reportes por entidad gestora
    - Dashboard por centro de responsabilidad
    - Análisis de distribución institucional
    - Seguimiento de contratos por secretaría/departamento
    
    ### 🔍 Filtrado:
    - **Campo**: `nombre_centro_gestor` (coincidencia exacta)
    - **Tipo**: String - Nombre completo del centro gestor
    - **Sensible a mayúsculas**: Sí
    - **Espacios**: Sensible a espacios adicionales
    
    ### 📊 Información incluida:
    - Todos los campos de los contratos del centro gestor
    - ID del documento para referencia
    - Conteo de registros encontrados
    - Información del filtro aplicado
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const centroGestor = "Secretaría de Salud";
    const response = await fetch(`/contratos_emprestito/${encodeURIComponent(centroGestor)}`);
    const data = await response.json();
    if (data.success && data.count > 0) {
        console.log(`${data.count} contratos encontrados para:`, centroGestor);
        const valorTotal = data.data.reduce((sum, c) => sum + (parseFloat(c.valor_contrato) || 0), 0);
        console.log('Valor total:', valorTotal);
    }
    ```
    
    ### 💡 Notas:
    - Típicamente retorna múltiples contratos por centro gestor
    - El nombre debe ser exacto (use `/centros-gestores/nombres-unicos` para obtener nombres válidos)
    - Para nombres con espacios, usar `encodeURIComponent()` en el frontend
    - Si no se encuentra ningún contrato, retorna array vacío
    
    ### 🔗 Endpoint relacionado:
    - `GET /centros-gestores/nombres-unicos` - Para obtener lista de centros gestores válidos
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        result = await get_contratos_emprestito_by_centro_gestor(nombre_centro_gestor)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo contratos por centro gestor: {result.get('error', 'Error desconocido')}"
            )
        
        return create_utf8_response({
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "filter": result["filter"],
            "timestamp": datetime.now().isoformat(),
            "last_updated": "2025-10-10T00:00:00Z",
            "message": result["message"]
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta por centro gestor: {str(e)}"
        )

@app.get("/emprestito/ordenes-compra", tags=["Gestión de Empréstito"])
async def get_ordenes_compra_todas():
    """
    ## 📋 Consultar Todas las Órdenes de Compra Existentes
    
    **Propósito**: Obtiene todas las órdenes de compra almacenadas en la colección 
    `ordenes_compra_emprestito` para revisar los datos disponibles.
    
    ### ✅ Información que proporciona:
    - **Listado completo**: Todas las órdenes de compra existentes
    - **Campos disponibles**: Estructura de datos actual
    - **Números de orden**: Para debugging del matching con TVEC
    """
    try:
        from api.scripts.ordenes_compra_operations import get_ordenes_compra_emprestito_all
        resultado = await get_ordenes_compra_emprestito_all()
        return resultado
        
    except Exception as e:
        logger.error(f"❌ Error consultando órdenes: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error consultando órdenes: {str(e)}"
        )

@app.post("/emprestito/obtener-ordenes-compra-TVEC", tags=["Gestión de Empréstito"])
async def obtener_ordenes_compra_tvec_endpoint():
    """
    ## 🛒 Obtener y Enriquecer Órdenes de Compra con Datos de TVEC
    
    **Propósito**: Enriquece todas las órdenes de compra existentes en la colección 
    `ordenes_compra_emprestito` con datos adicionales de la API de TVEC.
    
    ### ✅ Funcionalidades principales:
    - **Enriquecimiento de datos**: Obtiene datos adicionales de TVEC usando `numero_orden`
    - **Conservación de campos**: Mantiene todos los campos existentes en la colección
    - **Datos adicionales**: Agrega campos con prefijo `tvec_` para datos de la tienda virtual
    - **API Integration**: Usa la API oficial de datos abiertos de Colombia (rgxm-mmea)
    
    ### 📝 No requiere parámetros:
    Este endpoint procesa automáticamente todas las órdenes existentes en `ordenes_compra_emprestito`.
    
    ### 📤 Envío:
    ```http
    POST /emprestito/obtener-ordenes-compra-TVEC
    ```
    **No es necesario enviar ningún cuerpo JSON**.
    
    ### 🔄 Proceso:
    1. Obtener todas las órdenes de la colección `ordenes_compra_emprestito`
    2. Conectar con la API de TVEC (www.datos.gov.co/rgxm-mmea)
    3. Para cada orden, buscar datos adicionales usando `numero_orden`
    4. Enriquecer órdenes con campos adicionales con prefijo `tvec_`
    5. Actualizar registros en Firebase conservando campos originales
    6. Retornar resumen completo del enriquecimiento
    
    ### 📊 Campos adicionales agregados (estructura similar a contratos):
    
    **Campos principales (estructura estándar):**
    - `referencia_orden`: Referencia de la orden (similar a referencia_contrato)
    - `id_orden`: Identificador único de la orden (similar a id_contrato)
    - `estado_orden`: Estado de la orden (similar a estado_contrato)
    - `modalidad_contratacion`: Modalidad de la compra (mapeado desde tipo_compra)
    - `tipo_orden`: Tipo de compra (similar a tipo_contrato)
    - `fecha_publicacion_orden`: Fecha de publicación (similar a fecha_firma_contrato)
    - `fecha_vencimiento_orden`: Fecha de vencimiento (similar a fecha_fin_contrato)
    - `entidad_compradora`: Entidad que compra (similar a entidad_contratante)
    - `nombre_proveedor`: Nombre del proveedor (similar a nombre_contratista)
    - `nit_proveedor`: NIT del proveedor (similar a nit_contratista)
    - `descripcion_orden`: Descripción detallada (similar a descripcion_proceso)
    - `objeto_orden`: Objeto de la orden (similar a objeto_contrato)
    - `sector`: Sector/categoría principal
    - `valor_orden`: Valor total como número (similar a valor_contrato)
    - `_dataset_source`: "rgxm-mmea" (similar a "jbjy-vk9h" para contratos)
    - `fuente_datos`: "TVEC_API" (similar a "SECOP_API")
    - `fecha_guardado`: Timestamp de procesamiento
    - `version_esquema`: "1.0" (versión del esquema TVEC)
    
    **Campos específicos TVEC (con prefijo):**
    - `tvec_agregacion`: Tipo de agregación
    - `tvec_codigo_categoria`: Código de categoría
    - `tvec_unidad_medida`: Unidad de medida
    - `tvec_cantidad`: Cantidad
    - `tvec_precio_unitario`: Precio unitario
    
    ### 🔐 Snippet utilizado:
    El endpoint usa exactamente el snippet proporcionado:
    ```python
    import pandas as pd
    from sodapy import Socrata
    
    client = Socrata("www.datos.gov.co", None)
    results = client.get("rgxm-mmea", limit=2000)
    results_df = pd.DataFrame.from_records(results)
    ```
    
    ### ✅ Respuesta exitosa:
    ```json
    {
        "success": true,
        "message": "Enriquecimiento completado: 15/20 órdenes enriquecidas",
        "resumen": {
            "total_ordenes_procesadas": 20,
            "ordenes_enriquecidas": 15,
            "ordenes_sin_datos_tvec": 3,
            "ordenes_con_errores": 2,
            "tasa_enriquecimiento": "75.0%"
        },
        "fuente_datos": {
            "api_tvec": "www.datos.gov.co",
            "dataset": "rgxm-mmea",
            "registros_tvec_disponibles": 1850
        },
        "operacion_firebase": {
            "coleccion": "ordenes_compra_emprestito",
            "documentos_actualizados": 15,
            "campos_preservados": true,
            "campos_agregados_prefijo": "tvec_"
        },
        "ordenes_actualizadas": [
            {
                "doc_id": "abc123",
                "numero_orden": "OC-2024-001",
                "campos_agregados": [
                    "referencia_orden", "estado_orden", "valor_orden", 
                    "entidad_compradora", "nombre_proveedor", "nit_proveedor",
                    "descripcion_orden", "objeto_orden", "sector", "_dataset_source",
                    "fuente_datos", "fecha_guardado", "version_esquema"
                ],
                "datos_enriquecidos": {
                    "numero_orden": "OC-2024-001",
                    "referencia_orden": "OC-2024-001",
                    "estado_orden": "Activa",
                    "valor_orden": 1500000,
                    "entidad_compradora": "ALCALDIA DE SANTIAGO DE CALI",
                    "nombre_proveedor": "PROVEEDOR EJEMPLO S.A.S",
                    "nit_proveedor": "900123456-1",
                    "descripcion_orden": "Suministro de equipos tecnológicos",
                    "sector": "Tecnología",
                    "_dataset_source": "rgxm-mmea",
                    "fuente_datos": "TVEC_API",
                    "version_esquema": "1.0"
                }
            }
        ],
        "tiempo_total_segundos": 45.2,
        "timestamp": "2025-10-16T..."
    }
    ```
    
    ### 🚨 Requisitos:
    - Tener órdenes de compra registradas en `ordenes_compra_emprestito`
    - Cada orden debe tener el campo `numero_orden` 
    - Conexión a internet para acceder a la API de TVEC
    - Librerías: `sodapy` y `pandas` instaladas
    
    ### 💡 Características especiales:
    - **Preserva datos originales**: No modifica campos existentes
    - **Prefijo tvec_**: Evita conflictos con campos originales
    - **Matching por numero_orden**: Usa identificador único para relacionar datos
    - **Tolerante a errores**: Continúa procesando aunque algunas órdenes fallen
    - **Sin duplicados**: Solo agrega campos si no existen ya
    
    ### 🔗 Endpoints relacionados:
    - `POST /emprestito/cargar-orden-compra` - Para crear nuevas órdenes
    - `GET /ordenes_compra_emprestito_all` - Para consultar órdenes enriquecidas (si existe)
    """
    # Verificar disponibilidad de servicios
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    if not TVEC_ENRICH_OPERATIONS_AVAILABLE:
        raise HTTPException(
            status_code=503, 
            detail={
                "success": False,
                "error": "TVEC enrich operations not available",
                "message": "Las operaciones de enriquecimiento TVEC no están disponibles",
                "requirements": [
                    "pip install sodapy pandas",
                    "Verificar conectividad a internet",
                    "Confirmar acceso a www.datos.gov.co"
                ],
                "code": "TVEC_SERVICES_UNAVAILABLE"
            }
        )
    
    try:
        # Ejecutar enriquecimiento de órdenes de compra con datos de TVEC
        resultado = await obtener_ordenes_compra_tvec_enriquecidas()
        
        # Determinar código de estado basado en el resultado
        status_code = 200 if resultado.get("success") else 500
        
        # Retornar resultado con información detallada
        return JSONResponse(
            content={
                **resultado,
                "api_info": {
                    "endpoint_name": "obtener-ordenes-compra-TVEC",
                    "version": "1.0",
                    "snippet_based": True,
                    "preserves_original_data": True
                },
                "last_updated": "2025-10-16T00:00:00Z"
            },
            status_code=status_code,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint TVEC enriquecimiento: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error ejecutando enriquecimiento con datos de TVEC",
                "detalles": str(e),
                "code": "TVEC_INTERNAL_ERROR"
            }
        )

@app.get("/bancos_emprestito_all", tags=["Gestión de Empréstito"])
async def get_all_bancos_emprestito():
    """
    ## Obtener Todos los Bancos de Empréstito
    
    **Propósito**: Retorna todos los bancos disponibles en la colección "bancos_emprestito".
    
    ### ✅ Casos de uso:
    - Poblar dropdowns y selectores en formularios de empréstito
    - Obtener listado completo de bancos para validación
    - Integración con sistemas de gestión de procesos
    - Reportes y dashboards de bancos disponibles
    
    ### 📊 Información incluida:
    - Todos los campos disponibles de cada banco
    - ID del documento para referencia
    - Conteo total de bancos disponibles
    - Lista ordenada por nombre de banco
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const response = await fetch('/bancos_emprestito_all');
    const data = await response.json();
    if (data.success) {
        console.log('Bancos disponibles:', data.count);
        const bancoOptions = data.data.map(banco => ({
            value: banco.nombre_banco,
            label: banco.nombre_banco
        }));
    }
    ```
    
    ### 💡 Características:
    - **Ordenamiento**: Lista alfabética por nombre de banco
    - **Validación**: Datos limpios y serializados correctamente
    - **Compatibilidad**: UTF-8 completo para nombres con caracteres especiales
    - **Performance**: Optimizado para carga rápida de opciones
    
    ### 🔗 Endpoints relacionados:
    - `POST /emprestito/cargar-proceso` - Para crear nuevos procesos de empréstito usando estos bancos
    - `GET /contratos_emprestito_all` - Para consultar contratos por banco
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    if not EMPRESTITO_OPERATIONS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Emprestito operations not available")
    
    try:
        result = await get_bancos_emprestito_all()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo bancos de empréstito: {result.get('error', 'Error desconocido')}"
            )
        
        return create_utf8_response({
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "timestamp": result["timestamp"],
            "last_updated": "2025-10-11T00:00:00Z",  # Endpoint creation date
            "message": result["message"],
            "metadata": {
                "sorted": True,
                "utf8_enabled": True,
                "spanish_support": True,
                "purpose": "Banco selection for emprestito processes"
            }
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta de bancos: {str(e)}"
        )

@app.get("/procesos_emprestito_all", tags=["Gestión de Empréstito"])
@async_cache(ttl_seconds=300)  # Cache de 5 minutos
async def get_all_procesos_emprestito():
    """
    ## Obtener Todos los Procesos de Empréstito
    
    **Propósito**: Retorna todo el contenido de la colección "procesos_emprestito" en Firebase.
    
    ### ✅ Casos de uso:
    - Obtener listado completo de procesos de empréstito
    - Exportación de datos para análisis
    - Integración con sistemas externos
    - Reportes y dashboards de procesos
    - Monitoreo del estado de procesos
    
    ### 📊 Información incluida:
    - Todos los campos disponibles en la colección
    - ID del documento para referencia
    - Conteo total de registros
    - Timestamp de la consulta
    - Datos serializados correctamente para JSON
    
    ### 🗄️ Campos principales esperados:
    - **referencia_proceso**: Referencia única del proceso
    - **nombre_centro_gestor**: Entidad responsable
    - **nombre_banco**: Entidad bancaria
    - **plataforma**: SECOP, SECOP II, TVEC, etc.
    - **bp**: Código de proyecto base
    - **proceso_contractual**: Código del proceso contractual
    - **nombre_proceso**: Nombre del procedimiento
    - **estado_proceso**: Estado actual del proceso
    - **valor_publicacion**: Valor del proceso
    - **fecha_publicacion**: Fecha de publicación
    - **nombre_resumido_proceso**: Nombre resumido (opcional)
    - **id_paa**: ID del PAA (opcional)
    - **valor_proyectado**: Valor proyectado (opcional)
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const response = await fetch('/procesos_emprestito_all');
    const data = await response.json();
    if (data.success) {
        console.log('Procesos encontrados:', data.count);
        console.log('Datos:', data.data);
        
        // Filtrar por estado
        const activos = data.data.filter(p => p.estado_proceso === 'Activo');
        
        // Sumar valores
        const valorTotal = data.data.reduce((sum, p) => sum + (p.valor_publicacion || 0), 0);
    }
    ```
    
    ### 💡 Características:
    - **Serialización**: Datos de Firebase convertidos correctamente a JSON
    - **UTF-8**: Soporte completo para caracteres especiales
    - **Fechas**: Timestamps convertidos a formato ISO
    - **Performance**: Consulta optimizada de toda la colección
    - **Consistencia**: Estructura de datos uniforme
    
    ### 🔗 Endpoints relacionados:
    - `POST /emprestito/cargar-proceso` - Para crear nuevos procesos
    - `GET /contratos_emprestito_all` - Para consultar contratos relacionados
    - `GET /bancos_emprestito_all` - Para obtener bancos disponibles
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    if not EMPRESTITO_OPERATIONS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Emprestito operations not available")
    
    try:
        result = await get_procesos_emprestito_all()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo procesos de empréstito: {result.get('error', 'Error desconocido')}"
            )
        
        return create_utf8_response({
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "timestamp": result["timestamp"],
            "last_updated": "2025-10-18T00:00:00Z",  # Endpoint creation date
            "message": result["message"],
            "metadata": {
                "data_serialized": True,
                "utf8_enabled": True,
                "spanish_support": True,
                "firebase_timestamps_converted": True,
                "purpose": "Complete procesos_emprestito collection data"
            }
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta de procesos de empréstito: {str(e)}"
        )

@app.get("/ordenes_compra_emprestito/numero/{numero_orden}", tags=["Gestión de Empréstito"])
async def obtener_ordenes_por_numero(numero_orden: str):
    """
    ## 🔍 Obtener Órdenes de Compra por Número de Orden
    
    **Propósito**: Retorna órdenes de compra filtradas por número de orden específico.
    
    ### ✅ Casos de uso:
    - Búsqueda de órdenes por número específico
    - Consulta de detalles de orden individual
    - Validación de existencia de orden
    - Verificar datos enriquecidos de una orden específica
    
    ### 🔍 Filtrado:
    - **Campo**: `numero_orden` (coincidencia exacta)
    - **Tipo**: String - Número único de la orden
    - **Sensible a mayúsculas**: Sí
    
    ### 📊 Información incluida:
    - Todos los campos de las órdenes que coincidan con el número
    - Datos enriquecidos de TVEC (si están disponibles)
    - ID del documento para referencia
    - Información del filtro aplicado
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const numeroOrden = "OC-2024-001";
    const response = await fetch(`/ordenes_compra_emprestito/numero/${numeroOrden}`);
    const data = await response.json();
    if (data.success && data.count > 0) {
        const orden = data.data[0];
        console.log('Orden encontrada:', orden.numero_orden);
        if (orden._dataset_source === 'rgxm-mmea') {
            console.log('Orden enriquecida con TVEC:', orden.valor_orden);
        }
    }
    ```
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        result = await get_ordenes_compra_emprestito_by_referencia(numero_orden)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo órdenes por número: {result.get('error', 'Error desconocido')}"
            )
        
        return create_utf8_response({
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "filter": result["filter"],
            "timestamp": datetime.now().isoformat(),
            "last_updated": "2025-10-16T00:00:00Z",
            "message": result["message"]
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta por número de orden: {str(e)}"
        )

@app.get("/ordenes_compra_emprestito/centro-gestor/{nombre_centro_gestor}", tags=["Gestión de Empréstito"])
async def obtener_ordenes_por_centro_gestor(nombre_centro_gestor: str):
    """
    ## 🏢 Obtener Órdenes de Compra por Centro Gestor
    
    **Propósito**: Retorna órdenes de compra filtradas por nombre del centro gestor específico.
    
    ### ✅ Casos de uso:
    - Consulta de órdenes por dependencia responsable
    - Reportes por entidad gestora
    - Dashboard por centro de responsabilidad
    - Análisis de distribución institucional de órdenes de compra
    
    ### 🔍 Filtrado:
    - **Campo**: `nombre_centro_gestor` (coincidencia exacta)
    - **Tipo**: String - Nombre completo del centro gestor
    - **Sensible a mayúsculas**: Sí
    
    ### 📊 Información incluida:
    - Todas las órdenes del centro gestor especificado
    - Datos enriquecidos de TVEC (si están disponibles)
    - Conteo de registros encontrados
    - Información del filtro aplicado
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const centroGestor = "Secretaría de Salud";
    const response = await fetch(`/ordenes_compra_emprestito/centro-gestor/${encodeURIComponent(centroGestor)}`);
    const data = await response.json();
    if (data.success && data.count > 0) {
        console.log(`${data.count} órdenes encontradas para:`, centroGestor);
        const valorTotal = data.data.reduce((sum, o) => sum + (o.valor_orden || 0), 0);
        console.log('Valor total de órdenes:', valorTotal);
    }
    ```
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase or scripts not available")
    
    try:
        result = await get_ordenes_compra_emprestito_by_centro_gestor(nombre_centro_gestor)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo órdenes por centro gestor: {result.get('error', 'Error desconocido')}"
            )
        
        return create_utf8_response({
            "success": True,
            "data": result["data"],
            "count": result["count"],
            "collection": result["collection"],
            "filter": result["filter"],
            "timestamp": datetime.now().isoformat(),
            "last_updated": "2025-10-16T00:00:00Z",
            "message": result["message"]
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta por centro gestor: {str(e)}"
        )

@app.post("/emprestito/obtener-procesos-secop", tags=["Gestión de Empréstito"])
async def obtener_procesos_secop_completo_endpoint():
    """
    ## 🔄 Obtener y Actualizar Datos Completos de SECOP para Todos los Procesos
    
    Endpoint para complementar los datos de TODA la colección "procesos_emprestito" con información 
    adicional desde la API de SECOP, sin alterar los campos existentes ni los nombres de variables.
    
    ### ✅ Funcionalidades principales:
    - **Procesamiento masivo**: Actualiza TODOS los procesos de la colección automáticamente
    - **Actualización selectiva**: Solo actualiza campos que han cambiado por proceso
    - **Preservación de datos**: Mantiene todos los campos existentes intactos
    - **Mapeo desde SECOP**: Obtiene datos adicionales usando la API oficial
    - **Sin parámetros**: Lee automáticamente todas las referencias_proceso de Firebase
    
    ### 📊 Campos que se actualizan/complementan:
    **Campos básicos:**
    - `adjudicado` ← adjudicado (SECOP)
    - `fase` ← fase (SECOP)
    - `estado_proceso` ← estado_del_procedimiento (SECOP)
    
    **Campos adicionales agregados:**
    - `fecha_publicacion_fase` ← fecha_de_publicacion_del (SECOP)
    - `fecha_publicacion_fase_1` ← null (no disponible en SECOP)
    - `fecha_publicacion_fase_2` ← null (no disponible en SECOP)
    - `fecha_publicacion_fase_3` ← fecha_de_publicacion_fase_3 (SECOP)
    - `proveedores_invitados` ← proveedores_invitados (SECOP)
    - `proveedores_con_invitacion` ← proveedores_con_invitacion (SECOP)
    - `visualizaciones_proceso` ← visualizaciones_del (SECOP)
    - `proveedores_que_manifestaron` ← proveedores_que_manifestaron (SECOP)
    - `numero_lotes` ← numero_de_lotes (SECOP)
    - `fecha_adjudicacion` ← null (no disponible en SECOP)
    - `estado_resumen` ← estado_resumen (SECOP)
    - `fecha_recepcion_respuestas` ← null (no disponible en SECOP)
    - `fecha_apertura_respuestas` ← null (no disponible en SECOP)
    - `fecha_apertura_efectiva` ← null (no disponible en SECOP)
    - `respuestas_procedimiento` ← respuestas_al_procedimiento (SECOP)
    - `respuestas_externas` ← respuestas_externas (SECOP)
    - `conteo_respuestas_ofertas` ← conteo_de_respuestas_a_ofertas (SECOP)
    
    ### 🔐 Validaciones:
    - Verificar que el proceso existe en la colección `procesos_emprestito`
    - Conectar con API de SECOP usando la referencia_proceso
    - Solo actualizar si hay cambios reales en los datos
    - Mantener estructura de variables sin cambios
    
    ### 📝 Ejemplo de request:
    ```http
    POST /emprestito/obtener-procesos-secop
    ```
    **No requiere parámetros - procesamiento automático**
    
    ### ✅ Respuesta exitosa:
    ```json
    {
        "success": true,
        "message": "Se procesaron 5 procesos de empréstito exitosamente",
        "resumen_procesamiento": {
            "total_procesos_encontrados": 5,
            "procesos_procesados": 4,
            "procesos_actualizados": 3,
            "procesos_sin_cambios": 1,
            "procesos_con_errores": 1
        },
        "resultados_detallados": [
            {
                "referencia_proceso": "4163.001.32.1.718-2024",
                "success": true,
                "changes_count": 8,
                "changes_summary": [
                    "adjudicado: 'No' → 'Sí'",
                    "estado_proceso: 'En evaluación' → 'Seleccionado'"
                ]
            },
            {
                "referencia_proceso": "4164.001.32.1.719-2024",
                "success": true,
                "changes_count": 0,
                "message": "Ya está actualizado"
            }
        ],
        "estadisticas": {
            "total_campos_actualizados": 25,
            "tiempo_procesamiento": "45.2 segundos"
        },
        "timestamp": "2024-10-18T..."
    }
    ```
    
    ### 📋 Respuesta sin procesos:
    ```json
    {
        "success": false,
        "error": "No se encontraron procesos en la colección procesos_emprestito",
        "total_procesos_encontrados": 0
    }
    ```
    
    ### 🔍 API de SECOP utilizada:
    - **Dominio**: www.datos.gov.co
    - **Dataset**: p6dx-8zbt (Procesos de contratación)
    - **Filtro**: nit_entidad='890399011' AND referencia_del_proceso='{referencia_proceso}'
    
    ### ⏱️ Tiempo de procesamiento:
    - **Timeout extendido**: 5 minutos (300 segundos)
    - **Tiempo estimado**: ~10-15 segundos por proceso
    - **Progreso**: Se reporta en logs con ETA para procesos restantes
    - **Recomendación**: Monitor logs del servidor para ver progreso en tiempo real
    """
    try:
        check_emprestito_availability()
        
        # Procesar todos los procesos de empréstito automáticamente
        resultado = await procesar_todos_procesos_emprestito_completo()
        
        # Manejar respuesta según el resultado
        if not resultado.get("success"):
            # Si no se encontraron procesos
            if "No se encontraron procesos" in resultado.get("error", ""):
                raise HTTPException(
                    status_code=404,
                    detail=resultado
                )
            else:
                # Otros errores
                raise HTTPException(
                    status_code=500,
                    detail=resultado
                )
        
        # Respuesta exitosa
        return JSONResponse(
            content=resultado,
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint obtener procesos SECOP completo: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error obteniendo datos completos de SECOP para todos los procesos"
            }
        )


@app.get("/asignaciones-emprestito-banco-centro-gestor", tags=["Gestión de Empréstito"], summary="🔵 Obtener Asignaciones Banco-Centro Gestor")
async def get_all_asignaciones_emprestito_banco_centro_gestor():
    """
    ## 🔵 GET | 📋 Consultas | Obtener Todas las Asignaciones de Empréstito Banco-Centro Gestor
    
    Endpoint para obtener todas las asignaciones de montos de empréstito por banco y centro gestor
    almacenadas en la colección `montos_emprestito_asignados_centro_gestor`.
    
    ### ✅ Funcionalidades principales:
    - **Listado completo**: Retorna todas las asignaciones registradas
    - **Datos completos**: Incluye todos los campos de cada asignación
    - **Metadatos**: Incluye ID del documento, conteo total y timestamp
    
    ### 📊 Información incluida:
    - Todos los campos de la asignación
    - ID del documento para referencia
    - Conteo total de registros
    - Timestamp de la consulta
    
    ### 🗄️ Campos principales esperados:
    - **banco**: Nombre del banco financiador
    - **nombre_centro_gestor**: Nombre del centro gestor
    - **bp**: Código del proyecto presupuestal (BP)
    - **monto_programado**: Monto programado para el banco y centro gestor
    - **anio**: Año de la asignación
    - **created_at**: Fecha de creación del registro
    - **updated_at**: Fecha de última actualización
    - **data_hash**: Hash para control de duplicados
    
    ### ✅ Respuesta exitosa (200):
    ```json
    {
        "success": true,
        "data": [
            {
                "id": "BBVA_BP26004701_2026",
                "banco": "BBVA",
                "nombre_centro_gestor": "Secretaría de Educación",
                "bp": "BP26004701",
                "monto_programado": 1500000.00,
                "anio": 2026,
                "created_at": "2024-11-17T...",
                "updated_at": "2024-11-17T...",
                "data_hash": "abc123..."
            }
        ],
        "count": 83,
        "collection": "montos_emprestito_asignados_centro_gestor",
        "timestamp": "2024-11-17T...",
        "message": "Se obtuvieron 83 asignaciones de empréstito banco-centro gestor exitosamente"
    }
    ```
    """
    try:
        check_emprestito_availability()
        
        result = await get_asignaciones_emprestito_banco_centro_gestor_all()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo asignaciones de empréstito banco-centro gestor: {result.get('error', 'Error desconocido')}"
            )
        
        return JSONResponse(
            content={
                "success": True,
                "data": result["data"],
                "count": result["count"],
                "collection": result["collection"],
                "timestamp": result["timestamp"],
                "message": f"Se obtuvieron {result['count']} asignaciones de empréstito banco-centro gestor exitosamente"
            },
            status_code=200,
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint de asignaciones de empréstito banco-centro gestor: {e}")
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error": "Error interno del servidor",
                "message": "Error al obtener asignaciones de empréstito banco-centro gestor",
                "code": "INTERNAL_SERVER_ERROR"
            }
        )


# ============================================================================
# ENDPOINTS DE FLUJO DE CAJA EMPRÉSTITO
# ============================================================================

@app.post("/emprestito/flujo-caja/cargar-excel", tags=["Gestión de Empréstito"], summary="🟢 Cargar Flujos de Caja Excel")
async def cargar_flujo_caja_excel(
    archivo_excel: UploadFile = File(..., description="Archivo Excel con flujos de caja"),
    update_mode: str = Form(default="merge", description="Modo de actualización: merge, replace, append")
):
    """
    ## � POST | �📊 Carga de Archivos | Cargar Flujos de Caja desde Excel
    
    Endpoint para procesar archivos Excel con información de flujos de caja de proyectos
    y cargarlos en la colección "flujo_caja_emprestito".
    
    ### 📁 Archivo Excel esperado:
    - **Hoja**: "CONTRATOS - Seguimiento" 
    - **Columnas requeridas**: Responsable, Organismo, Banco, BP Proyecto, Descripcion BP
    - **Columnas de datos**: Todas las columnas que contengan "Desembolso" en su nombre
    - **Formato de fechas**: Las columnas de desembolso deben contener fechas como jul-25, ago-25, etc.
    
    ### 🔧 Modos de actualización:
    - **merge**: Actualiza existentes y crea nuevos (por defecto)
    - **replace**: Reemplaza toda la colección
    - **append**: Solo agrega nuevos registros
    
    ### 📊 Procesamiento:
    1. Lee datos del Excel
    2. Separa columnas de Desembolso normal y REAL
    3. Convierte a formato largo (un registro por mes)
    4. Crea campo Periodo en formato fecha
    5. Guarda en Firebase con ID único por organismo_banco_mes
    
    ### 🎯 Cómo usar:
    1. Selecciona archivo .xlsx con formato correcto
    2. Elige modo de actualización
    3. Haz clic en "Execute"
    
    ### ✅ Validaciones:
    - Solo archivos .xlsx
    - Columnas Organismo y Banco requeridas
    - Al menos una columna de Desembolso
    - Tamaño máximo: 10MB
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase o scripts no disponibles")
    
    if not FLUJO_CAJA_OPERATIONS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Operaciones de flujo de caja no disponibles")
    
    # Validar modo de actualización
    if update_mode not in ["merge", "replace", "append"]:
        raise HTTPException(status_code=400, detail="update_mode debe ser: merge, replace o append")
    
    # Validar tipo de archivo
    if not archivo_excel.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx, .xls)")
    
    # Validar tamaño del archivo (10MB máximo)
    max_size = 10 * 1024 * 1024  # 10MB
    file_content = await archivo_excel.read()
    if len(file_content) > max_size:
        raise HTTPException(status_code=400, detail="El archivo no puede exceder 10MB")
    
    try:
        # Procesar el archivo Excel
        result = process_flujo_caja_excel(file_content, archivo_excel.filename)
        
        if not result["success"]:
            raise HTTPException(status_code=400, detail=result.get('error', 'Error procesando Excel'))
        
        # Guardar en Firebase
        save_result = await save_flujo_caja_to_firebase(result["data"], update_mode)
        
        if not save_result["success"]:
            raise HTTPException(status_code=500, detail=save_result.get('error', 'Error guardando en Firebase'))
        
        # Combinar resultados
        final_result = {
            "success": True,
            "message": "Flujos de caja cargados exitosamente",
            "archivo_info": {
                "nombre_archivo": archivo_excel.filename,
                "tamaño_bytes": len(file_content),
                "modo_actualizacion": update_mode
            },
            "procesamiento": result["summary"],
            "guardado": save_result["summary"],
            "timestamp": datetime.now().isoformat(),
            "last_updated": "2025-10-20T00:00:00Z"
        }
        
        return create_utf8_response(final_result)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")

@app.get("/emprestito/flujo-caja/all", tags=["Gestión de Empréstito"], summary="🔵 Flujos de Caja")
async def get_flujos_caja_all(
    responsable: Optional[str] = Query(None, description="Filtrar por responsable específico"),
    organismo: Optional[str] = Query(None, description="Filtrar por organismo específico"),
    banco: Optional[str] = Query(None, description="Filtrar por banco específico"),
    bp_proyecto: Optional[str] = Query(None, description="Filtrar por BP Proyecto específico"),
    mes: Optional[str] = Query(None, description="Filtrar por mes específico (ej: jul-25)"),
    periodo_desde: Optional[str] = Query(None, description="Periodo desde (formato: YYYY-MM-DD)"),
    periodo_hasta: Optional[str] = Query(None, description="Periodo hasta (formato: YYYY-MM-DD)"),
    limit: Optional[int] = Query(None, ge=1, le=1000, description="Límite de registros")
):
    """
    ## � GET | �📊 Consultas con Filtros | Obtener Todos los Flujos de Caja
    
    Endpoint para consultar flujos de caja almacenados en la colección "flujo_caja_emprestito".
    
    ### ✅ Casos de uso:
    - Consultar flujos de caja por organismo o banco
    - Filtrar por períodos específicos
    - Analizar desembolsos planeados vs reales
    - Generar reportes de flujo de caja
    - Exportar datos para dashboards
    
    ### 🔍 Filtros disponibles:
    - **responsable**: Filtrar por responsable específico
    - **organismo**: Filtrar por organismo específico
    - **banco**: Filtrar por banco específico
    - **bp_proyecto**: Filtrar por BP Proyecto específico  
    - **mes**: Filtrar por mes específico (ej: "jul-25")
    - **periodo_desde**: Desde fecha específica (YYYY-MM-DD)
    - **periodo_hasta**: Hasta fecha específica (YYYY-MM-DD)
    - **limit**: Limitar número de resultados (máx: 1000)
    
    ### 📊 Información incluida:
    - Responsable, organismo, banco y BP proyecto
    - Descripción del BP proyecto
    - Mes y período en formato fecha
    - Monto de desembolso
    - Columna origen del Excel
    - ID único del registro y metadatos de archivo origen
    
    ### 📝 Ejemplo de uso:
    ```javascript
    // Obtener todos los flujos
    const response = await fetch('/emprestito/flujo-caja/all');
    
    // Filtrar por banco específico
    const response = await fetch('/emprestito/flujo-caja/all?banco=Banco Popular');
    
    // Filtrar por período
    const response = await fetch('/emprestito/flujo-caja/all?periodo_desde=2025-07-01&periodo_hasta=2025-12-31');
    ```
    
    ### 💡 Características:
    - **Ordenamiento**: Por período (cronológico)
    - **Resumen**: Estadísticas agregadas incluidas
    - **Metadatos**: Organismos, bancos y meses únicos
    - **UTF-8**: Soporte completo para caracteres especiales
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase o scripts no disponibles")
    
    if not FLUJO_CAJA_OPERATIONS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Operaciones de flujo de caja no disponibles")
    
    try:
        # Construir filtros
        filters = {}
        
        if responsable:
            filters['responsable'] = responsable
        if organismo:
            filters['organismo'] = organismo
        if banco:
            filters['banco'] = banco
        if bp_proyecto:
            filters['bp_proyecto'] = bp_proyecto
        if mes:
            filters['mes'] = mes
        if periodo_desde:
            filters['periodo_desde'] = periodo_desde
        if periodo_hasta:
            filters['periodo_hasta'] = periodo_hasta
        if limit:
            filters['limit'] = limit
        
        # Obtener datos de Firebase
        result = await get_flujo_caja_from_firebase(filters)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error obteniendo flujos de caja: {result.get('error', 'Error desconocido')}"
            )
        
        # Agregar información del endpoint
        result["last_updated"] = "2025-10-20T00:00:00Z"
        result["endpoint_info"] = {
            "filtros_aplicados": len([k for k, v in filters.items() if v is not None]),
            "total_filtros_disponibles": 6,
            "ordenamiento": "por_periodo_cronologico"
        }
        
        return create_utf8_response(result)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando consulta de flujos de caja: {str(e)}"
        )

@app.post("/emprestito/crear-tabla-proyecciones", tags=["Gestión de Empréstito"], summary="🟢 Crear Tabla Proyecciones")
async def crear_tabla_proyecciones_endpoint():
    """
    ## � POST | 🔗 Integración Externa | Crear Tabla de Proyecciones desde Google Sheets
    
    **Propósito**: Lee datos de Google Sheets específico y los carga en la colección "proyecciones_emprestito".
    
    ### 🔧 Proceso automático:
    1. **Lee datos** desde Google Sheets específico (Publicados Emprestitos nuevo)
    2. **Mapea campos** según especificaciones definidas
    3. **Procesa BP** agregando prefijo "BP" automáticamente
    4. **Guarda en Firebase** en colección "proyecciones_emprestito"
    5. **Elimina temporal** y registra fecha de actualización
    
    ### 📋 Mapeo de campos:
    - `Item` → `item`
    - `Nro de Proceso` → `referencia_proceso`
    - `NOMBRE ABREVIADO` → `nombre_organismo_reducido`
    - `Banco` → `nombre_banco`
    - `BP` → `BP` (con prefijo "BP" agregado)
    - `DESCRIPCION BP` → `descripcion_bp`
    - `Proyecto` → `nombre_generico_proyecto`
    - `Proyecto con su respectivo contrato` → `nombre_resumido_proceso`
    - `ID PAA` → `id_paa`
    - `LINK DEL PROCESO` → `urlProceso`
    - `valor_proyectado` → `valor_proyectado` (mapeo directo)
    
    **NOTA**: La columna en Google Sheets ahora se llama "valor_proyectado" directamente
    
    ### ✅ Características:
    - **Reemplazo completo**: Elimina datos existentes y carga nuevos
    - **Validación automática**: Verifica campos obligatorios
    - **Manejo de errores**: Reporta filas con problemas
    - **Metadatos**: Registra fecha de carga y estadísticas
    - **UTF-8**: Soporte completo para caracteres especiales
    - **URL fija**: Usa Google Sheets predefinido
    - **Service Account**: Autenticación con service account configurado
    
    ### 🔐 Autenticación:
    - **Service Account**: `unidad-cumplimiento-sheets@unidad-cumplimiento.iam.gserviceaccount.com`
    - **Permisos**: Debe tener acceso de lectura al Google Sheets configurado
    - **Scopes**: `spreadsheets.readonly` y `drive.readonly`
    - **Credenciales**: Configuradas en el sistema usando ADC o variable de entorno
    
    ### 📝 Ejemplo de respuesta:
    ```json
    {
        "success": true,
        "message": "Tabla de proyecciones creada exitosamente",
        "resumen_operacion": {
            "filas_leidas": 150,
            "registros_procesados": 148,
            "registros_guardados": 148,
            "docs_eliminados_previos": 145
        }
    }
    ```
    
    ### 💡 Notas importantes:
    - **URL fija**: Usa Google Sheets predefinido internamente
    - **Automático**: No requiere parámetros de entrada
    - **Destructivo**: Reemplaza todos los datos existentes
    - **Auditable**: Mantiene registro de fecha de última actualización
    - **Permisos**: Requiere service account con acceso al Google Sheets
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase o scripts no disponibles")
    
    if not EMPRESTITO_OPERATIONS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Operaciones de empréstito no disponibles")
    
    try:
        # URL fija del Google Sheets según especificación del usuario
        sheet_url = "https://docs.google.com/spreadsheets/d/11-sdLwINHHwRit8b9jnnXcO2phhuEVUpXM6q6yv8DYo/edit?usp=sharing"
        
        # Ejecutar proceso completo
        result = await crear_tabla_proyecciones_desde_sheets(sheet_url)
        
        if not result["success"]:
            # Verificar si es error de autorización para dar mejor mensaje
            error_msg = result.get('error', 'Error desconocido')
            
            if 'Unauthorized' in error_msg or '401' in error_msg:
                raise HTTPException(
                    status_code=403,
                    detail={
                        "error": "El Google Sheets no es público o no tiene permisos de lectura",
                        "solucion": "Para resolver este problema:",
                        "pasos": [
                            "1. Abrir el Google Sheets",
                            "2. Hacer clic en 'Compartir' (botón azul superior derecho)",
                            "3. En 'Obtener enlace', cambiar a 'Cualquier persona con el enlace'",
                            "4. Cambiar permisos a 'Lector'",
                            "5. Copiar el enlace y usarlo en el parámetro sheet_url"
                        ],
                        "error_original": error_msg
                    }
                )
            else:
                raise HTTPException(
                    status_code=500,
                    detail=f"Error creando tabla de proyecciones: {error_msg}"
                )
        
        # Agregar información del endpoint
        result["last_updated"] = "2025-10-22T00:00:00Z"
        result["endpoint_info"] = {
            "sheet_url_fija": True,
            "operacion": "reemplazo_completo",
            "campos_mapeados": 10,
            "validaciones": "campos_obligatorios",
            "service_account": "unidad-cumplimiento-sheets@unidad-cumplimiento.iam.gserviceaccount.com"
        }
        
        return create_utf8_response(result)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando creación de tabla de proyecciones: {str(e)}"
        )

@app.get("/emprestito/leer-tabla-proyecciones", tags=["Gestión de Empréstito"], summary="🔵 Tabla de Proyecciones")
async def leer_tabla_proyecciones_endpoint(
    sheet_url: Optional[str] = Query(
        None, 
        description="URL de Google Sheets para detectar registros con Nro de Proceso que NO están en procesos_emprestito."
    ),
    solo_no_guardados: bool = Query(
        False,
        description="Si es True y se proporciona sheet_url, devuelve solo registros que NO están en procesos_emprestito pero tienen Nro de Proceso válido"
    )
):
    """
    ## 📋 GET | 📋 Listados | Leer Tabla de Proyecciones de Empréstito
    
    **Propósito**: 
    - **Sin parámetros**: Obtiene todos los registros de la colección "proyecciones_emprestito".
    - **Con sheet_url**: Detecta registros de Google Sheets que NO están en procesos_emprestito.
    
    ### ✅ Casos de uso:
    
    #### Modo 1: Lectura de BD (sin parámetros)
    - Consultar proyecciones cargadas desde Google Sheets
    - Verificar datos después de carga
    - Exportar proyecciones para análisis
    - Integrar con dashboards y reportes
    - Auditar última fecha de actualización
    
    #### Modo 2: Detección de no guardados en procesos_emprestito (con sheet_url)
    - **Identifica registros pendientes**: Encuentra qué datos de Sheets tienen Nro de Proceso pero NO están en procesos_emprestito
    - **Validación de sincronización**: Verifica qué procesos faltan por crear en la BD
    - **Detección de pendientes**: Lista proyecciones que necesitan ser guardadas como procesos
    - **Control de calidad**: Asegura que todos los procesos válidos estén registrados
    
    ### 🔍 Condiciones para Modo 2 (Registros devueltos):
    1. ✅ Tienen valor en columna "Nro de Proceso" (no vacío, no null)
    2. ❌ El valor de "Nro de Proceso" NO existe en la colección `procesos_emprestito` con campo `referencia_proceso`
    
    ### 📊 Información incluida (Modo 1 - Sin sheet_url):
    - **Datos mapeados**: Todos los campos según mapeo definido
    - **Metadatos**: Fecha de carga, fuente, fila origen
    - **Timestamps**: Fecha de guardado y última actualización
    - **ID único**: Identificador de Firebase para cada registro
    - **Estadísticas**: Información de la última carga realizada
    
    ### 🔍 Información incluida (Modo 2 - Con sheet_url):
    - **Registros no guardados**: Solo los que tienen Nro de Proceso válido pero NO existen en procesos_emprestito
    - **Comparación precisa**: Verifica contra la colección procesos_emprestito
    - **Metadata de comparación**: Estadísticas sobre registros encontrados/no encontrados
    - **Optimización**: Usa mapas en memoria para comparación rápida O(1)
    
    ### 🔍 Campos de respuesta:
    - `item`: Número de ítem
    - `referencia_proceso`: Número de proceso (Nro de Proceso de Sheets)
    - `nombre_organismo_reducido`: Nombre abreviado del organismo
    - `nombre_banco`: Banco asociado
    - `BP`: Código BP con prefijo agregado
    - `descripcion_bp`: Descripción del BP
    - `nombre_generico_proyecto`: Nombre del proyecto
    - `nombre_resumido_proceso`: Proyecto con contrato
    - `id_paa`: ID del PAA
    - `urlProceso`: Enlace al proceso
    - `valor_proyectado`: Valor total del proyecto (única columna de valor)
    - `_es_nuevo`: (Solo Modo 2) Indica que es un registro no guardado
    - `_motivo`: (Solo Modo 2) Razón por la cual no está guardado
    
    **NOTA**: NO se incluyen campos duplicados como "VALOR TOTAL" o "Valor Adjudicado"
    
    ### 📝 Ejemplos de uso:
    
    #### Ejemplo 1: Leer todos los registros guardados en proyecciones_emprestito
    ```javascript
    const response = await fetch('/emprestito/leer-tabla-proyecciones');
    const data = await response.json();
    
    if (data.success) {
        console.log(`Proyecciones encontradas: ${data.count}`);
        data.data.forEach(proyeccion => {
            console.log(`${proyeccion.referencia_proceso}: ${proyeccion.valor_proyectado}`);
        });
    }
    ```
    
    #### Ejemplo 2: Detectar registros pendientes de guardar en procesos_emprestito
    ```javascript
    const sheetUrl = 'https://docs.google.com/spreadsheets/d/ABC123/edit';
    const response = await fetch(
        `/emprestito/leer-tabla-proyecciones?sheet_url=${encodeURIComponent(sheetUrl)}&solo_no_guardados=true`
    );
    const data = await response.json();
    
    if (data.success) {
        console.log(`Registros pendientes: ${data.count}`);
        console.log(`Total en Sheets: ${data.metadata.total_sheets}`);
        console.log(`Ya en procesos_emprestito: ${data.metadata.ya_en_procesos}`);
        console.log(`Sin Nro de Proceso: ${data.metadata.sin_proceso}`);
        
        // Procesar registros pendientes
        data.data.forEach(registro => {
            console.log(`Pendiente: ${registro.referencia_proceso} - ${registro._motivo}`);
        });
    }
    ```
    
    ### 💡 Características:
    - **Ordenamiento** (Modo 1): Por fecha de carga (más recientes primero)
    - **Filtrado inteligente** (Modo 2): Solo registros con Nro Proceso válido que NO están en procesos_emprestito
    - **Validación estricta**: Verifica que referencia_proceso no sea null, vacío o solo espacios
    - **UTF-8**: Soporte completo para caracteres especiales
    - **Auditoría**: Incluye información de trazabilidad
    - **Optimización**: Búsqueda O(1) usando sets en memoria
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase o scripts no disponibles")
    
    if not EMPRESTITO_OPERATIONS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Operaciones de empréstito no disponibles")
    
    try:
        # Modo 2: Comparar con Google Sheets y devolver no guardados en procesos_emprestito
        if sheet_url and solo_no_guardados:
            result = await leer_proyecciones_no_guardadas(sheet_url)
            
            if not result["success"]:
                raise HTTPException(
                    status_code=500,
                    detail=f"Error comparando con Google Sheets: {result.get('error', 'Error desconocido')}"
                )
            
            # Agregar información del endpoint
            result["last_updated"] = "2025-11-01T00:00:00Z"
            result["endpoint_info"] = {
                "modo": "deteccion_no_guardados",
                "sheet_url": sheet_url,
                "filtro": "no_en_procesos_emprestito_con_nro_proceso_valido",
                "coleccion_comparada": "procesos_emprestito",
                "campo_comparado": "referencia_proceso",
                "optimizado": True
            }
            
            return create_utf8_response(result)
        
        # Modo 1: Obtener proyecciones de Firebase (comportamiento original)
        result = await leer_proyecciones_emprestito()
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=f"Error leyendo tabla de proyecciones: {result.get('error', 'Error desconocido')}"
            )
        
        # Agregar información del endpoint
        result["last_updated"] = "2025-11-01T00:00:00Z"
        result["endpoint_info"] = {
            "modo": "lectura_bd",
            "coleccion_fuente": "proyecciones_emprestito",
            "ordenamiento": "por_fecha_carga_desc",
            "incluye_metadatos": True,
            "trazabilidad_completa": True
        }
        
        return create_utf8_response(result)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando lectura de tabla de proyecciones: {str(e)}"
        )


@app.get("/emprestito/proyecciones-sin-proceso", tags=["Gestión de Empréstito"])
async def endpoint_proyecciones_sin_proceso():
    """Devuelve proyecciones cuya 'referencia_proceso' no exista en 'procesos_emprestito'."""
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase o scripts no disponibles")

    if not EMPRESTITO_OPERATIONS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Operaciones de empréstito no disponibles")

    try:
        result = await get_proyecciones_sin_proceso()

        if not result.get("success", False):
            raise HTTPException(status_code=500, detail=result.get("error", "Error desconocido"))

        # Agregar metadata del endpoint
        result["last_updated"] = "2025-10-23T00:00:00Z"
        result["endpoint_info"] = {
            "coleccion_origen": "proyecciones_emprestito",
            "coleccion_comparacion": "procesos_emprestito",
            "filter_field": "referencia_proceso",
            "returned_count": result.get("count", 0)
        }

        return create_utf8_response(result)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando endpoint: {str(e)}")


@app.put("/emprestito/proyecciones/{referencia_proceso}", tags=["Gestión de Empréstito"], summary="🟡 Actualizar Proyección")
async def actualizar_proyeccion_emprestito_endpoint(
    referencia_proceso: str,
    datos_actualizacion: ProyeccionEmprestitoUpdateRequest
):
    """
    ## � PUT | ✏️ Actualización | Actualizar Proyección de Empréstito
    
    **Propósito**: Actualiza cualquier campo de un registro específico en la colección "proyecciones_emprestito" 
    según su "referencia_proceso".
    
    ### ✅ Casos de uso:
    - Actualizar datos específicos de una proyección existente
    - Corregir información incorrecta en proyecciones
    - Modificar valores proyectados o información del banco
    - Actualizar enlaces de procesos o información PAA
    - Mantener datos sincronizados con fuentes externas
    
    ### 🎯 Funcionamiento:
    1. **Busca** el registro por `referencia_proceso` (parámetro de ruta)
    2. **Actualiza** solo los campos proporcionados en el body
    3. **Mantiene** los campos no especificados sin cambios
    4. **Registra** timestamp de última actualización
    5. **Retorna** datos previos y actualizados para auditoría
    
    ### 📋 Campos actualizables:
    - `item`: Número de ítem
    - `nombre_organismo_reducido`: Nombre abreviado del organismo
    - `nombre_banco`: Banco asociado
    - `BP`: Código BP
    - `nombre_generico_proyecto`: Nombre del proyecto
    - `nombre_resumido_proceso`: Proyecto con contrato
    - `id_paa`: ID del PAA
    - `urlProceso`: Enlace al proceso
    - `valor_proyectado`: Valor total del proyecto
    
    ### 🔒 Validaciones:
    - **referencia_proceso**: Debe existir en la colección
    - **valor_proyectado**: Debe ser >= 0 si se proporciona
    - **strings**: Se limpian automáticamente de espacios
    - **campos opcionales**: Solo se actualizan los proporcionados
    
    ### 📝 Ejemplo de uso:
    ```javascript
    const referencia = "PROC-2024-001";
    const datosActualizar = {
        valor_proyectado: 500000000,
        nombre_banco: "Banco de Occidente",
        urlProceso: "https://nuevo-enlace.com"
    };
    
    const response = await fetch(`/emprestito/proyecciones/${referencia}`, {
        method: 'PUT',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(datosActualizar)
    });
    ```
    
    ### ✅ Respuesta exitosa:
    ```json
    {
        "success": true,
        "message": "Proyección actualizada exitosamente",
        "referencia_proceso": "PROC-2024-001",
        "doc_id": "abc123",
        "datos_previos": { ... },
        "datos_actualizados": { ... },
        "campos_modificados": ["valor_proyectado", "nombre_banco", "urlProceso"]
    }
    ```
    
    ### 💡 Características:
    - **Actualización parcial**: Solo modifica campos especificados
    - **Auditoría completa**: Guarda datos previos y nuevos
    - **Búsqueda exacta**: Por referencia_proceso únicamente
    - **UTF-8**: Soporte completo para caracteres especiales
    - **Timestamp automático**: Registra fecha de modificación
    - **Validación robusta**: Verifica existencia y tipos de datos
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase o scripts no disponibles")
    
    if not EMPRESTITO_OPERATIONS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Operaciones de empréstito no disponibles")
    
    try:
        # Convertir el modelo Pydantic a diccionario, excluyendo campos None
        datos_dict = datos_actualizacion.dict(exclude_none=True)
        
        # Verificar que se proporcionen al menos algunos datos para actualizar
        if not datos_dict:
            raise HTTPException(
                status_code=400,
                detail="Debe proporcionar al menos un campo para actualizar"
            )
        
        # Ejecutar actualización
        result = await actualizar_proyeccion_emprestito(referencia_proceso, datos_dict)
        
        if not result["success"]:
            # Manejo específico de errores
            if "No se encontró" in result.get('error', ''):
                raise HTTPException(
                    status_code=404,
                    detail=f"No se encontró proyección con referencia_proceso: {referencia_proceso}"
                )
            else:
                raise HTTPException(
                    status_code=500,
                    detail=f"Error actualizando proyección: {result.get('error', 'Error desconocido')}"
                )
        
        # Agregar información del endpoint
        result["last_updated"] = "2025-10-23T00:00:00Z"
        result["endpoint_info"] = {
            "metodo": "PUT",
            "operacion": "actualizacion_parcial",
            "campos_actualizables": [
                "item", "nombre_organismo_reducido", "nombre_banco", "BP",
                "nombre_generico_proyecto", "nombre_resumido_proceso", 
                "id_paa", "urlProceso", "valor_proyectado"
            ],
            "validaciones_aplicadas": True,
            "auditoria_completa": True
        }
        
        return create_utf8_response(result)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando actualización de proyección: {str(e)}"
        )


@app.post("/emprestito/registrar-proyeccion", tags=["Gestión de Empréstito"], summary="🟢 Registrar Nueva Proyección")
async def registrar_proyeccion_emprestito_endpoint(
    referencia_proceso: str = Form(..., description="Referencia única del proceso"),
    nombre_centro_gestor: str = Form(..., description="Nombre del centro gestor"),
    nombre_banco: str = Form(..., description="Nombre del banco"),
    bp: str = Form(..., description="Código BP", alias="BP"),
    proyecto_generico: str = Form(..., description="Proyecto genérico"),
    estado_proyeccion: Optional[str] = Form(None, description="Estado de la proyección"),
    nombre_resumido_proceso: Optional[str] = Form(None, description="Nombre resumido del proceso"),
    id_paa: Optional[str] = Form(None, description="ID del PAA"),
    valor_proyectado: Optional[float] = Form(None, ge=0, description="Valor proyectado (debe ser >= 0)"),
    urlProceso: Optional[str] = Form(None, description="URL del proceso")
):
    """
    ## 🟢 POST | ➕ Creación | Registrar Nueva Proyección de Empréstito
    
    **Propósito**: Crea un nuevo registro en la colección "proyecciones_emprestito" con todos los 
    campos necesarios para el seguimiento de proyecciones de empréstito.
    
    ### ✅ Casos de uso:
    - Registrar nuevas proyecciones de empréstito
    - Crear registros preliminares antes de la formalización
    - Documentar proyecciones en etapas tempranas
    - Vincular proyecciones con procesos PAA
    - Establecer valores proyectados para presupuestación
    
    ### ✅ Casos de uso:
    - Registrar nuevas proyecciones de empréstito
    - Crear registros preliminares antes de la formalización
    - Documentar proyecciones en etapas tempranas
    - Vincular proyecciones con procesos PAA
    - Establecer valores proyectados para presupuestación
    
    ### 🎯 Funcionamiento:
    1. **Valida** que no exista una proyección con la misma referencia_proceso
    2. **Verifica** que todos los campos requeridos estén presentes
    3. **Limpia** y normaliza los datos ingresados
    4. **Crea** el registro en Firebase con timestamp
    5. **Retorna** confirmación con ID del documento creado
    
    ### 📋 Campos del registro:
    
    #### Campos Requeridos:
    - `referencia_proceso`: Identificador único del proceso
    - `nombre_centro_gestor`: Nombre del centro gestor responsable
    - `nombre_banco`: Entidad bancaria asociada
    - `bp`: Código BP del proyecto
    - `proyecto_generico`: Nombre genérico del proyecto
    
    #### Campos Opcionales:
    - `estado_proyeccion`: Estado actual de la proyección
    - `nombre_resumido_proceso`: Nombre resumido para identificación
    - `id_paa`: Identificador del Plan Anual de Adquisiciones
    - `valor_proyectado`: Monto proyectado (debe ser >= 0)
    - `urlProceso`: URL del proceso en plataforma SECOP
    
    ### 🔒 Validaciones:
    - **referencia_proceso**: No debe existir previamente en la colección
    - **valor_proyectado**: Debe ser >= 0 si se proporciona
    - **strings**: Se limpian automáticamente de espacios
    - **campos requeridos**: Todos los marcados como obligatorios deben proporcionarse
    """
    if not FIREBASE_AVAILABLE or not SCRIPTS_AVAILABLE:
        raise HTTPException(status_code=503, detail="Firebase o scripts no disponibles")
    
    try:
        # Construir diccionario con los datos del formulario
        datos_dict = {
            'referencia_proceso': referencia_proceso.strip() if referencia_proceso else None,
            'nombre_centro_gestor': nombre_centro_gestor.strip() if nombre_centro_gestor else None,
            'nombre_banco': nombre_banco.strip() if nombre_banco else None,
            'BP': bp.strip() if bp else None,
            'proyecto_generico': proyecto_generico.strip() if proyecto_generico else None,
        }
        
        # Agregar campos opcionales solo si tienen valor
        if estado_proyeccion:
            datos_dict['estado_proyeccion'] = estado_proyeccion.strip()
        if nombre_resumido_proceso:
            datos_dict['nombre_resumido_proceso'] = nombre_resumido_proceso.strip()
        if id_paa:
            datos_dict['id_paa'] = id_paa.strip()
        if valor_proyectado is not None:
            if valor_proyectado < 0:
                raise HTTPException(
                    status_code=400,
                    detail="El valor_proyectado debe ser mayor o igual a 0"
                )
            datos_dict['valor_proyectado'] = valor_proyectado
        if urlProceso:
            datos_dict['urlProceso'] = urlProceso.strip()
        
        # Verificar que la referencia_proceso no exista ya
        db = firestore.Client()
        coleccion = db.collection('proyecciones_emprestito')
        
        # Buscar si ya existe
        existing_docs = coleccion.where('referencia_proceso', '==', datos_dict['referencia_proceso']).limit(1).stream()
        
        if any(existing_docs):
            raise HTTPException(
                status_code=409,
                detail=f"Ya existe una proyección con referencia_proceso: {datos_dict['referencia_proceso']}"
            )
        
        # Agregar timestamp de creación
        from datetime import datetime
        datos_dict['created_at'] = datetime.utcnow().isoformat()
        datos_dict['updated_at'] = datetime.utcnow().isoformat()
        
        # Crear el documento
        doc_ref = coleccion.document()
        doc_ref.set(datos_dict)
        
        # Preparar respuesta exitosa
        response = {
            "success": True,
            "message": "Proyección registrada exitosamente",
            "referencia_proceso": datos_dict['referencia_proceso'],
            "doc_id": doc_ref.id,
            "datos_registrados": datos_dict,
            "timestamp": datos_dict['created_at'],
            "coleccion": "proyecciones_emprestito",
            "endpoint_info": {
                "metodo": "POST",
                "operacion": "registro_nuevo",
                "campos_registrados": list(datos_dict.keys()),
                "validaciones_aplicadas": True
            }
        }
        
        return create_utf8_response(response)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error registrando proyección: {str(e)}"
        )


# ============================================================================
# SERVIDOR
# ============================================================================

# ============================================================================
# INCLUIR ROUTERS DE ADMINISTRACIÓN
# ============================================================================

# Incluir router de administración de usuarios, roles y permisos
if AUTH_SYSTEM_AVAILABLE:
    try:
        from api.routers.auth_admin import router as auth_admin_router
        app.include_router(auth_admin_router)
        print("✅ Auth admin router included successfully")
    except Exception as e:
        print(f"⚠️ Warning: Could not include auth admin router: {e}")
else:
    print("⚠️ Auth admin router not included - Auth system not available")

# Incluir router de control de calidad de unidades de proyecto
try:
    from api.routers.quality_control import router as quality_control_router
    app.include_router(quality_control_router)
    print("✅ Quality control router included successfully")
except Exception as e:
    print(f"⚠️ Warning: Could not include quality control router: {e}")

# Incluir router de Artefacto de Captura #360
try:
    from api.routers.captura_360_router import router as captura_360_router
    app.include_router(captura_360_router)
    print("✅ Captura 360 router included successfully")
except Exception as e:
    print(f"⚠️ Warning: Could not include captura 360 router: {e}")

# ============================================================================

# Ejecutar servidor si se llama directamente
if __name__ == "__main__":
    port = int(os.getenv("PORT", 8000))
    print(f"Starting server on port: {port}")
    print(f"Environment: {os.getenv('ENVIRONMENT', 'development')}")
    print(f"Firebase Project: {PROJECT_ID}")
    
    uvicorn.run(
        "main:app", 
        host="0.0.0.0", 
        port=port, 
        reload=False
    )
