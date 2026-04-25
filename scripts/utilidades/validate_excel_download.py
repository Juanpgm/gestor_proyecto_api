"""
Script para validar el contenido del archivo Excel descargado
"""

import openpyxl
from pathlib import Path

def validate_excel_content():
    """Validar el contenido del archivo Excel descargado"""
    
    print("=" * 80)
    print("🔍 VALIDACIÓN DEL ARCHIVO EXCEL")
    print("=" * 80)
    
    # Buscar el archivo más reciente
    files = list(Path(".").glob("test_unidades_proyecto_*.xlsx"))
    
    if not files:
        print("❌ No se encontraron archivos de prueba")
        return
    
    # Tomar el archivo completo
    excel_file = "test_unidades_proyecto_completo.xlsx"
    
    if not Path(excel_file).exists():
        excel_file = str(files[0])
    
    print(f"\n📄 Archivo: {excel_file}")
    print(f"   Tamaño: {Path(excel_file).stat().st_size / 1024:.2f} KB")
    
    try:
        # Abrir el archivo
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        print(f"\n📊 INFORMACIÓN DE LA HOJA:")
        print(f"   Nombre: {ws.title}")
        print(f"   Dimensiones: {ws.dimensions}")
        print(f"   Filas totales: {ws.max_row}")
        print(f"   Columnas totales: {ws.max_column}")
        
        # Leer encabezados
        print(f"\n📋 ENCABEZADOS (primeras 15 columnas):")
        headers = []
        for col in range(1, min(16, ws.max_column + 1)):
            header = ws.cell(row=1, column=col).value
            headers.append(header)
            print(f"   {col}. {header}")
        
        # Mostrar primera fila de datos
        print(f"\n📄 PRIMERA FILA DE DATOS:")
        if ws.max_row > 1:
            for col in range(1, min(11, ws.max_column + 1)):
                header = ws.cell(row=1, column=col).value
                value = ws.cell(row=2, column=col).value
                print(f"   {header}: {value}")
        
        # Estadísticas
        print(f"\n📈 ESTADÍSTICAS:")
        print(f"   Total registros (sin encabezado): {ws.max_row - 1}")
        print(f"   Total columnas: {ws.max_column}")
        print(f"   Congelación de paneles: {'Sí' if ws.freeze_panes else 'No'}")
        
        # Validar que hay datos
        if ws.max_row > 1:
            print(f"\n✅ El archivo contiene datos válidos")
            
            # Contar campos con datos
            print(f"\n📊 COMPLETITUD DE DATOS (primera fila):")
            for col in range(1, min(11, ws.max_column + 1)):
                header = ws.cell(row=1, column=col).value
                value = ws.cell(row=2, column=col).value
                status = "✅" if value else "⚠️"
                print(f"   {status} {header}: {value if value else 'NULL'}")
        else:
            print(f"\n⚠️ El archivo no contiene datos (solo encabezados)")
        
        # Verificar columnas críticas
        print(f"\n🔍 COLUMNAS CRÍTICAS:")
        critical_columns = ["UPID", "Nombre UP", "Centro Gestor", "Estado"]
        for critical in critical_columns:
            if critical in headers:
                col_idx = headers.index(critical) + 1
                value = ws.cell(row=2, column=col_idx).value if ws.max_row > 1 else None
                print(f"   ✅ {critical}: {value}")
            else:
                print(f"   ❌ {critical}: NO ENCONTRADA")
        
        print("\n" + "=" * 80)
        print("✅ Validación completada")
        print("=" * 80)
        
    except Exception as e:
        print(f"\n❌ Error validando archivo: {str(e)}")
        import traceback
        print(traceback.format_exc())

if __name__ == "__main__":
    validate_excel_content()
